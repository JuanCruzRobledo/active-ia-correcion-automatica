"""
Tests de `excel_cierre_cursada.generar_excel_cierre` — layout del modelo CORREGIDO
(bugfix cierre-cursada-fix, tarea 12).

Verifica contra el modelo `docs/modelos/modelo planilla de cierre CORREGIDO.xlsx`
(ver diseño, sección "Layout exacto del modelo CORREGIDO"):
  - título "TOTAL MATERIA {NOMBRE}" (merge A1:B1, mayúsculas)
  - resumen de 3 conteos PROMOCIONADOS/REGULARES/RECURSANTES en A2:B4
  - columnas dinámicas Nombre y Apellido | Email | Parcial n... | Global TPI |
    Estado Alumno | Nota Final (Nota Final es una adición deliberada, última columna)
  - barra de bloque por comisión mergeando desde la columna C hasta la última
  - SIN columnas de TPs y SIN gráfico de dona (0 charts)
  - Nota Final = entero o "N/E"; Global TPI = "N/E" cuando `global_valor is None`
"""

import io
from datetime import datetime

from hypothesis import given, strategies as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.models.cierre_cursada import CierreCursadaAlumno, CierreCursadaRun
from app.models.enums import EstadoCierreEnum
from app.services.excel_cierre_cursada import (
    _agrupar_por_comision,
    _fmt_nota_final,
    _fmt_valor,
    generar_excel_cierre,
)

# Config congelada de 2 parciales + 1 global (examenes_snapshot), igual shape que el
# service arma en `generar` (ver cierre_cursada_service.py).
_EXAMENES_SNAPSHOT = [
    {"id": 1, "tipo": "PARCIAL", "moodle_cmid": 101, "modo_aprobacion": "NUMERICO",
     "nota_minima": 60, "recupera_examen_id": None, "orden": 1},
    {"id": 2, "tipo": "PARCIAL", "moodle_cmid": 102, "modo_aprobacion": "NUMERICO",
     "nota_minima": 60, "recupera_examen_id": None, "orden": 2},
    {"id": 3, "tipo": "GLOBAL", "moodle_cmid": 103, "modo_aprobacion": "NUMERICO",
     "nota_minima": 6, "recupera_examen_id": None, "orden": 3},
]


def _alumno(
    apellido, nombre, email, estado, *,
    p1=80.0, p2=60.0, global_valor=9.0, nota_final=8,
    comision_nombre="Comisión 1", tutor_nombre="Juan Pérez",
):
    resultados_examenes = [
        {"examen_id": 1, "valor_real": p1},
        {"examen_id": 2, "valor_real": p2},
        {"examen_id": 3, "valor_real": global_valor},
    ]
    return CierreCursadaAlumno(
        moodle_user_id=1, nombre=nombre, apellido=apellido, email=email,
        comision_id=None, comision_nombre=comision_nombre, tutor_nombre=tutor_nombre,
        resultados_examenes=resultados_examenes,
        global_valor=global_valor,
        estado=EstadoCierreEnum(estado),
        nota_final=nota_final,
    )


def _run(alumnos, *, promociona=1, regulariza=0, recursa=0):
    run = CierreCursadaRun(
        id=1, materia_id=1, cuatrimestre_id=1,
        examenes_snapshot=_EXAMENES_SNAPSHOT,
        generado_por_id=1, total_alumnos=len(alumnos),
        total_promociona=promociona, total_regulariza=regulariza, total_recursa=recursa,
    )
    run.created_at = datetime(2026, 1, 1, 12, 0)
    run.alumnos = alumnos
    return run


def _wb(data: bytes):
    return load_workbook(io.BytesIO(data))


def test_titulo_total_materia_merge_a1_b1_mayusculas():
    run = _run([_alumno("Gómez", "Ana", "ana@x.com", "PROMOCIONA")])
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    assert ws["A1"].value == "TOTAL MATERIA PROGRAMACIÓN 1"
    merges = {str(m) for m in ws.merged_cells.ranges}
    assert "A1:B1" in merges


def test_headers_exactos_y_en_orden_2_parciales_1_global():
    run = _run([_alumno("Gómez", "Ana", "ana@x.com", "PROMOCIONA")])
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    fila_headers = next(
        [c.value for c in row]
        for row in ws.iter_rows()
        if row[0].value == "Nombre y Apellido"
    )
    # Bug 4 (tarea 10.3): se agrega la columna auxiliar "Recuperable" al final del
    # layout "por Comisiones" (col H con 2 parciales + 1 global).
    assert fila_headers == [
        "Nombre y Apellido", "Email", "Parcial 1", "Parcial 2",
        "Global TPI", "Estado Alumno", "Nota Final", "Recuperable",
    ]


def test_resumen_de_conteos_como_formulas_nativas_a2_b6():
    """Bug 4 (tarea 10.4): el resumen deja de ser 3 conteos estáticos y pasa a 5
    conteos como FÓRMULAS nativas (`COUNTIF`, sintaxis de archivo en inglés + coma) en
    A2:B6 — PROMOCIONADOS/REGULARES/RECURSANTES sobre la columna Estado (F) y las nuevas
    filas RECUPERABLES (sobre H) y ABANDONOS (sobre F). Preserva la intención original (el
    resumen muestra los conteos) asertando la naturaleza de fórmula recalculable."""
    run = _run(
        [
            _alumno("Gómez", "Ana", "ana@x.com", "PROMOCIONA"),
            _alumno("Pérez", "Bruno", "bruno@x.com", "REGULARIZA"),
        ],
        promociona=1, regulariza=1, recursa=0,
    )
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    assert (ws["A2"].value, ws["B2"].value) == ("PROMOCIONADOS", '=COUNTIF(F:F,"PROMOCIONA")')
    assert (ws["A3"].value, ws["B3"].value) == ("REGULARES", '=COUNTIF(F:F,"REGULARIZA")')
    assert (ws["A4"].value, ws["B4"].value) == ("RECURSANTES", '=COUNTIF(F:F,"RECURSA")')
    assert (ws["A5"].value, ws["B5"].value) == ("RECUPERABLES", '=COUNTIF(H:H,"RECUPERABLE*")')
    assert (ws["A6"].value, ws["B6"].value) == ("ABANDONOS", '=COUNTIF(F:F,"ABANDONO")')


def test_barra_de_bloque_mergea_c_hasta_ultima_columna():
    """Bug 4 (tarea 10.3): al agregar la columna "Recuperable" al final, la barra de
    bloque merge desde C hasta la nueva última columna (H con 2 parciales + 1 global),
    no hasta G. Se preserva que la barra arranca en la columna C."""
    run = _run([_alumno("Gómez", "Ana", "ana@x.com", "PROMOCIONA")])
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    # Última columna = "Recuperable" (encabezado de la tabla de detalle).
    ultima_col = next(
        c.column
        for row in ws.iter_rows()
        for c in row
        if c.value == "Recuperable"
    )
    ultima_letra = get_column_letter(ultima_col)

    barra_cell = next(
        c for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.startswith("Comisión 1")
    )
    fila = barra_cell.row
    merges = {str(m) for m in ws.merged_cells.ranges}
    assert f"C{fila}:{ultima_letra}{fila}" in merges
    assert ultima_letra == "H"
    # La barra arranca en la columna C (3), no en A.
    assert barra_cell.column == 3


def test_sin_columnas_de_tp_y_sin_dona():
    run = _run([_alumno("Gómez", "Ana", "ana@x.com", "PROMOCIONA")])
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    valores = {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}
    columnas_viejas = {"TPs Aprobados", "Autoeval OK", "TPI", "Habilitado para Final"}
    assert not (valores & columnas_viejas)
    assert len(ws._charts) == 0


def test_celda_nota_final_entero_o_ne():
    alumno_con_nf = _alumno("Gómez", "Ana", "ana@x.com", "PROMOCIONA", nota_final=8)
    alumno_sin_nf = _alumno("Pérez", "Bruno", "bruno@x.com", "REGULARIZA", nota_final=None)
    run = _run([alumno_con_nf, alumno_sin_nf], promociona=1, regulariza=1)
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    # Bug 4 (tarea 10.3): "Nota Final" ya no es la última columna (ahora le sigue
    # "Recuperable"). Se referencia la columna "Nota Final" por su posición en el
    # encabezado, en vez de `[-1]`.
    fila_headers = next(
        [c.value for c in row]
        for row in ws.iter_rows()
        if row[0].value == "Nombre y Apellido"
    )
    idx_nota_final = fila_headers.index("Nota Final")

    fila_ana = next(row for row in ws.iter_rows() if row[0].value == "Gómez, Ana")
    fila_bruno = next(row for row in ws.iter_rows() if row[0].value == "Pérez, Bruno")
    assert fila_ana[idx_nota_final].value == 8
    # REGULARIZA con nota_final=None renderiza la celda "Nota Final" en blanco, no
    # "N/E" (ese formateo queda reservado a las columnas de examen vía `_fmt_valor`).
    assert fila_bruno[idx_nota_final].value == "" or fila_bruno[idx_nota_final].value is None


def test_global_tpi_ne_cuando_global_valor_es_none():
    alumno = _alumno(
        "Gómez", "Ana", "ana@x.com", "REGULARIZA",
        global_valor=None, nota_final=None,
    )
    run = _run([alumno], promociona=0, regulariza=1)
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    fila_ana = next(row for row in ws.iter_rows() if row[0].value == "Gómez, Ana")
    # Columnas: Nombre, Email, Parcial 1, Parcial 2, Global TPI, Estado, Nota Final
    assert fila_ana[4].value == "N/E"


# ============================================================================
# _agrupar_por_comision — orden de los bloques de comisión (Bug 3, tarea 5)
# ============================================================================
#
# Property 5: Bug Condition - Bloques de comisión ordenados alfabéticamente (design.md).
# CRITICAL: estos tests DEBEN FALLAR sobre el código sin arreglar — hoy
# `_agrupar_por_comision` arma el dict en el orden de LLEGADA de los alumnos (orden de
# inserción), no alfabético, y el bloque "Sin comisión asignada" queda en la posición del
# primer alumno sin comisión en vez de al final. La falla confirma que el Bug 3 existe.
# NO se debe modificar `_agrupar_por_comision` (ni ningún otro código fuente) para hacer
# pasar estos tests.
# **Validates: Requirements 1.6, 1.7**


def test_bug_bloques_desordenados_no_quedan_alfabeticos():
    """Alumnos que llegan en el orden de comisiones "M26 C1-03", "M26 C1-01",
    "M26 C1-02" -> esperado (tras el fix): las claves de `_agrupar_por_comision`
    empiezan por el bloque de "M26 C1-01", luego "M26 C1-02", luego "M26 C1-03".
    HOY: `_agrupar_por_comision` arma el dict en el orden de llegada de los alumnos
    ("M26 C1-03", "M26 C1-01", "M26 C1-02"), no alfabético."""
    alumnos = [
        _alumno(
            "Gómez", "Ana", "ana@x.com", "PROMOCIONA",
            comision_nombre="M26 C1-03", tutor_nombre=None,
        ),
        _alumno(
            "Pérez", "Bruno", "bruno@x.com", "PROMOCIONA",
            comision_nombre="M26 C1-01", tutor_nombre=None,
        ),
        _alumno(
            "Díaz", "Carla", "carla@x.com", "PROMOCIONA",
            comision_nombre="M26 C1-02", tutor_nombre=None,
        ),
    ]

    titulos = list(_agrupar_por_comision(alumnos).keys())

    assert titulos == ["M26 C1-01", "M26 C1-02", "M26 C1-03"]


def test_bug_sin_comision_asignada_no_queda_al_final():
    """Un alumno sin comisión (`comision_nombre=None`) que llega PRIMERO, seguido de
    dos comisiones reales -> esperado (tras el fix): "Sin comisión asignada" queda
    como ÚLTIMA clave. HOY: `_agrupar_por_comision` lo deja primero, en el orden de
    llegada."""
    alumnos = [
        _alumno(
            "Gómez", "Ana", "ana@x.com", "PROMOCIONA",
            comision_nombre=None, tutor_nombre=None,
        ),
        _alumno(
            "Pérez", "Bruno", "bruno@x.com", "PROMOCIONA",
            comision_nombre="M26 C1-01", tutor_nombre=None,
        ),
        _alumno(
            "Díaz", "Carla", "carla@x.com", "PROMOCIONA",
            comision_nombre="M26 C1-02", tutor_nombre=None,
        ),
    ]

    titulos = list(_agrupar_por_comision(alumnos).keys())

    assert titulos[-1] == "Sin comisión asignada"


# ===================== Property 5 (property-based): orden siempre alfabético =====================
#
# Genera listas de alumnos con `comision_nombre` aleatorios (incluyendo `None`) en orden
# de llegada aleatorio. Pool con distinta capitalización ("m26 c1-01" vs "M26 C1-03") para
# ejercitar el orden case-insensitive, sin que dos nombres del pool colisionen al
# `casefold()` (evita ambigüedad sobre qué bloque "debería" aparecer primero).
# **Validates: Requirements 1.6, 1.7**

_POOL_COMISIONES = ["M26 C1-03", "m26 c1-01", "A27 C1-02", None]


@given(st.lists(st.sampled_from(_POOL_COMISIONES), min_size=1, max_size=8))
def test_prop5_bloques_siempre_alfabeticos_con_sin_comision_al_final(nombres):
    alumnos = [
        _alumno(
            f"Apellido{i}", f"Nombre{i}", f"a{i}@x.com", "PROMOCIONA",
            comision_nombre=nombre, tutor_nombre=None,
        )
        for i, nombre in enumerate(nombres)
    ]

    titulos = list(_agrupar_por_comision(alumnos).keys())

    nombres_reales = sorted(
        {a.comision_nombre for a in alumnos if a.comision_nombre is not None},
        key=str.casefold,
    )
    esperados = list(nombres_reales)
    if any(a.comision_nombre is None for a in alumnos):
        esperados.append("Sin comisión asignada")

    assert titulos == esperados
    if "Sin comisión asignada" in titulos:
        assert titulos[-1] == "Sin comisión asignada"


# ============================================================================
# _fmt_nota_final — celda "Nota Final" en blanco para no promocionados (Bug 4, tarea 6)
# ============================================================================
#
# Property 6: Bug Condition - Nota Final sólo para PROMOCIONA (design.md,
# isBugCondition_nota_no_promociona). CRITICAL: este test DEBE FALLAR sobre el código
# sin arreglar -- hoy `_fmt_nota_final(None)` devuelve "N/E" (mismo formateo que
# `_fmt_valor`), en vez de una celda en blanco (`""`). La falla confirma que el Bug 4
# existe. NO se debe modificar `_fmt_nota_final` (ni ningún otro código fuente) para
# hacer pasar este test.
# **Validates: Requirements 1.8, 1.9**


def test_bug_fmt_nota_final_none_deberia_ser_blanco_no_ne():
    """Bug 4: `_fmt_nota_final(None)` -> esperado (tras el fix) `""` (celda en blanco).
    HOY: devuelve "N/E", igual que `_fmt_valor(None)` (ambos formateadores deberían
    DIVERGER tras el fix: sólo `_fmt_nota_final` pasa a blanco-en-`None`)."""
    assert _fmt_nota_final(None) == ""


# ============================================================================
# Preservación (tarea 7, ANTES del fix de Bug 3 + Bug 4)
# ============================================================================
#
# Property 5: Preservation - Orden intra-bloque y encabezado (Bug 3, design.md).
# Property 6: Preservation - PROMOCIONA numérico, columna N/E de examen y estado (Bug 4).
# Metodología observation-first: se corre el código SIN arreglar con inputs que NO
# disparan cada bug y se captura el comportamiento observado, que debe seguir intacto
# después del fix de la tarea 8 (Bug 3 sólo reordena los BLOQUES por comisión, no los
# alumnos dentro de cada bloque ni el formato del encabezado; Bug 4 sólo cambia el
# render de `_fmt_nota_final`, no `_fmt_valor` ni la clasificación de estado).
#
# - PROMOCIONA conserva su nota (3.9) y estado sin cambios (3.11): ya cubiertos, sin
#   duplicar, por `test_promociona_todos_incluido_global` y los tests de
#   `calcular_estado_cierre` en `test_cierre_cursada_calculo.py`.


def test_preservacion_orden_intrabloque_por_apellido_nombre():
    """Orden intra-bloque (3.7): dentro de un bloque, los alumnos siguen ordenados por
    `(apellido, nombre)` -- comportamiento ya existente en `_escribir_detalle`
    (`sorted(del_grupo, key=lambda x: (x.apellido or "", x.nombre or ""))`). Observado
    sobre el código SIN arreglar: este orden es independiente del orden de llegada de
    los alumnos y no se ve afectado por el Bug 3 (que sólo reordena los BLOQUES).
    **Validates: Requirements 3.7**"""
    alumnos = [
        _alumno("Zeta", "Zoe", "zoe@x.com", "PROMOCIONA", comision_nombre="Comisión 1", tutor_nombre=None),
        _alumno("Alfa", "Ana", "ana@x.com", "PROMOCIONA", comision_nombre="Comisión 1", tutor_nombre=None),
        _alumno("Alfa", "Beto", "beto@x.com", "PROMOCIONA", comision_nombre="Comisión 1", tutor_nombre=None),
    ]
    run = _run(alumnos, promociona=3)
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    nombres_en_orden = [
        row[0].value
        for row in ws.iter_rows()
        if isinstance(row[0].value, str) and "," in row[0].value
    ]
    assert nombres_en_orden == ["Alfa, Ana", "Alfa, Beto", "Zeta, Zoe"]


def test_preservacion_encabezado_de_bloque_incluye_tutor():
    """Encabezado de bloque (3.8): la barra de cada bloque sigue con el formato
    "{comisión} — Tutor: {tutor}" (extiende `test_barra_de_bloque_mergea_c_hasta_g`,
    que sólo verifica el merge/posición, para observar el texto compuesto completo).
    **Validates: Requirements 3.8**"""
    run = _run(
        [
            _alumno(
                "Gómez", "Ana", "ana@x.com", "PROMOCIONA",
                comision_nombre="Comisión 1", tutor_nombre="Juan Pérez",
            )
        ]
    )
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    barra_cell = next(
        c for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.startswith("Comisión 1")
    )
    assert barra_cell.value == "Comisión 1 — Tutor: Juan Pérez"


def test_preservacion_una_sola_comision_orden_no_se_altera():
    """Edge (3.7/3.8-scope): con una sola comisión ya no hay otros bloques para
    reordenar -- `_agrupar_por_comision` sigue devolviendo un único bloque con el
    título compuesto intacto. **Validates: Requirements 3.7, 3.8**"""
    alumnos = [
        _alumno(
            "Gómez", "Ana", "ana@x.com", "PROMOCIONA",
            comision_nombre="M26 C1-01", tutor_nombre="Juan Pérez",
        )
    ]
    titulos = list(_agrupar_por_comision(alumnos).keys())
    assert titulos == ["M26 C1-01 — Tutor: Juan Pérez"]


def test_preservacion_solo_sin_comision_asignada_bloque_unico():
    """Edge (3.7/3.8-scope): si TODOS los alumnos están sin comisión, el único bloque
    es "Sin comisión asignada" -- no hay otros bloques con los que reordenar.
    **Validates: Requirements 3.7, 3.8**"""
    alumnos = [
        _alumno("Gómez", "Ana", "ana@x.com", "PROMOCIONA", comision_nombre=None, tutor_nombre=None),
        _alumno("Pérez", "Bruno", "bruno@x.com", "PROMOCIONA", comision_nombre=None, tutor_nombre=None),
    ]
    titulos = list(_agrupar_por_comision(alumnos).keys())
    assert titulos == ["Sin comisión asignada"]


def test_preservacion_fmt_valor_none_sigue_siendo_ne():
    """Columna de examen "N/E" bajo Bug 4 (3.10): `_fmt_valor(None)` sigue devolviendo
    "N/E" sin cambios -- diverge a propósito de `_fmt_nota_final`, que el fix de Bug 4
    (tarea 8) cambiará a celda en blanco (`""`) ante `None`.
    **Validates: Requirements 3.10**"""
    assert _fmt_valor(None) == "N/E"


# ============================================================================
# Bug 2 — El Excel de cierre debe tener DOS hojas (tarea 2, exploración)
# ============================================================================
#
# Property 2: Bug Condition - Excel con dos hojas (design.md, isBugCondition de Bug 2).
# CRITICAL: estos tests DEBEN FALLAR sobre el código sin arreglar — hoy
# `generar_excel_cierre` produce UNA sola hoja (`wb.active`) con los bloques por
# comisión, sin la segunda hoja plana "Crudo" (lista alfabética con columnas Comision y
# Tutor) del modelo `docs/modelos/Cierre_Programacin_3 FIX.xlsx`. La falla confirma que
# el Bug 2 existe. NO se debe modificar `excel_cierre_cursada` (ni ningún otro código
# fuente) para hacer pasar estos tests.
# Detalle de la Bug Condition:
#   cantidadDeHojas(wb) != 2
#   OR NOT existeHoja(wb, "{materia} Crudo")
#   OR NOT hojaCrudoTieneColumnas(wb, ["Comision", "Tutor"])
# **Validates: Requirements 1.4, 1.5**


def _run_prog3():
    """Corrida mockeada de "Programación 3" con alumnos de dos comisiones y su tutor,
    para ejercitar ambas hojas (por Comisiones + Crudo)."""
    alumnos = [
        _alumno(
            "Gómez", "Ana", "ana@x.com", "PROMOCIONA",
            comision_nombre="M25 C3-01", tutor_nombre="Juan Pérez",
        ),
        _alumno(
            "Díaz", "Carla", "carla@x.com", "REGULARIZA",
            comision_nombre="M25 C3-02", tutor_nombre="María López",
        ),
    ]
    return _run(alumnos, promociona=1, regulariza=1, recursa=0)


def test_bug_excel_dos_hojas_por_comisiones_y_crudo():
    """Bug 2: el Excel de cierre debe tener DOS hojas —
    "Programación 3 por Comisiones" y "Programación 3 Crudo".
    HOY: `generar_excel_cierre` devuelve UNA sola hoja (`len(wb.worksheets) == 1`).
    **Validates: Requirements 1.4**"""
    data, _ = generar_excel_cierre("Programación 3", _run_prog3())
    wb = _wb(data)

    assert len(wb.worksheets) == 2
    assert wb.sheetnames == [
        "Programación 3 por Comisiones",
        "Programación 3 Crudo",
    ]


def test_bug_hoja_crudo_tiene_columnas_comision_y_tutor():
    """Bug 2: la hoja "Crudo" debe existir y tener encabezados que incluyan "Comision"
    y "Tutor". HOY: la hoja "Crudo" no existe (sólo hay una hoja con los bloques por
    comisión, sin esas columnas).
    **Validates: Requirements 1.5**"""
    data, _ = generar_excel_cierre("Programación 3", _run_prog3())
    wb = _wb(data)

    assert "Programación 3 Crudo" in wb.sheetnames

    ws_crudo = wb["Programación 3 Crudo"]
    fila_headers = next(
        [c.value for c in row]
        for row in ws_crudo.iter_rows()
        if row[0].value == "Nombre y Apellido"
    )
    assert "Comision" in fila_headers
    assert "Tutor" in fila_headers


# ============================================================================
# Bug 4 — Conteos del resumen y columna "Recuperable" con FÓRMULAS NATIVAS
# (tarea 4, exploración)
# ============================================================================
#
# Property 4: Bug Condition - Conteos y Recuperable con fórmulas nativas (design.md,
# isBugCondition de Bug 4).
# CRITICAL: estos tests DEBEN FALLAR sobre el código sin arreglar — hoy
# `_escribir_resumen` escribe los conteos del resumen (PROMOCIONADOS/REGULARES/
# RECURSANTES) como enteros fijos (`run.total_*`), NO existe una columna/fórmula
# "Recuperable" por alumno, ni conteos de Recuperables/Abandonos. La falla confirma que
# el Bug 4 existe. NO se debe modificar `excel_cierre_cursada` (ni ningún otro código
# fuente) para hacer pasar estos tests.
#
# Detalle de la Bug Condition:
#   esValorEstatico(celda_resumen)
#   OR NOT existeColumnaRecuperable
#   OR NOT esFormulaNativa(celda_recuperable)
#
# Fórmulas nativas de la hoja "por Comisiones" (sintaxis de archivo: inglés + coma;
# Excel es-AR las muestra traducidas): estado en F, parciales en C/D, aux Recuperable en H.
#   Recuperable (fila n):
#     =IF(F{n}<>"RECURSA","",IF(AND(IFERROR(VALUE(C{n}),0)>=40,IFERROR(VALUE(D{n}),0)<40),
#      "RECUPERABLE CON PARCIAL 2",IF(AND(IFERROR(VALUE(D{n}),0)>=40,IFERROR(VALUE(C{n}),0)<40),
#      "RECUPERABLE CON PARCIAL 1","")))
#   Promocionados: =COUNTIF(F:F,"PROMOCIONA")
#   Regulares:     =COUNTIF(F:F,"REGULARIZA")
#   Recursantes:   =COUNTIF(F:F,"RECURSA")
#   Recuperables:  =COUNTIF(H:H,"RECUPERABLE*")
#   Abandonos:     =COUNTIF(F:F,"ABANDONO")
# **Validates: Requirements 1.8, 1.9, 1.10**


def _run_bug4():
    """Corrida mockeada con un RECURSA (P1>=40, P2<40 -> recuperable con parcial 1) para
    ejercitar tanto el resumen como la columna Recuperable de la hoja "por Comisiones"."""
    alumnos = [
        _alumno("Gómez", "Ana", "ana@x.com", "PROMOCIONA"),
        _alumno(
            "Pérez", "Bruno", "bruno@x.com", "RECURSA",
            p1=80.0, p2=30.0, global_valor=None, nota_final=None,
        ),
    ]
    return _run(alumnos, promociona=1, regulariza=0, recursa=1)


def _valor_junto_a_etiqueta(ws, etiqueta: str):
    """Valor de la celda inmediatamente a la derecha de `etiqueta` (layout resumen:
    etiqueta en la col N, conteo en la col N+1)."""
    for row in ws.iter_rows():
        for c in row:
            if c.value == etiqueta:
                return ws.cell(c.row, c.column + 1).value
    return None


def _formulas(ws) -> list[str]:
    """Todas las celdas cuyo valor es un string que empieza con '=' (fórmulas nativas)."""
    return [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]


def test_bug_conteo_promocionados_es_formula_no_entero():
    """Bug 4: el conteo "PROMOCIONADOS" del resumen debe ser una fórmula nativa
    `=COUNTIF(F:F,"PROMOCIONA")`. HOY: `_escribir_resumen` escribe `run.total_promociona`
    como un `int` estático (no recalcula si se edita el "Estado Alumno").
    **Validates: Requirements 1.8, 1.9**"""
    data, _ = generar_excel_cierre("Programación 3", _run_bug4())
    ws = _wb(data).active

    valor = _valor_junto_a_etiqueta(ws, "PROMOCIONADOS")
    assert isinstance(valor, str), f"esperaba una fórmula, obtuve {valor!r} (valor estático)"
    assert valor.startswith('=COUNTIF(F:F,"PROMOCIONA")')


def test_bug_columna_recuperable_es_formula_por_fila():
    """Bug 4: debe existir una columna "Recuperable" cuya celda por fila sea la fórmula
    nativa `=IF(F{n}<>"RECURSA","",IF(AND(...` (sintaxis de archivo: inglés + coma) con
    las referencias de la hoja "por Comisiones" (estado F, parciales C/D). HOY: no existe
    la columna ni la fórmula.
    **Validates: Requirements 1.10**"""
    data, _ = generar_excel_cierre("Programación 3", _run_bug4())
    ws = _wb(data).active

    # Localizar el encabezado "Recuperable" y su columna.
    header_cell = next(
        (
            c
            for row in ws.iter_rows()
            for c in row
            if c.value == "Recuperable"
        ),
        None,
    )
    assert header_cell is not None, 'no existe la columna "Recuperable"'
    col_recuperable = header_cell.column

    # Primera fila de datos bajo el encabezado (fila con un alumno "Apellido, Nombre").
    fila_datos_alumno = next(
        row
        for row in ws.iter_rows(min_row=header_cell.row + 1)
        if isinstance(row[0].value, str) and "," in row[0].value
    )
    n = fila_datos_alumno[0].row
    celda = ws.cell(n, col_recuperable).value

    assert isinstance(celda, str) and celda.startswith("="), (
        f"la celda Recuperable no es una fórmula: {celda!r}"
    )
    assert celda.startswith(f'=IF(F{n}<>"RECURSA","",IF(AND(')
    assert f"IFERROR(VALUE(C{n}),0)>=40" in celda
    assert f"IFERROR(VALUE(D{n}),0)<40" in celda
    assert "RECUPERABLE CON PARCIAL 1" in celda
    assert "RECUPERABLE CON PARCIAL 2" in celda


def test_bug_conteo_recuperables_y_abandonos_son_formulas():
    """Bug 4: el resumen debe incluir el conteo de recuperables
    `=COUNTIF(H:H,"RECUPERABLE*")` y el de abandonos `=COUNTIF(F:F,"ABANDONO")`.
    HOY: ninguno de los dos existe (el resumen sólo tiene 3 conteos estáticos).
    **Validates: Requirements 1.10**"""
    data, _ = generar_excel_cierre("Programación 3", _run_bug4())
    ws = _wb(data).active

    formulas = _formulas(ws)
    assert any(f.startswith('=COUNTIF(H:H,"RECUPERABLE*")') for f in formulas), (
        f"falta el conteo de recuperables; fórmulas halladas: {formulas!r}"
    )
    assert any(f.startswith('=COUNTIF(F:F,"ABANDONO")') for f in formulas), (
        f"falta el conteo de abandonos; fórmulas halladas: {formulas!r}"
    )


# ============================================================================
# Property 7 (Preservación) — spec cierre-cursada-quiz-recuperables-fix
# ============================================================================
#
# Metodología observation-first: se corre el generador SIN arreglar (hoy UNA sola hoja
# "por Comisiones") con inputs que NO disparan los Bugs 2/4 y se captura el comportamiento
# a preservar. El fix de Bug 2 agrega una segunda hoja "Crudo" y el de Bug 4 cambia los
# conteos a fórmulas, pero NO deben alterar: el orden numérico natural de las comisiones
# con "Sin comisión asignada" al final (3.6), ni el `N/E` en las columnas de examen sin
# nota (3.8).
#
# Nota (3.10 — estilos de la casa): la preservación de los estilos de `excel_estilos.py`
# (banda de título, headers, filas, barra de bloque) sobre la hoja actual ya está cubierta
# por `test_excel_cierre_usa_paleta_y_bordes_de_excel_estilos` en
# `test_cierre_cursada_service.py`; no se duplica acá.


def test_prop7_preservacion_orden_numerico_natural_de_comisiones():
    """3.6: los bloques de comisión se ordenan por orden numérico NATURAL del nombre
    ("M26 C1-2" antes de "M26 C1-10", no orden lexicográfico), con "Sin comisión asignada"
    siempre al final — garantía de `comision-orden-numerico-fix` que el fix de esta spec
    debe preservar. **Validates: Requirements 3.6**"""
    alumnos = [
        _alumno("A", "a", "a@x.com", "PROMOCIONA", comision_nombre="M26 C1-10", tutor_nombre=None),
        _alumno("B", "b", "b@x.com", "PROMOCIONA", comision_nombre="M26 C1-2", tutor_nombre=None),
        _alumno("C", "c", "c@x.com", "PROMOCIONA", comision_nombre="M26 C1-1", tutor_nombre=None),
        _alumno("D", "d", "d@x.com", "PROMOCIONA", comision_nombre=None, tutor_nombre=None),
    ]

    titulos = list(_agrupar_por_comision(alumnos).keys())

    assert titulos == ["M26 C1-1", "M26 C1-2", "M26 C1-10", "Sin comisión asignada"]


def test_prop7_preservacion_columnas_examen_muestran_ne_sin_nota():
    """3.8: las columnas de examen (`Parcial n`, `Global TPI`) siguen mostrando `N/E`
    cuando el alumno no tiene nota en ese examen — comportamiento de `_fmt_valor` que el
    fix no toca. **Validates: Requirements 3.8**"""
    alumno = _alumno(
        "Gómez", "Ana", "ana@x.com", "REGULARIZA",
        p1=80.0, p2=None, global_valor=None, nota_final=None,
    )
    run = _run([alumno], promociona=0, regulariza=1)
    data, _ = generar_excel_cierre("Programación 1", run)
    ws = _wb(data).active

    fila = next(row for row in ws.iter_rows() if row[0].value == "Gómez, Ana")
    # Columnas: Nombre, Email, Parcial 1, Parcial 2, Global TPI, Estado, Nota Final
    assert fila[2].value == 80.0     # Parcial 1 con nota -> valor numérico
    assert fila[3].value == "N/E"    # Parcial 2 sin nota -> N/E
    assert fila[4].value == "N/E"    # Global TPI sin nota -> N/E


@given(
    p1=st.one_of(st.none(), st.floats(min_value=0.0, max_value=100.0, allow_nan=False, allow_infinity=False)),
    p2=st.one_of(st.none(), st.floats(min_value=0.0, max_value=100.0, allow_nan=False, allow_infinity=False)),
)
def test_prop7_preservacion_fmt_valor_none_es_ne_resto_pasa_directo(p1, p2):
    """3.8 (property-based): `_fmt_valor` devuelve `N/E` exactamente cuando no hay nota
    (`None`), y el valor numérico tal cual en caso contrario. **Validates: Requirements 3.8**"""
    for valor in (p1, p2):
        resultado = _fmt_valor(valor)
        if valor is None:
            assert resultado == "N/E"
        else:
            assert resultado == valor
