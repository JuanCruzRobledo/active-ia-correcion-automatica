"""PERF-009 — los artefactos oficiales NO se truncan silenciosamente.

El Excel de notas y el ZIP de devoluciones usaban ``get_all(per_page=1000)`` con el
comentario "# Get all". Era mentira: una comisión con más de 1000 entregas/correcciones
exportaba INCOMPLETA sin avisar. Esto es un problema de CORRECTITUD (actas oficiales),
no sólo de performance.

El fix es un método de repo dedicado —``get_all_for_export``— que itera en lotes hasta
agotar el conjunto (sin el tope de 1000) reutilizando ``get_all``. Estos tests fijan la
garantía: el export trae TODAS las filas.

- REPO (entregas y correcciones): con un ``batch_size`` chico y N > batch, prueba que la
  iteración multi-página devuelve el conjunto COMPLETO (no cortado en un lote).
- SERVICIO (excel y zip): monkeypatchea ``get_all`` para reventar si alguien vuelve a
  pedir ``per_page >= 1000`` (el viejo tope truncante) y verifica que el artefacto refleja
  TODAS las entregas/correcciones sembradas.
"""

import io
import json
import sqlite3
import zipfile
from decimal import Decimal

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy import ARRAY
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

# sqlite3 no bindea dict/list (JSONB/ARRAY) directo; estos adapters los serializan a JSON.
sqlite3.register_adapter(dict, json.dumps)
sqlite3.register_adapter(list, json.dumps)

from app.models.base import Base
from app.models.comision import Comision
from app.models.correccion import Correccion
from app.models.entrega import Entrega
from app.models.enums import EstadoEntregaEnum
from app.models.materia import Materia
from app.models.rubrica import Rubrica
from app.repositories.correccion_repository import CorreccionRepository
from app.repositories.entrega_repository import EntregaRepository
from app.services.excel_service import ExcelService
from app.services.pdf_service import PDFService


# SQLite no soporta JSONB/ARRAY nativos; se compilan a JSON sólo en el dialecto de tests.
@compiles(JSONB, "sqlite")
def _jsonb_como_json_en_sqlite(element, compiler, **kw):  # noqa: D401
    return "JSON"


@compiles(ARRAY, "sqlite")
def _array_como_json_en_sqlite(element, compiler, **kw):  # noqa: D401
    return "JSON"


@pytest_asyncio.fixture
async def db_session():
    import app.models  # noqa: F401 — puebla el registry de mappers/tablas

    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False, future=True)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session_maker = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with session_maker() as session:
        yield session
    await engine.dispose()


async def _seed(db, *, n: int, corregidas: int | None = None) -> dict:
    """Siembra una materia/comisión/rúbrica y ``n`` entregas.

    Las primeras ``corregidas`` (default: todas) llevan una Correccion asociada.
    """
    if corregidas is None:
        corregidas = n

    materia = Materia(nombre="Programación 1", codigo="P1")
    db.add(materia)
    await db.flush()
    comi = Comision(materia_id=materia.id, nombre="C1-1", anio=2026, activa=True)
    db.add(comi)
    await db.flush()
    rubrica = Rubrica(
        materia_id=materia.id,
        titulo="Trabajo Práctico 1",
        descripcion="Rúbrica de prueba",
        numero=1,
        anio=2026,
        tipo="TP",
        puntaje_maximo=100,
        criterios_json=[],
    )
    db.add(rubrica)
    await db.flush()

    entregas = []
    for i in range(n):
        estado = (
            EstadoEntregaEnum.CORREGIDA if i < corregidas else EstadoEntregaEnum.SUBIDA
        )
        entregas.append(
            Entrega(
                comision_id=comi.id,
                rubrica_id=rubrica.id,
                alumno_nombre=f"Alumno {i:03d}",
                archivo_nombre=f"tp{i}.zip",
                archivo_ruta=f"/tmp/tp{i}.zip",
                archivo_tipo="zip",
                contenido_preview="print('hola')",
                estado=estado,
            )
        )
    db.add_all(entregas)
    await db.flush()

    for i in range(corregidas):
        db.add(
            Correccion(
                entrega_id=entregas[i].id,
                nota=Decimal("90"),
                criterios_json={
                    "criterios": [
                        {
                            "nombre": "Funcionalidad",
                            "puntaje_obtenido": 40,
                            "puntaje_maximo": 40,
                            "estado": "OK",
                            "feedback": "Bien",
                        }
                    ]
                },
                comentario_general="OK",
                editado_manualmente=False,
            )
        )
    await db.commit()
    return {"comision_id": comi.id, "rubrica_id": rubrica.id, "n": n}


# ---------------------------------------------------------------------------
# REPO — la iteración multi-página trae el conjunto COMPLETO (no trunca)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_entrega_get_all_for_export_no_trunca_multi_pagina(db_session):
    """Con batch_size=2 y 5 entregas, get_all_for_export devuelve las 5 (3 páginas)."""
    ids = await _seed(db_session, n=5, corregidas=0)
    repo = EntregaRepository(db_session)

    entregas = await repo.get_all_for_export(
        comision_id=ids["comision_id"], rubrica_id=ids["rubrica_id"], batch_size=2
    )

    assert len(entregas) == 5
    nombres = {e.alumno_nombre for e in entregas}
    assert nombres == {f"Alumno {i:03d}" for i in range(5)}


@pytest.mark.asyncio
async def test_correccion_get_all_for_export_no_trunca_multi_pagina(db_session):
    """Con batch_size=2 y 5 correcciones, get_all_for_export devuelve las 5 (3 páginas)."""
    ids = await _seed(db_session, n=5)
    repo = CorreccionRepository(db_session)

    correcciones = await repo.get_all_for_export(
        comision_id=ids["comision_id"], rubrica_id=ids["rubrica_id"], batch_size=2
    )

    assert len(correcciones) == 5
    # IDs únicos: no hay filas repetidas ni faltantes entre páginas.
    assert len({c.id for c in correcciones}) == 5


# ---------------------------------------------------------------------------
# SERVICIO — el export ya NO usa el tope de 1000 y sale completo
# ---------------------------------------------------------------------------


def _prohibir_tope_1000(repo):
    """Envuelve repo.get_all para que reviente si alguien pide per_page >= 1000.

    Devuelve un dict con el flag 'llamado' para inspección adicional si hiciera falta.
    """
    original = repo.get_all
    estado = {"max_per_page": 0}

    async def guard(*args, per_page=20, **kwargs):
        estado["max_per_page"] = max(estado["max_per_page"], per_page)
        assert per_page < 1000, (
            "regresión PERF-009: el export volvió a usar el tope de 1000 que trunca"
        )
        return await original(*args, per_page=per_page, **kwargs)

    repo.get_all = guard
    return estado


@pytest.mark.asyncio
async def test_exportar_notas_excel_no_trunca_y_no_usa_tope_1000(db_session):
    """El acta de notas trae TODAS las entregas y no pasa por el viejo per_page=1000."""
    ids = await _seed(db_session, n=7, corregidas=4)
    service = ExcelService(db_session)
    _prohibir_tope_1000(service.entrega_repo)

    data, _filename = await service.exportar_notas_excel(
        comision_id=ids["comision_id"], rubrica_id=ids["rubrica_id"]
    )

    ws = load_workbook(io.BytesIO(data)).active
    # Filas de datos: de la 5 en adelante, una por entrega.
    alumnos = {
        ws.cell(row=r, column=1).value for r in range(5, 5 + 7)
    }
    assert alumnos == {f"Alumno {i:03d}" for i in range(7)}
    # La fila de total confirma el conteo COMPLETO (7).
    total_row = 5 + 7 + 1  # una fila en blanco tras la última + fila de total
    assert ws.cell(row=total_row, column=1).value == "Total de entregas: 7"


@pytest.mark.asyncio
async def test_generar_zip_pdfs_no_trunca_y_no_usa_tope_1000(db_session):
    """El ZIP de devoluciones trae un PDF por corrección y no usa el viejo per_page=1000."""
    ids = await _seed(db_session, n=6)
    service = PDFService(db_session)
    _prohibir_tope_1000(service.correccion_repo)

    zip_bytes, _filename = await service.generar_zip_pdfs(
        comision_id=ids["comision_id"], rubrica_id=ids["rubrica_id"]
    )

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        nombres = zf.namelist()
        assert len(nombres) == 6
        # Todos son PDFs válidos (render efectivo, no placeholders).
        for nombre in nombres:
            assert nombre.endswith(".pdf")
            assert zf.read(nombre).startswith(b"%PDF")
