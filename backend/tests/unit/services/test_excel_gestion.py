"""Tests de ExcelService.exportar_gestion (T6).

Genera el .xlsx en memoria y lo vuelve a cargar con openpyxl para verificar:
una hoja por Regional, columnas pedidas, orden por comisión, total por hoja y
hoja "Resumen" con los totales.
"""

import io

from openpyxl import load_workbook

from unittest.mock import AsyncMock

from app.services.excel_service import ExcelService
from app.services.gestion_service import AlumnoGestion, ResultadoGestion


def _alumno(nombre, apellido, email, regional, comision, dias, tiempo, suspendido=False):
    return AlumnoGestion(nombre, apellido, email, regional, comision, dias, tiempo, suspendido)


def _resultado():
    items = [
        _alumno("Ana", "Lopez", "ana@x", "Mendoza", "M26 C1-09", 45, "45 días"),
        _alumno("Zoe", "Diaz", "zoe@x", "Mendoza", "M26 C1-01", 10, "10 días"),
        _alumno("Beto", "Ruiz", "beto@x", "Rosario", "M26 C1-02", None, "Nunca", True),
    ]
    return ResultadoGestion(items=items, total=3)


def _wb(data: bytes):
    return load_workbook(io.BytesIO(data))


def test_una_hoja_por_regional_mas_resumen():
    data, filename = ExcelService(AsyncMock()).exportar_gestion(_resultado(), "Prog I")
    wb = _wb(data)
    assert set(wb.sheetnames) == {"Resumen", "Mendoza", "Rosario"}
    assert filename.endswith(".xlsx")


def test_columnas_y_orden_por_comision():
    data, _ = ExcelService(AsyncMock()).exportar_gestion(_resultado(), "Prog I")
    ws = _wb(data)["Mendoza"]
    rows = list(ws.iter_rows(values_only=True))
    assert ("Nombre", "Apellido", "Email", "Regional", "Comisión", "Tiempo de inactividad") in rows
    # dentro de la regional, ordenado por comisión: C1-01 (Zoe) antes que C1-09 (Ana)
    emails = [r[2] for r in rows if r[2] and "@" in str(r[2])]
    assert emails == ["zoe@x", "ana@x"]


def test_total_por_hoja():
    data, _ = ExcelService(AsyncMock()).exportar_gestion(_resultado(), "Prog I")
    ws = _wb(data)["Mendoza"]
    flat = [str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None]
    assert "Total de alumnos: 2" in flat


def test_hoja_resumen_con_totales_por_regional_y_global():
    data, _ = ExcelService(AsyncMock()).exportar_gestion(_resultado(), "Prog I")
    ws = _wb(data)["Resumen"]
    pares = [(r[0], r[1]) for r in ws.iter_rows(values_only=True)]
    assert ("Mendoza", 2) in pares
    assert ("Rosario", 1) in pares
    assert ("Total", 3) in pares


def test_tiempo_inactividad_nunca_se_escribe():
    data, _ = ExcelService(AsyncMock()).exportar_gestion(_resultado(), "Prog I")
    ws = _wb(data)["Rosario"]
    flat = [str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None]
    assert "Nunca" in flat


def test_resultado_vacio_solo_resumen():
    data, _ = ExcelService(AsyncMock()).exportar_gestion(ResultadoGestion(items=[], total=0), "Prog I")
    assert _wb(data).sheetnames == ["Resumen"]
