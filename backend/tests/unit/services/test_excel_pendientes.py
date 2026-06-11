"""Tests de ExcelService.exportar_pendientes (entregas pendientes, dos ejes + tutor).

agrupar_por='trabajo' (por Práctico): hoja por trabajo; columnas
Comisión·Tutor·Nombre·Apellido·Email·Fecha; resumen por práctico.
agrupar_por='comision' (por Comisión): hoja por comisión (título con tutor);
columnas Trabajo·Nombre·Apellido·Email·Fecha; resumen Comisión·Tutor·Pendientes.
La primera hoja siempre es "Resumen".
"""

import io

from openpyxl import load_workbook
from unittest.mock import AsyncMock

from app.services.excel_service import ExcelService
from app.services.gestion_service import EntregaPendiente


def _wb(data: bytes):
    return load_workbook(io.BytesIO(data))


def _pendientes():
    # comisión C1-01 → tutor "Prof Zoe"; C1-02 → "Prof Ana"
    return [
        EntregaPendiente("TP 1 - Variables", "M26 C1-02", "Prof Ana", "Ana", "Lopez", "ana@x", "01/06/2026 10:00"),
        EntregaPendiente("TP 1 - Variables", "M26 C1-01", "Prof Zoe", "Zoe", "Diaz", "zoe@x", "02/06/2026 11:00"),
        EntregaPendiente("Parcial 1", "M26 C1-01", "Prof Zoe", "Beto", "Ruiz", "beto@x", "03/06/2026 09:00"),
    ]


# --- por trabajo (Pendientes por Práctico) ------------------------------------

def test_resumen_pendientes_muestra_fecha_de_creacion():
    data, _ = ExcelService(AsyncMock()).exportar_pendientes(_pendientes(), "Prog I")
    a2 = _wb(data)["Resumen"]["A2"].value
    assert "Documento generado el" in a2


def test_por_trabajo_columnas_con_tutor_y_email():
    data, fn = ExcelService(AsyncMock()).exportar_pendientes(_pendientes(), "Prog I", agrupar_por="trabajo")
    wb = _wb(data)
    assert wb.sheetnames[0] == "Resumen"
    assert "practico" in fn.lower()
    rows = list(wb["TP 1 - Variables"].iter_rows(values_only=True))
    assert ("Comisión", "Tutor", "Nombre", "Apellido", "Email", "Fecha de entrega") in rows
    # fila de Zoe: comisión, tutor al lado
    fila_zoe = next(r for r in rows if r[2] == "Zoe")
    assert fila_zoe[0] == "M26 C1-01" and fila_zoe[1] == "Prof Zoe"
    # orden por comisión: C1-01 (Zoe) antes que C1-02 (Ana)
    assert [r[2] for r in rows if r[2] in ("Ana", "Zoe")] == ["Zoe", "Ana"]


def test_por_trabajo_resumen_por_practico():
    data, _ = ExcelService(AsyncMock()).exportar_pendientes(_pendientes(), "Prog I", agrupar_por="trabajo")
    pares = [(r[0], r[1]) for r in _wb(data)["Resumen"].iter_rows(values_only=True)]
    assert ("Trabajo", "Pendientes") in pares
    assert ("TP 1 - Variables", 2) in pares
    assert ("Parcial 1", 1) in pares
    assert ("Total", 3) in pares


# --- por comisión (Pendientes por Comisión) -----------------------------------

def test_por_comision_titulo_con_tutor_y_columnas():
    data, fn = ExcelService(AsyncMock()).exportar_pendientes(_pendientes(), "Prog I", agrupar_por="comision")
    wb = _wb(data)
    assert wb.sheetnames[0] == "Resumen"
    assert set(wb.sheetnames) == {"Resumen", "M26 C1-01", "M26 C1-02"}
    assert "comision" in fn.lower()
    ws = wb["M26 C1-01"]
    flat = [str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None]
    assert any("Tutor: Prof Zoe" in s for s in flat)  # tutor en el título
    rows = list(ws.iter_rows(values_only=True))
    assert ("Trabajo", "Nombre", "Apellido", "Email", "Fecha de entrega") in rows
    # C1-01 ordenado por trabajo: "Parcial 1" (Beto) antes que "TP 1" (Zoe)
    assert [r[1] for r in rows if r[1] in ("Beto", "Zoe")] == ["Beto", "Zoe"]


def test_por_comision_resumen_con_tutor():
    data, _ = ExcelService(AsyncMock()).exportar_pendientes(_pendientes(), "Prog I", agrupar_por="comision")
    filas = [(r[0], r[1], r[2]) for r in _wb(data)["Resumen"].iter_rows(values_only=True)]
    assert ("Comisión", "Tutor", "Pendientes") in filas
    assert ("M26 C1-01", "Prof Zoe", 2) in filas
    assert ("M26 C1-02", "Prof Ana", 1) in filas
    # Total en la 3ra columna
    total = next(r for r in filas if r[0] == "Total")
    assert total[2] == 3


def test_pendientes_vacio_solo_resumen():
    data, _ = ExcelService(AsyncMock()).exportar_pendientes([], "Prog I", agrupar_por="trabajo")
    assert _wb(data).sheetnames == ["Resumen"]
