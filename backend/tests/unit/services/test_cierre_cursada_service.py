"""
Tests de los helpers puros de CierreCursadaService (parseo de porcentaje,
interpretación del TP en escala, resolución de comisión por grupo Moodle) y de
preservación del histórico append-only del repository (bugfix cierre-cursada-fix,
tarea 2 — Property 2: Preservation, Req 3.3/3.4/3.6).

No cubren el flujo completo de generar() (requiere Moodle real/mock pesado) —
ver test_cierre_cursada_calculo.py para el árbol de decisión, que es lo que
de verdad no se puede calcular mal.
"""

from types import SimpleNamespace

import pytest
import pytest_asyncio
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

from app.models.base import Base
from app.models.cierre_cursada import CierreCursadaAlumno, CierreCursadaRun
from app.models.cohorte import Cohorte, Cuatrimestre
from app.models.enums import EstadoCierreEnum
from app.models.materia import Materia
from app.models.usuario import Usuario
from app.repositories.cierre_cursada_repository import CierreCursadaRepository
from app.services.cierre_cursada_service import CierreCursadaService


# SQLite no tiene un tipo JSONB nativo; se lo compila como JSON (mismo comportamiento
# observable — guardar/leer un dict). Solo afecta al dialecto sqlite de estos tests —
# Postgres sigue usando JSONB real en producción.
@compiles(JSONB, "sqlite")
def _jsonb_como_json_en_sqlite(element, compiler, **kw):
    return "JSON"


@pytest_asyncio.fixture
async def db_session():
    """Sesión SQLite en memoria con SOLO las tablas que este test necesita
    (usuarios/cohortes/cuatrimestres/materias/cierre_cursada_*): el metadata GLOBAL
    (Base.metadata, todas las tablas de la app) trae columnas Postgres-only (ARRAY,
    JSONB de otras tablas) que SQLite no puede compilar, así que no se puede usar
    `Base.metadata.create_all` sin filtrar acá."""
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False, future=True)
    tablas = [
        Usuario.__table__,
        Cohorte.__table__,
        Cuatrimestre.__table__,
        Materia.__table__,
        CierreCursadaRun.__table__,
        CierreCursadaAlumno.__table__,
    ]
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all, tables=tablas)

    session_maker = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with session_maker() as session:
        yield session

    await engine.dispose()


# NOTA (tarea 11.1/11.4, bugfix cierre-cursada-fix): se quitaron los tests
# `test_parse_pct` y `test_tp_resultado` que ejercitaban `_parse_pct`/`_tp_resultado`.
# Esas funciones pertenecían al sistema PARALELO de mapeo manual + umbrales fijos
# (REGLAS_CIERRE_DEFAULT / CierreCursadaItem) que este bugfix elimina intencionalmente
# (tarea 11.1): el cierre ahora clasifica dirigido por `ExamenMateria`
# (`calcular_estado_cierre` + `calcular_nota_final`), no por parseo de porcentajes de
# ítems del calificador. No son parte de la Property 2 (Preservation) — esa property
# cubre EXCLUSIVAMENTE rescate/ESCALA/comisión/append-only/estilos (Req 3.1-3.6), no el
# sistema de umbrales fijos eliminado.


# ===================== _resolver_comision =====================


def _comision(id_, nombre, moodle_group_code, moodle_group_id=None, tutores=()):
    return SimpleNamespace(
        id=id_,
        nombre=nombre,
        moodle_group_code=moodle_group_code,
        moodle_group_id=moodle_group_id,
        tutores=tutores,
    )


def _tutor(nombre):
    return SimpleNamespace(tutor=SimpleNamespace(nombre=nombre))


# NOTA (tarea 3.1, bugfix cierre-cursada-comision-nota-fix): estos tests migraron a la
# nueva firma `_resolver_comision(grupos: list[dict], comisiones)`. El bridge que
# ejercitaban (matcheo por `moodle_group_code`, texto manual) fue REEMPLAZADO por el fix
# (Bug 1) por `moodle_group_id`/nombre derivado — ver design.md. Se adaptan los fixtures
# para usar los bridges nuevos, preservando la intención/aserciones originales de cada
# test (match único, case-insensitive, sin match, ambiguo, unión de tutores, sin bridge
# configurado -> sin match).


def test_resolver_comision_match_unico():
    comisiones = [
        _comision(1, "Comisión 1", None, moodle_group_id=100, tutores=[_tutor("Ana Pérez")]),
        _comision(2, "Comisión 2", None, moodle_group_id=200),
    ]
    grupos = [{"id": 100, "name": "Comisión 1"}]
    cid, nombre, tutor = CierreCursadaService._resolver_comision(grupos, comisiones)
    assert (cid, nombre, tutor) == (1, "Comisión 1", "Ana Pérez")


def test_resolver_comision_case_insensitive():
    # Case-insensitive vía el fallback por nombre derivado (parse_comision): el grupo
    # matchea el formato de comisión y el código resultante se compara case-insensitive
    # contra Comision.nombre.
    comisiones = [_comision(1, "m26 c1-01", None)]
    grupos = [{"id": None, "name": "M26 C1-01"}]
    cid, _, _ = CierreCursadaService._resolver_comision(grupos, comisiones)
    assert cid == 1


def test_resolver_comision_sin_match():
    comisiones = [_comision(1, "Comisión 1", None, moodle_group_id=100)]
    grupos = [{"id": 999, "name": "Otro Grupo"}]
    cid, nombre, tutor = CierreCursadaService._resolver_comision(grupos, comisiones)
    assert (cid, nombre, tutor) == (None, "Sin comisión asignada", None)


def test_resolver_comision_multiples_matches_no_rompe():
    # Alumno en dos grupos que matchean dos comisiones distintas -> ambiguo, sin comisión.
    comisiones = [
        _comision(1, "Comisión 1", None, moodle_group_id=100),
        _comision(2, "Comisión 2", None, moodle_group_id=200),
    ]
    grupos = [{"id": 100, "name": "Comisión 1"}, {"id": 200, "name": "Comisión 2"}]
    cid, nombre, tutor = CierreCursadaService._resolver_comision(grupos, comisiones)
    assert (cid, nombre, tutor) == (None, "Sin comisión asignada", None)


def test_resolver_comision_multiples_tutores_se_unen_con_barra():
    comisiones = [
        _comision(
            1, "Comisión 1", None, moodle_group_id=100,
            tutores=[_tutor("Ana Pérez"), _tutor("Juan Gómez")],
        ),
    ]
    grupos = [{"id": 100, "name": "Comisión 1"}]
    _, _, tutor = CierreCursadaService._resolver_comision(grupos, comisiones)
    assert tutor == "Ana Pérez / Juan Gómez"


def test_resolver_comision_sin_moodle_group_code_no_matchea():
    # Sin ningún bridge configurado (ni moodle_group_id ni nombre derivable) -> sin match.
    comisiones = [_comision(1, "Comisión 1", None)]
    grupos = [{"id": None, "name": "Cualquier Grupo"}]
    cid, nombre, _ = CierreCursadaService._resolver_comision(grupos, comisiones)
    assert (cid, nombre) == (None, "Sin comisión asignada")


# ===================== Bug 1 exploration (tarea 1, ANTES del fix) =====================
#
# Property 1: Bug Condition - Resolución de comisión por grupo de Moodle (design.md).
# CRITICAL: estos tests codifican la NUEVA firma de _resolver_comision(grupos, comisiones)
# donde `grupos` es la lista completa de dicts [{id, name}, ...] que trae Moodle (no sólo
# nombres) — ver Fix Implementation / bosquejo en design.md. El código SIN arreglar todavía
# tiene la firma vieja (`grupos: list[str]`, matcheo por `moodle_group_code`), así que
# invocarlo con la firma nueva DEBE FALLAR (ya sea por resultado incorrecto o porque el
# código no sabe iterar dicts) — la falla confirma que el Bug 1 existe.
# **Validates: Requirements 1.1, 1.2**


def test_bug_resolver_comision_match_por_moodle_group_id():
    """Bug 1: grupo {"id": 342, "name": "M26 C1-09"} matchea una Comision real por
    moodle_group_id=342 (moodle_group_code=None, el puente frágil que usa el código
    sin arreglar). Esperado (tras el fix): resuelve comision_id=1. HOY: el código sin
    arreglar sólo sabe matchear por moodle_group_code (texto manual, aquí None) y además
    espera `grupos` como lista de nombres (str), no de dicts -> no resuelve la comisión
    real (devuelve "Sin comisión asignada" o directamente no puede procesar el dict)."""
    comisiones = [_comision(1, "M26 C1-09", moodle_group_code=None, moodle_group_id=342)]
    grupos = [{"id": 342, "name": "M26 C1-09"}]

    cid, nombre, tutor = CierreCursadaService._resolver_comision(grupos, comisiones)

    assert (cid, nombre, tutor) == (1, "M26 C1-09", None)


def test_bug_resolver_comision_match_por_nombre_derivado():
    """Bug 1: grupo {"id": None, "name": "M26 C1-09"} matchea una Comision real por
    nombre derivado (parse_comision("M26 C1-09") == "M26 C1-09" == Comision.nombre),
    sin moodle_group_code configurado. Esperado (tras el fix): resuelve comision_id=1.
    HOY: el código sin arreglar no deriva el código de comisión del nombre del grupo
    (no usa parse_comision) y depende 100% de moodle_group_code -> no resuelve la
    comisión real."""
    comisiones = [_comision(1, "M26 C1-09", moodle_group_code=None)]
    grupos = [{"id": None, "name": "M26 C1-09"}]

    cid, nombre, tutor = CierreCursadaService._resolver_comision(grupos, comisiones)

    assert (cid, nombre, tutor) == (1, "M26 C1-09", None)


# ===================== Bug 1 preservation (tarea 2, ANTES del fix) =====================
#
# Property 3 (Preservation) — Req 3.1/3.2 (design.md, cierre-cursada-comision-nota-fix):
# un alumno SIN grupo de comisión válido, o con match AMBIGUO, sigue quedando
# "Sin comisión asignada" sin lanzar excepción, tanto antes como después del fix de la
# tarea 3.1. Metodología observation-first: se corre el código SIN arreglar con estos
# inputs (que NO disparan el Bug 1) y se captura el comportamiento observado.
#
# Los casos "sin match" (3.1) y "ambiguo" (3.2) con la firma VIEJA
# (`grupos: list[str]`, matcheo por `moodle_group_code`) ya están cubiertos arriba por
# `test_resolver_comision_sin_match` y `test_resolver_comision_multiples_matches_no_rompe`
# — no se duplican acá (ver instrucción de la tarea 2).
#
# NOTA sobre la firma NUEVA ([{id, name}], la que introduce el fix de la tarea 3.1): al
# ejercitar `_resolver_comision` SIN arreglar con una lista de dicts NO VACÍA, el código
# actual rompe con `AttributeError` (`g.strip()` sobre un dict) para CUALQUIER dict, sea
# cual sea su contenido — es decir, el código sin arreglar ni siquiera llega a evaluar
# "sin match" o "ambiguo" con ese shape de entrada (comportamiento confirmado corriendo
# el código). El único caso con la firma nueva que hoy se puede observar sin excepción
# es la lista de grupos VACÍA (alumno sin ningún grupo en Moodle): es shape-agnóstica
# (nunca itera un elemento individual) y sigue devolviendo "Sin comisión asignada" tanto
# antes como después del fix.


def test_resolver_comision_sin_grupos_es_sin_comision_asignada():
    """Alumno sin ningún grupo de Moodle (lista vacía) -> "Sin comisión asignada",
    sin excepción. Válido con cualquier shape de `grupos` (`list[str]` o
    `list[dict]`) porque una lista vacía nunca itera un elemento individual."""
    comisiones = [_comision(1, "M26 C1-09", moodle_group_code=None, moodle_group_id=342)]

    cid, nombre, tutor = CierreCursadaService._resolver_comision([], comisiones)

    assert (cid, nombre, tutor) == (None, "Sin comisión asignada", None)


# ===================== Histórico append-only (CierreCursadaRepository) =====================
#
# Property 2 (Preservación) — Req 3.4: dos corridas de la misma materia coexisten en el
# histórico, la más reciente aparece primero, y crear una NO sobrescribe la anterior.
# Metodología observación-primero: se observa el comportamiento YA EXISTENte de
# crear_run/listar_runs sobre el código sin arreglar (este repository no cambia en el fix
# salvo eliminar get_mapping/upsert_mapping — tarea 9).


def _run_minimo(materia_id: int, cuatrimestre_id: int, generado_por_id: int) -> CierreCursadaRun:
    """CierreCursadaRun con el shape MINIMO válido hoy (columnas legacy no-nullable
    todavía, tal como está el modelo ANTES del fix)."""
    alumno = CierreCursadaAlumno(
        moodle_user_id=1,
        nombre="Ana",
        apellido="Gómez",
        email="ana@x.com",
        comision_id=None,
        comision_nombre="Sin comisión asignada",
        tutor_nombre=None,
        tps=[],
        autoeval_pcts=[],
        parcial1_instancias=[],
        parcial2_instancias=[],
        tpi_instancias=[],
        tp_ok=True,
        autoeval_ok=True,
        p1_max=60.0,
        p2_max=60.0,
        tpi_max=60.0,
        estado=EstadoCierreEnum.PROMOCIONA,
        nota_final=6,
        habilitado_final="No aplica (promocionó)",
    )
    run = CierreCursadaRun(
        materia_id=materia_id,
        cuatrimestre_id=cuatrimestre_id,
        umbral_tp_pct=50.0,
        reglas_snapshot={"parcial_promocion_min_pct": 60},
        generado_por_id=generado_por_id,
        total_alumnos=1,
        total_promociona=1,
        total_regulariza=0,
        total_recursa=0,
    )
    run.alumnos = [alumno]
    return run


@pytest.mark.asyncio
async def test_pbt_historico_append_only_dos_corridas_coexisten(db_session):
    """Dos corridas de la MISMA materia persistidas una tras otra: ambas siguen
    existiendo (append-only), listar_runs las devuelve ordenadas por más reciente
    primero, y la segunda NO pisa/reemplaza los datos de la primera."""
    from app.models.enums import RolEnum

    # Fixtures mínimas de FK (materia + cuatrimestre + usuario) para poder persistir un run.
    usuario = Usuario(
        username="gestor1", password_hash="x", nombre="Gestor", rol=RolEnum.ADMIN,
    )
    cohorte = Cohorte(codigo="M26", nombre="M26")
    db_session.add_all([usuario, cohorte])
    await db_session.flush()
    cuatrimestre = Cuatrimestre(nombre="1C", cohorte_id=cohorte.id, numero=1)
    materia = Materia(nombre="Programación 1", codigo="P1")
    db_session.add_all([cuatrimestre, materia])
    await db_session.flush()

    repo = CierreCursadaRepository(db_session)

    primera = await repo.crear_run(_run_minimo(materia.id, cuatrimestre.id, usuario.id))
    segunda = await repo.crear_run(_run_minimo(materia.id, cuatrimestre.id, usuario.id))

    assert primera.id != segunda.id  # dos filas distintas, no un upsert

    runs = await repo.listar_runs(materia.id)
    assert len(runs) == 2  # ambas coexisten
    assert {r.id for r in runs} == {primera.id, segunda.id}
    # Más reciente primero (created_at desc; con ids autoincrementales en el mismo test
    # la segunda corrida es la más reciente).
    assert runs[0].id == segunda.id
    assert runs[1].id == primera.id

    # La primera corrida sigue intacta (no se sobrescribió al crear la segunda).
    recargada = await repo.get_run(primera.id)
    assert recargada is not None
    assert recargada.total_alumnos == 1
    assert recargada.alumnos[0].apellido == "Gómez"


# ===================== Estilos de la casa en el Excel del cierre (Req 3.5) =====================
#
# Property 2 (Preservación): el .xlsx generado sigue usando la banda de título, los
# encabezados, las filas de datos y la barra de bloque de excel_estilos.py (paleta y
# bordes compartidos con el resto de reportes de la plataforma). Se observa el
# comportamiento YA EXISTENTE del generador actual (columnas TP/Autoeval/etc. — este
# archivo se reescribe en la tarea 12, pero los estilos de la casa deben seguir).


def test_excel_cierre_usa_paleta_y_bordes_de_excel_estilos():
    import io
    from datetime import datetime

    from openpyxl import load_workbook

    from app.services import excel_estilos
    from app.services.excel_cierre_cursada import generar_excel_cierre

    alumno = CierreCursadaAlumno(
        moodle_user_id=1, nombre="Ana", apellido="Gómez", email="ana@x.com",
        comision_id=None, comision_nombre="Comisión 1", tutor_nombre="Juan Pérez",
        tps=[], autoeval_pcts=[], parcial1_instancias=[], parcial2_instancias=[], tpi_instancias=[],
        tp_ok=True, autoeval_ok=True, p1_max=60.0, p2_max=60.0, tpi_max=60.0,
        estado=EstadoCierreEnum.PROMOCIONA, nota_final=6, habilitado_final="No aplica (promocionó)",
    )
    run = CierreCursadaRun(
        id=1, materia_id=1, cuatrimestre_id=1, umbral_tp_pct=50.0,
        reglas_snapshot={}, generado_por_id=1, total_alumnos=1,
        total_promociona=1, total_regulariza=0, total_recursa=0,
    )
    run.created_at = datetime(2026, 1, 1, 12, 0)
    run.alumnos = [alumno]

    data, _ = generar_excel_cierre("Programación 1", run)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    # Título (R1, banda_titulo): fill AZUL claro (FILL_TITULO) + fuente AZUL_TITULO.
    titulo = ws["A1"]
    assert titulo.font.color.rgb.endswith(excel_estilos.AZUL_TITULO)
    assert titulo.fill.fgColor.rgb.endswith("DCE6F1")

    # Encabezados de columna (celda_header): fondo AZUL_TITULO, fuente blanca, con borde.
    header_cell = next(c for row in ws.iter_rows() for c in row if c.value == "Nombre y Apellido")
    assert header_cell.fill.fgColor.rgb.endswith(excel_estilos.AZUL_TITULO)
    assert header_cell.font.color.rgb.endswith("FFFFFF")
    assert header_cell.border.left.style == "thin"

    # Barra de bloque por comisión (barra_bloque): fondo AZUL_BLOQUE, fuente blanca bold.
    barra_cell = next(
        c for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.startswith("Comisión 1")
    )
    assert barra_cell.fill.fgColor.rgb.endswith(excel_estilos.AZUL_BLOQUE)
    assert barra_cell.font.bold is True

    # Fila de datos (fila_datos): borde en cada celda.
    dato_cell = next(c for row in ws.iter_rows() for c in row if c.value == "Gómez, Ana")
    assert dato_cell.border.left.style == "thin"
