"""PERF-004 — no-regresión del export de NOTAS (calificaciones oficiales).

El render openpyxl de `ExcelService.exportar_notas_excel` pasó a correr en un thread
(`asyncio.to_thread`) para no bloquear el event loop. La REGLA DE ORO es que el archivo
generado sea IDÉNTICO al de antes: mismas celdas, formato y orden. Estos tests fijan esa
salida.

- TEST A (unidad pura): `_build_notas_workbook_sync` sobre `filas` ya materializadas
  (sin DB) → verifica título, subtítulo, headers, filas de alumnos (corregida + sin
  corregir), nota, estado, fecha, editado, comentario y la fila de total.
- TEST B (extremo a extremo, no-MissingGreenlet): sembrando SQLite async y llamando a
  `exportar_notas_excel`, que atraviesa `asyncio.to_thread` con una sesión async real.
  Si algún dato del acta se cargara lazy dentro del thread, esto reventaría con
  MissingGreenlet. Comprueba que el .xlsx trae los datos y que el nombre del archivo es el
  esperado.
"""

import io
import json
import sqlite3
from datetime import datetime
from decimal import Decimal

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy import ARRAY
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

# sqlite3 no bindea dict/list (JSONB/ARRAY) directo; estos adapters los serializan a JSON.
sqlite3.register_adapter(dict, json.dumps)
sqlite3.register_adapter(list, json.dumps)

from app.models.base import Base
from app.models.comision import Comision
from app.models.correccion import Correccion
from app.models.entrega import Entrega
from app.models.enums import EstadoEntregaEnum
from app.models.materia import Materia
from app.models.rubrica import Rubrica
from app.services.excel_service import ExcelService

UNIV_ID = 1  # multi-tenant-scoping-queries: universidad_id ahora NOT NULL



# SQLite no soporta JSONB/ARRAY nativos; se compilan a JSON sólo en el dialecto de tests.
@compiles(JSONB, "sqlite")
def _jsonb_como_json_en_sqlite(element, compiler, **kw):  # noqa: D401
    return "JSON"


@compiles(ARRAY, "sqlite")
def _array_como_json_en_sqlite(element, compiler, **kw):  # noqa: D401
    return "JSON"


@pytest_asyncio.fixture
async def db_session():
    import app.models  # noqa: F401 — puebla el registry de mappers/tablas

    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False, future=True)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session_maker = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with session_maker() as session:
        yield session
    await engine.dispose()


# ---------------------------------------------------------------------------
# TEST A — builder puro (sin DB): la salida openpyxl es la esperada
# ---------------------------------------------------------------------------


def test_build_notas_workbook_sync_contenido_y_formato():
    """`_build_notas_workbook_sync` produce el acta con el layout y los datos exactos."""
    filas = [
        {
            "alumno_nombre": "Ada Lovelace",
            "estado": EstadoEntregaEnum.CORREGIDA,
            "correccion": {
                "nota": Decimal("85"),
                "created_at": datetime(2026, 7, 18, 10, 30),
                "editado_manualmente": True,
                "comentario_general": "Buen trabajo",
            },
        },
        {
            "alumno_nombre": "Alan Turing",
            "estado": EstadoEntregaEnum.SUBIDA,  # sin corrección
            "correccion": None,
        },
    ]

    svc = ExcelService(db=None)  # el builder no toca la DB
    data = svc._build_notas_workbook_sync(
        filas,
        len(filas),
        "Programación 1",  # materia_nombre
        "P1",  # materia_codigo
        "C1-1",  # comision_nombre
        "TP Numero 1",  # rubrica_nombre
        "TP",  # rubrica_tipo
        "1",  # rubrica_numero
    )
    assert isinstance(data, bytes) and len(data) > 0

    ws = load_workbook(io.BytesIO(data)).active
    assert ws.title == "Notas"

    # Título + subtítulo
    assert ws["A1"].value == "Programación 1 (P1)"
    assert ws["A2"].value == "Comisión: C1-1  |  TP: TP 1 - TP Numero 1"

    # Headers (fila 4)
    headers = [ws.cell(row=4, column=c).value for c in range(1, 7)]
    assert headers == [
        "Alumno", "Nota", "Estado", "Fecha Corrección", "Editado", "Comentario General",
    ]

    # Fila 5 → alumno corregido
    assert ws.cell(row=5, column=1).value == "Ada Lovelace"
    assert float(ws.cell(row=5, column=2).value) == 85.0
    assert ws.cell(row=5, column=3).value == "CORREGIDA"
    assert ws.cell(row=5, column=4).value == "18/07/2026 10:30"
    assert ws.cell(row=5, column=5).value == "Sí"
    assert ws.cell(row=5, column=6).value == "Buen trabajo"

    # Fila 6 → alumno sin corregir (SUBIDA → "PENDIENTE" en la columna Estado)
    assert ws.cell(row=6, column=1).value == "Alan Turing"
    assert ws.cell(row=6, column=2).value == "-"
    assert ws.cell(row=6, column=3).value == "PENDIENTE"
    assert ws.cell(row=6, column=4).value == "-"
    assert ws.cell(row=6, column=5).value == "-"

    # Fila de total (una fila en blanco después de la última): row 8
    assert ws.cell(row=8, column=1).value == "Total de entregas: 2"


def test_build_notas_workbook_sync_lista_vacia_es_valida():
    """Sin filas, el workbook sigue siendo válido (headers + total 0)."""
    svc = ExcelService(db=None)
    data = svc._build_notas_workbook_sync(
        [], 0, "Mat", "M1", "C1", "Rub", "TP", "2"
    )
    ws = load_workbook(io.BytesIO(data)).active
    assert ws["A1"].value == "Mat (M1)"
    assert ws.cell(row=4, column=1).value == "Alumno"
    # Sin datos, current_row = 5; total en row 6.
    assert ws.cell(row=6, column=1).value == "Total de entregas: 0"


# ---------------------------------------------------------------------------
# TEST B — extremo a extremo: exportar_notas_excel a través de asyncio.to_thread
# ---------------------------------------------------------------------------


async def _seed(db) -> dict:
    """Materia + comisión + rúbrica + dos entregas (una corregida, una sin corregir)."""
    materia = Materia(universidad_id=UNIV_ID, nombre="Programación 1", codigo="P1")
    db.add(materia)
    await db.flush()
    comi = Comision(universidad_id=UNIV_ID, materia_id=materia.id, nombre="C1-1", anio=2026, activa=True)
    db.add(comi)
    await db.flush()
    rubrica = Rubrica(universidad_id=UNIV_ID, 
        materia_id=materia.id,
        titulo="Trabajo Práctico 1",
        descripcion="Rúbrica de prueba",
        numero=1,
        anio=2026,
        tipo="TP",
        puntaje_maximo=100,
        criterios_json=[],
    )
    db.add(rubrica)
    await db.flush()

    entrega_corregida = Entrega(universidad_id=UNIV_ID, 
        comision_id=comi.id,
        rubrica_id=rubrica.id,
        alumno_nombre="Ada Lovelace",
        archivo_nombre="tp.zip",
        archivo_tipo="zip",
        contenido_preview="print('hola')",
        estado=EstadoEntregaEnum.CORREGIDA,
    )
    entrega_pendiente = Entrega(universidad_id=UNIV_ID, 
        comision_id=comi.id,
        rubrica_id=rubrica.id,
        alumno_nombre="Alan Turing",
        archivo_nombre="tp2.zip",
        archivo_tipo="zip",
        contenido_preview="print('mundo')",
        estado=EstadoEntregaEnum.SUBIDA,
    )
    db.add_all([entrega_corregida, entrega_pendiente])
    await db.flush()

    correccion = Correccion(universidad_id=UNIV_ID, 
        entrega_id=entrega_corregida.id,
        nota=Decimal("90"),
        criterios_json={"criterios": []},
        comentario_general="Excelente",
        editado_manualmente=False,
    )
    db.add(correccion)
    await db.commit()
    return {"comision_id": comi.id, "rubrica_id": rubrica.id}


@pytest.mark.asyncio
async def test_exportar_notas_excel_end_to_end(db_session):
    """exportar_notas_excel corre el render en un thread SIN MissingGreenlet y el .xlsx
    trae los datos del acta (esta es la corrida con sesión async real)."""
    ids = await _seed(db_session)
    service = ExcelService(db_session)

    data, filename = await service.exportar_notas_excel(
        comision_id=ids["comision_id"], rubrica_id=ids["rubrica_id"]
    )

    assert isinstance(data, bytes) and len(data) > 0
    assert filename.startswith("notas_P1_")
    assert "TP1" in filename
    assert filename.endswith(".xlsx")

    ws = load_workbook(io.BytesIO(data)).active
    assert ws["A1"].value == "Programación 1 (P1)"
    assert ws.cell(row=4, column=1).value == "Alumno"

    # Las dos filas de datos están (orden por created_at desc del repo; comprobamos
    # presencia, no orden, para no acoplar el test al orden interno del repositorio).
    alumnos = {ws.cell(row=r, column=1).value for r in (5, 6)}
    assert alumnos == {"Ada Lovelace", "Alan Turing"}

    # La nota 90 de Ada aparece en alguna de las dos filas de datos.
    notas = {
        str(ws.cell(row=r, column=2).value) for r in (5, 6)
    }
    assert "90" in notas or "90.0" in notas

    # Comentario general de la corrección.
    comentarios = {ws.cell(row=r, column=6).value for r in (5, 6)}
    assert "Excelente" in comentarios

    # Fila de total (dos entregas → row 8).
    assert ws.cell(row=8, column=1).value == "Total de entregas: 2"
