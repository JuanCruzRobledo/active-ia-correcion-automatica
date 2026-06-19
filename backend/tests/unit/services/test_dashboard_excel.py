"""Tests de construir_excel_avance (Excel de avance del Dashboard de Gestores).

Verifica el fix del placeholder: la columna ahora muestra LAS DEUDAS del alumno
(igual que los emails), con la etiqueta de la materia (Unidad/Semana), en vez del
viejo "Unidad 8 (Unidad 8)".
"""

import io
from types import SimpleNamespace

from openpyxl import load_workbook

from app.models.enums import EstadoAvanceEnum
from app.services.dashboard_excel import agregar_examenes, construir_excel_avance


def _alumno(apellido, nombre, alcanzada, deudas, desaprobada=False, examenes=None):
    return SimpleNamespace(
        apellido=apellido, nombre=nombre, email=f"{nombre}@x.com",
        comision="M26 C1-09", unidad_alcanzada=alcanzada,
        actividades_faltantes={"deudas": deudas} if deudas else None,
        actividad_actual_desaprobada=desaprobada,
        resultados_examenes={"examenes": examenes} if examenes else None,
    )


def _build(etiqueta):
    alumnos = {
        EstadoAvanceEnum.AL_DIA: [
            _alumno("Gómez", "Ana", 8, None, examenes=[
                {"etiqueta": "Parcial 1", "resultado": "aprobado", "rescatado": False},
            ]),
        ],
        EstadoAvanceEnum.RIESGO_ALTO: [
            _alumno("Páez", "Beto", 8, [
                {"unidad": 8, "tipo": "QUIZ", "estado": "pendiente"},
                {"unidad": 8, "tipo": "TP", "estado": "desaprobado"},
            ], desaprobada=True),
        ],
        EstadoAvanceEnum.RIESGO_MEDIO: [],
        EstadoAvanceEnum.SIN_ACTIVIDAD: [],
    }
    conteos = {EstadoAvanceEnum.AL_DIA: 1, EstadoAvanceEnum.RIESGO_ALTO: 1}
    data = construir_excel_avance("Test", conteos, alumnos, etiqueta)
    return load_workbook(io.BytesIO(data))


def test_header_usa_etiqueta_semana():
    wb = _build("Semana")
    ws = wb["Riesgo alto"]
    headers = [c.value for c in ws[1]]
    assert "Semana alcanzada" in headers
    assert "Le falta" in headers


def test_columna_le_falta_muestra_deudas_no_placeholder():
    wb = _build("Semana")
    ws = wb["Riesgo alto"]
    fila = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    le_falta = fila[5]
    # Antes salía "Unidad 8 (Unidad 8)"; ahora las deudas reales con etiqueta "Semana".
    assert le_falta == "Semana 8: Quiz, TP (desaprobado)"
    assert "(Unidad" not in (le_falta or "")


def test_al_dia_sin_deudas_celda_vacia():
    wb = _build("Unidad")
    ws = wb["Al día"]
    fila = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert (fila[5] or "") == ""  # al día → no le falta nada


def test_columna_examenes():
    wb = _build("Unidad")
    headers = [c.value for c in wb["Al día"][1]]
    assert "Exámenes" in headers
    fila = list(wb["Al día"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert fila[6] == "Parcial 1: Aprobó"  # columna 7 (índice 6)


def test_resumen_muestra_fecha_creacion_y_corte():
    alumnos = {e: [] for e in EstadoAvanceEnum}
    data = construir_excel_avance("Test", {}, alumnos, "Semana", 8, "Parcial 2")
    resumen = load_workbook(io.BytesIO(data))["Resumen"]
    # El corte va en el título (A1, mismo peso); la fecha como metadata (A2).
    assert "Tareas requeridas: Semana 8, Parcial 2" in resumen["A1"].value
    assert "Documento generado el" in resumen["A2"].value


def test_resumen_sin_unidad_actual_solo_fecha():
    alumnos = {e: [] for e in EstadoAvanceEnum}
    data = construir_excel_avance("Test", {}, alumnos, "Unidad", None)
    resumen = load_workbook(io.BytesIO(data))["Resumen"]
    assert "Tareas requeridas" not in resumen["A1"].value
    assert "Documento generado el" in resumen["A2"].value


# ===================== hoja de exámenes (estado de parciales) =====================


def _ex(etiqueta, resultado, rescatado=False):
    return {"etiqueta": etiqueta, "resultado": resultado, "rescatado": rescatado}


def test_agregar_examenes_cuenta_aprobado_en_cualquier_instancia():
    # El rescate YA viene aplicado en `resultado`: un parcial rescatado figura "aprobado".
    alumnos = [
        SimpleNamespace(resultados_examenes={"examenes": [
            _ex("Parcial 1", "aprobado"),
            _ex("Parcial 2", "desaprobado"),
        ]}),
        SimpleNamespace(resultados_examenes={"examenes": [
            _ex("Parcial 1", "aprobado", rescatado=True),  # rescatado → cuenta aprobado
            _ex("Parcial 2", "ausente"),
        ]}),
        SimpleNamespace(resultados_examenes={"examenes": [
            _ex("Parcial 1", "desaprobado"),
            _ex("Parcial 2", "aprobado"),
        ]}),
    ]
    filas = agregar_examenes(alumnos)
    # Orden estable por aparición: Parcial 1, luego Parcial 2.
    assert [f["etiqueta"] for f in filas] == ["Parcial 1", "Parcial 2"]
    assert filas[0] == {
        "etiqueta": "Parcial 1", "aprobado": 2, "desaprobado": 1, "ausente": 0,
        "sin_corregir": 0, "total": 3,
    }
    assert filas[1] == {
        "etiqueta": "Parcial 2", "aprobado": 1, "desaprobado": 1, "ausente": 1,
        "sin_corregir": 0, "total": 3,
    }


def test_agregar_examenes_cuenta_sin_corregir():
    alumnos = [
        SimpleNamespace(resultados_examenes={"examenes": [_ex("Global 1", "sin_corregir")]}),
        SimpleNamespace(resultados_examenes={"examenes": [_ex("Global 1", "sin_corregir")]}),
        SimpleNamespace(resultados_examenes={"examenes": [_ex("Global 1", "aprobado")]}),
    ]
    fila = agregar_examenes(alumnos)[0]
    assert fila["sin_corregir"] == 2
    assert fila["aprobado"] == 1
    assert fila["total"] == 3


def test_agregar_examenes_sin_examenes_devuelve_vacio():
    alumnos = [SimpleNamespace(resultados_examenes=None)]
    assert agregar_examenes(alumnos) == []


def test_hoja_examenes_existe_con_conteos():
    alumnos = {
        EstadoAvanceEnum.AL_DIA: [
            _alumno("Gómez", "Ana", 8, None, examenes=[
                _ex("Parcial 1", "aprobado"), _ex("Parcial 2", "aprobado"),
            ]),
        ],
        EstadoAvanceEnum.RIESGO_ALTO: [
            _alumno("Páez", "Beto", 8, None, examenes=[
                _ex("Parcial 1", "desaprobado"), _ex("Parcial 2", "ausente"),
            ]),
        ],
        EstadoAvanceEnum.RIESGO_MEDIO: [],
        EstadoAvanceEnum.SIN_ACTIVIDAD: [],
    }
    data = construir_excel_avance("Test", {}, alumnos, "Unidad", 8, "Parcial 2")
    wb = load_workbook(io.BytesIO(data))
    assert "Exámenes" in wb.sheetnames
    ws = wb["Exámenes"]
    # Bloque 1 (Parcial 1) arranca en fila 4: barra, header, 4 resultados, total.
    assert ws["A4"].value == "Parcial 1"
    assert ws["A6"].value == "Aprobado" and ws["B6"].value == 1
    assert ws["A7"].value == "Desaprobado" and ws["B7"].value == 1
    assert ws["A8"].value == "Sin corregir" and ws["B8"].value == 0
    assert ws["A9"].value == "Ausente" and ws["B9"].value == 0
    assert ws["A10"].value == "Total" and ws["B10"].value == 2


def test_sin_examenes_no_crea_hoja():
    alumnos = {e: [] for e in EstadoAvanceEnum}
    data = construir_excel_avance("Test", {}, alumnos, "Unidad", 8)
    wb = load_workbook(io.BytesIO(data))
    assert "Exámenes" not in wb.sheetnames
