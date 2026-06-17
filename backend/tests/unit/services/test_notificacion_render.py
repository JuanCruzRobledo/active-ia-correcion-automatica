"""
Tests de notificacion_render (T4 — adjuntos de las notificaciones).

- formatear_faltantes: lógica PURA texto → TDD fino.
- construir_excel_avance: devuelve bytes; se verifica estructura (hojas, resumen, bloques).
"""

import io

import pytest
from openpyxl import load_workbook

from app.services.notificacion_render import (
    construir_excel_avance,
    construir_html_alumno,
    examen_para_atender,
    formatear_examenes,
    formatear_faltantes,
    label_resultado_examen,
    subtitulo_documento,
    tiene_examen_desaprobado,
    titulo_con_corte,
)


def test_subtitulo_documento_solo_fecha():
    assert subtitulo_documento().startswith("Documento generado el")


def test_titulo_con_corte():
    assert titulo_con_corte("Prog 1", "Unidad", 8, "Parcial 2") == "Prog 1  —  Tareas requeridas: Unidad 8, Parcial 2"
    assert titulo_con_corte("PYE", "Semana", 13, None) == "PYE  —  Tareas requeridas: Semana 13"
    assert titulo_con_corte("Prog 1", "Unidad", None, None) == "Prog 1"  # sin corte → solo el nombre


# ===================== construir_excel_avance (formato de referencia) =====================


def _al(comision, apellido, nombre, deudas, *, estado="RIESGO_ALTO", examenes=None):
    return {
        "comision": comision, "apellido": apellido, "nombre": nombre,
        "deudas": deudas, "estado": estado, "examenes": examenes or [],
    }


MATERIAS_AVANCE = [
    {
        "materia": "Programación 1",
        "etiqueta": "Unidad",
        "alumnos": [
            _al("M26 C1-01", "FERREIRA", "HERNAN", [
                {"unidad": 3, "tipo": "TP", "estado": "no entregado"},
                {"unidad": 3, "tipo": "AUTOEVALUACION", "estado": "pendiente"},
                {"unidad": 4, "tipo": "TP", "estado": "desaprobado"},
            ]),
            _al("M26 C1-02", "REYNOSO", "MARIA", [
                {"unidad": 3, "tipo": "QUIZ", "estado": "pendiente"},
            ]),
            _al("M26 C1-03", "ARIAS", "NICOLAS", []),  # sin deudas → NO aparece
        ],
    },
]


def _todos_los_valores(ws):
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


def test_excel_avance_una_hoja_por_materia():
    wb = load_workbook(io.BytesIO(construir_excel_avance(MATERIAS_AVANCE, agrupar="unidad")))
    assert wb.sheetnames == ["Programación 1"]
    assert wb["Programación 1"]["A1"].value == "Programación 1"


def test_excel_avance_resumen_por_unidad_cuenta_deudores():
    wb = load_workbook(io.BytesIO(construir_excel_avance(MATERIAS_AVANCE, agrupar="unidad")))
    ws = wb["Programación 1"]
    assert ws["A4"].value == "Unidad"
    assert ws["B4"].value == "Alumnos que deben"
    resumen = {}
    for r in range(5, 9):
        k = ws.cell(r, 1).value
        if k and str(k).startswith("Unidad"):
            resumen[k] = ws.cell(r, 2).value
    assert resumen.get("Unidad 3") == 2  # Ferreira + Reynoso
    assert resumen.get("Unidad 4") == 1  # solo Ferreira


def test_excel_avance_listado_bloques_por_unidad_y_pendiente_de_esa_unidad():
    wb = load_workbook(io.BytesIO(construir_excel_avance(MATERIAS_AVANCE, agrupar="unidad")))
    vals = _todos_los_valores(wb["Programación 1"])
    # Encabezado de bloque por unidad con conteo.
    assert any("Unidad 3" in v and "2 alumnos" in v for v in vals)
    assert any("Unidad 4" in v and "1 alumno" in v for v in vals)
    # Header del listado y columnas.
    assert any("LISTADO DE ALUMNOS POR UNIDAD" in v for v in vals)
    assert "Comisión" in vals and "Alumno" in vals and "Pendiente" in vals
    # En el bloque Unidad 3, Ferreira muestra solo lo de la U3 (TP, Autoevaluación), NO el TP de U4.
    assert any(v == "TP, Autoevaluación" for v in vals)
    assert any(v == "TP (desaprobado)" for v in vals)  # bloque Unidad 4


def test_excel_avance_alumno_sin_deudas_no_aparece():
    wb = load_workbook(io.BytesIO(construir_excel_avance(MATERIAS_AVANCE, agrupar="unidad")))
    vals = _todos_los_valores(wb["Programación 1"])
    assert not any("ARIAS" in v for v in vals)


def test_excel_avance_respeta_etiqueta_semana():
    materias = [{"materia": "PYE", "etiqueta": "Semana", "alumnos": [
        _al("A25 C2-01", "ALBARRACIN", "TAMARA", [{"unidad": 1, "tipo": "QUIZ", "estado": "pendiente"}]),
    ]}]
    wb = load_workbook(io.BytesIO(construir_excel_avance(materias, agrupar="unidad")))
    ws = wb["PYE"]
    assert ws["A4"].value == "Semana"
    vals = _todos_los_valores(ws)
    assert any("Semana 1" in v for v in vals)
    assert any("LISTADO DE ALUMNOS POR SEMANA" in v for v in vals)


def test_excel_avance_modo_comision_bloque_por_comision():
    wb = load_workbook(io.BytesIO(construir_excel_avance(MATERIAS_AVANCE, agrupar="comision")))
    ws = wb["Programación 1"]
    assert ws["A4"].value == "Comisión"
    vals = _todos_los_valores(ws)
    # Un bloque por comisión.
    assert any("M26 C1-01" in v and "1 alumno" in v for v in vals)
    # En modo comisión el Pendiente del alumno trae TODAS sus unidades juntas.
    assert any("Unidad 3" in v and "Unidad 4" in v for v in vals)


def test_excel_avance_modo_ambos_dos_hojas_por_materia():
    # Académico: 1 materia → 2 hojas (por unidad + por comisión), pestañas prefijadas.
    wb = load_workbook(io.BytesIO(construir_excel_avance(MATERIAS_AVANCE, agrupar="ambos")))
    assert len(wb.sheetnames) == 2
    por_u = next((wb[n] for n in wb.sheetnames if n.startswith("Por unidad")), None)
    por_c = next((wb[n] for n in wb.sheetnames if n.startswith("Por comisión")), None)
    assert por_u is not None and por_c is not None
    # La hoja por unidad tiene el header de resumen "Unidad"; la de comisión, "Comisión".
    assert por_u["A4"].value == "Unidad"
    assert por_c["A4"].value == "Comisión"


def test_excel_avance_bloque_examenes_y_tabla_estado():
    materias = [{"materia": "Prog 1", "etiqueta": "Unidad", "alumnos": [
        _al("M26 C1-01", "GOMEZ", "Ana", [{"unidad": 3, "tipo": "TP", "estado": "no entregado"}]),
        # Solo debe un parcial (al día en actividades): entra al bloque Exámenes igual.
        _al("M26 C1-02", "DIAZ", "Beto", [], estado="AL_DIA", examenes=["Parcial 1"]),
        # Al día y sin nada que deber: cuenta en la torta, NO en el listado.
        _al("M26 C1-03", "RUIZ", "Eva", [], estado="AL_DIA"),
    ]}]
    wb = load_workbook(io.BytesIO(construir_excel_avance(materias, agrupar="unidad")))
    vals = _todos_los_valores(wb["Prog 1"])
    assert any("Exámenes" in v for v in vals)        # bloque/fila de exámenes
    assert any("Parcial 1" in v for v in vals)
    assert any("DIAZ" in v for v in vals)            # aparece por el parcial
    assert "Estado" in vals                          # tabla por estado (alimenta la torta)
    assert not any("RUIZ" in v for v in vals)        # al día sin deudas → no en el listado


def test_excel_avance_academico_por_comision_columna_examenes():
    materias = [{"materia": "Prog 1", "etiqueta": "Unidad", "alumnos": [
        _al("M26 C1-02", "DIAZ", "Beto", [], estado="RIESGO_ALTO", examenes=["Parcial 1", "Parcial 2"]),
    ]}]
    wb = load_workbook(io.BytesIO(construir_excel_avance(materias, agrupar="comision")))
    vals = _todos_los_valores(wb["Prog 1"])
    assert "Exámenes" in vals                         # header de columna
    assert any("Parcial 1, Parcial 2" in v for v in vals)


def test_excel_avance_corte_en_titulo():
    materias = [{"materia": "Prog 1", "etiqueta": "Unidad", "unidad_actual": 8, "corte_examen": "Parcial 2",
                 "alumnos": [_al("C1", "G", "A", [{"unidad": 3, "tipo": "TP", "estado": "pendiente"}])]}]
    wb = load_workbook(io.BytesIO(construir_excel_avance(materias, agrupar="unidad")))
    ws = wb["Prog 1"]
    assert "Tareas requeridas: Unidad 8, Parcial 2" in ws["A1"].value  # corte al lado del nombre
    assert ws["A2"].value.startswith("Documento generado el")        # fecha como metadata


def test_excel_avance_sin_materias_no_rompe():
    wb = load_workbook(io.BytesIO(construir_excel_avance([], agrupar="unidad")))
    assert len(wb.sheetnames) == 1


def test_examen_para_atender_incluye_desaprobado_y_ausente():
    assert examen_para_atender(None) is False
    assert examen_para_atender({"examenes": [{"resultado": "aprobado"}]}) is False
    assert examen_para_atender({"examenes": [{"resultado": "ausente"}]}) is True
    assert examen_para_atender({"examenes": [{"resultado": "desaprobado"}]}) is True


def test_label_resultado_examen():
    assert label_resultado_examen(None) == "—"
    assert label_resultado_examen({"resultado": "aprobado"}) == "Aprobó"
    assert label_resultado_examen({"resultado": "desaprobado"}) == "Desaprobó"
    assert label_resultado_examen({"resultado": "ausente"}) == "Ausente"
    # rescatado → aprobado con marca (R).
    assert label_resultado_examen({"resultado": "aprobado", "rescatado": True}) == "Aprobó (R)"


# ===================== tiene_examen_desaprobado (disparo) =====================


def test_tiene_examen_desaprobado_true_solo_con_desaprobado():
    assert tiene_examen_desaprobado(None) is False
    assert tiene_examen_desaprobado({"examenes": []}) is False
    # Ausente NO dispara (antes del examen todos están ausentes).
    assert tiene_examen_desaprobado(
        {"examenes": [{"resultado": "ausente"}, {"resultado": "aprobado"}]}
    ) is False
    assert tiene_examen_desaprobado(
        {"examenes": [{"resultado": "desaprobado"}]}
    ) is True


# ===================== formatear_examenes =====================


def test_formatear_examenes_none_devuelve_vacio():
    assert formatear_examenes(None) == ""
    assert formatear_examenes({"examenes": []}) == ""


def test_formatear_examenes_aprobado_y_ausente():
    r = {"examenes": [
        {"etiqueta": "Parcial 1", "resultado": "aprobado", "rescatado": False},
        {"etiqueta": "Parcial 2", "resultado": "ausente", "rescatado": False},
    ]}
    assert formatear_examenes(r) == "Parcial 1: Aprobó · Parcial 2: Ausente"


def test_formatear_examenes_rescatado_marca_recuperado():
    r = {"examenes": [
        {"etiqueta": "Parcial 2", "resultado": "aprobado", "rescatado": True},
    ]}
    assert formatear_examenes(r) == "Parcial 2: Aprobó (recuperado)"


def test_formatear_examenes_desaprobado():
    r = {"examenes": [{"etiqueta": "Global 1", "resultado": "desaprobado", "rescatado": False}]}
    assert formatear_examenes(r) == "Global 1: Desaprobó"


# ===================== formatear_faltantes =====================


def test_formatear_none_devuelve_vacio():
    assert formatear_faltantes(None) == ""
    assert formatear_faltantes({"deudas": []}) == ""


def test_formatear_tp_desaprobado():
    f = {"deudas": [{"unidad": 3, "tipo": "TP", "estado": "desaprobado"}]}
    assert formatear_faltantes(f) == "Unidad 3: TP (desaprobado)"


def test_formatear_tp_no_entregado():
    f = {"deudas": [{"unidad": 3, "tipo": "TP", "estado": "no entregado"}]}
    assert formatear_faltantes(f) == "Unidad 3: TP"


def test_formatear_con_etiqueta_semana():
    # Materias por semana (ej. PYE) → "Semana" en vez de "Unidad".
    f = {"deudas": [
        {"unidad": 8, "tipo": "QUIZ", "estado": "pendiente"},
        {"unidad": 8, "tipo": "TP", "estado": "desaprobado"},
    ]}
    assert formatear_faltantes(f, "Semana") == "Semana 8: Quiz, TP (desaprobado)"


def test_formatear_agrupa_por_unidad():
    f = {"deudas": [
        {"unidad": 3, "tipo": "TP", "estado": "desaprobado"},
        {"unidad": 3, "tipo": "AUTOEVALUACION", "estado": "pendiente"},
        {"unidad": 5, "tipo": "CIERRE", "estado": "pendiente"},
    ]}
    assert formatear_faltantes(f) == "Unidad 3: TP (desaprobado), Autoevaluación · Unidad 5: Cierre"


def test_formatear_quiz_y_quiz_desaprobado():
    f = {"deudas": [
        {"unidad": 2, "tipo": "QUIZ", "estado": "pendiente"},
        {"unidad": 2, "tipo": "QUIZ", "estado": "desaprobado"},
    ]}
    assert formatear_faltantes(f) == "Unidad 2: Quiz, Quiz (desaprobado)"


def test_excel_avance_nombres_de_hoja_se_sanitizan_a_31_chars():
    materias = [{"materia": "X" * 50, "etiqueta": "Unidad", "alumnos": [
        _al("C1", "G", "A", [{"unidad": 1, "tipo": "TP", "estado": "pendiente"}]),
    ]}]
    wb = load_workbook(io.BytesIO(construir_excel_avance(materias, agrupar="unidad")))
    assert all(len(name) <= 31 for name in wb.sheetnames)


# ===================== construir_html_alumno =====================

FILAS_ALUMNO = [
    {"comision": "M26 C1-09", "materia": "Programación 1", "actividad": "Falta Unidad 4 y 5"},
    {"comision": "M26 C1-01", "materia": "Organización Empresarial", "actividad": "Actividad de cierre unidad 8"},
]


def test_html_alumno_incluye_nombre_y_headers():
    html = construir_html_alumno("Ana Gómez", FILAS_ALUMNO)
    assert "Ana Gómez" in html
    assert "Comisión" in html and "Materia" in html and "Actividad" in html
    assert "<table" in html


def test_html_alumno_una_fila_de_datos_por_materia():
    html = construir_html_alumno("Ana", FILAS_ALUMNO)
    # 1 <tr> de header + 1 por fila de datos
    assert html.count("<tr") == len(FILAS_ALUMNO) + 1


def test_html_alumno_incluye_contenido_de_cada_fila():
    html = construir_html_alumno("Ana", FILAS_ALUMNO)
    assert "Programación 1" in html
    assert "Falta Unidad 4 y 5" in html
    assert "Actividad de cierre unidad 8" in html


def test_html_alumno_escapa_caracteres_especiales():
    html = construir_html_alumno("Ana & <script>", [
        {"comision": "C1", "materia": "M & <b>", "actividad": "x"},
    ])
    assert "&amp;" in html
    assert "&lt;script&gt;" in html
    assert "<script>" not in html  # no debe inyectarse markup crudo


def test_html_alumno_usa_estilos_inline():
    # Los clientes de correo ignoran <style>/CSS externo → todo debe ir inline.
    html = construir_html_alumno("Ana", FILAS_ALUMNO)
    assert "style=" in html
    assert "<style" not in html


def test_html_alumno_sin_filas_mensaje_al_dia():
    html = construir_html_alumno("Ana", [])
    assert "<table" not in html  # sin tabla
    assert "Ana" in html


def test_html_alumno_incluye_corte_y_fecha_creacion():
    html = construir_html_alumno("Ana", [
        {"comision": "C1", "materia": "Prog 1", "actividad": "Falta U4",
         "examenes": "—", "corte": "Unidad 8"},
    ])
    assert "Hasta" in html  # header de la columna de corte
    assert "Unidad 8" in html  # valor del corte
    assert "Documento generado el" in html  # fecha de creación


def test_html_alumno_incluye_columna_examenes():
    html = construir_html_alumno("Ana", [
        {"comision": "C1", "materia": "Prog 1", "actividad": "Falta U4",
         "examenes": "Parcial 1: Desaprobó"},
    ])
    assert "Exámenes" in html  # header de la columna
    assert "Parcial 1: Desaprobó" in html  # valor de la fila
