"""
Tests unitarios, de propiedad e integración adicionales (bugfix
comision-orden-numerico-fix, tarea 4).

Complementa la cobertura ya existente en:
  - `test_orden_natural_utils.py` (tarea 3.1) — `natural_key` / `orden_natural_sql`.
  - `test_orden_natural_comision.py` (tareas 1 / 3.6) — Bug Condition por superficie.
  - `test_preservacion_orden_comision.py` (tareas 2 / 3.7) — Properties 2-5.

Este archivo NO duplica lo ya cubierto: cierra los huecos puntuales señalados en el
checklist de la tarea 4 (ver `.kiro/specs/comision-orden-numerico-fix/tasks.md`):
  1. `orden_natural_sql`: los TRES términos (no sólo el sufijo) aplican `NULLS LAST`
     explícito, según el ajuste de la tarea 3.7.
  2. `_agrupar_por_comision`: caso de una sola comisión (sin bucket "Sin comisión
     asignada" cuando no hace falta).
  3. PBT Property 4 (estabilidad/determinismo con duplicados de `(prefijo, n)`).
  4. PBT Property 5 (mezcla aleatoria de `None` + nombres → `None` al final).
  5. Integración Superficie 2: `get_all` con `COMI-1..COMI-12` en DOS años,
     paginación y filtro por año; `get_by_materia` / `get_by_materia_con_tutores` /
     `get_by_tutor` en orden natural.
  6. Integración Superficie 3: `get_alumnos_de_run` con el rango completo
     `M26 C1-01..C1-10` (+ nulos).
  7. Integración Superficie 4: `get_alumnos_por_estado` (no sólo
     `get_alumnos_de_snapshot`) con `COMI-2..COMI-11` (+ nulos).
  8. Integración Superficie 1: camino completo `generar_excel_cierre` → `.xlsx` →
     lectura del archivo, con comisiones de distinta longitud de sufijo.

**Validates: Requirements 2.1, 2.2, 2.3, 2.4, 3.1, 3.2, 3.3, 3.4, 3.5, 3.6, 3.7**
"""

import re
from io import BytesIO

import pytest
import pytest_asyncio
from hypothesis import given, strategies as st
from openpyxl import load_workbook
from sqlalchemy import ARRAY, Column, String, event
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

from app.models.actividad import Actividad
from app.models.avance import AvanceAlumno, AvanceSnapshot
from app.models.base import Base
from app.models.cierre_cursada import CierreCursadaAlumno, CierreCursadaRun
from app.models.cohorte import Cohorte, Cuatrimestre
from app.models.comision import Comision, ComisionTutor
from app.models.componente_unidad import ComponenteUnidad
from app.models.correccion import Correccion
from app.models.entrega import Entrega
from app.models.enums import EstadoAvanceEnum, EstadoCierreEnum
from app.models.examen_materia import ExamenMateria
from app.models.materia import CoordinadorMateria, Materia
from app.models.rubrica import Rubrica
from app.models.unidad import Unidad
from app.models.usuario import Usuario
from app.repositories.avance_repository import AvanceRepository
from app.repositories.cierre_cursada_repository import CierreCursadaRepository
from app.repositories.comision_repository import ComisionRepository
from app.services.excel_cierre_cursada import _agrupar_por_comision, generar_excel_cierre
from app.utils.orden_natural import natural_key, orden_natural_sql

UNIV_ID = 1  # multi-tenant-scoping-queries: universidad_id ahora NOT NULL



# SQLite no tiene JSONB nativo (mismo patrón que los otros archivos de esta suite).
@compiles(JSONB, "sqlite")
def _jsonb_como_json_en_sqlite(element, compiler, **kw):
    return "JSON"


@compiles(ARRAY, "sqlite")
def _array_como_json_en_sqlite(element, compiler, **kw):
    return "JSON"


def _regexp_replace_sqlite(source: str | None, pattern: str, replacement: str) -> str | None:
    """UDF que replica `regexp_replace` de PostgreSQL para que `orden_natural_sql`
    pueda ejecutarse contra SQLite en memoria (mismo shim que los demás archivos de
    esta suite; ver `test_orden_natural_comision.py` para el detalle completo)."""
    if source is None:
        return None
    return re.sub(pattern, replacement, source, count=1)


@pytest_asyncio.fixture
async def db_session():
    """Sesión SQLite en memoria con las tablas de las superficies 2, 3 y 4."""
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False, future=True)

    @event.listens_for(engine.sync_engine, "connect")
    def _registrar_regexp_replace(dbapi_connection, connection_record):
        dbapi_connection.create_function("regexp_replace", 3, _regexp_replace_sqlite)

    tablas = [
        Usuario.__table__,
        Cohorte.__table__,
        Cuatrimestre.__table__,
        Materia.__table__,
        CoordinadorMateria.__table__,
        Unidad.__table__,
        ComponenteUnidad.__table__,
        ExamenMateria.__table__,
        Rubrica.__table__,
        Comision.__table__,
        ComisionTutor.__table__,
        Entrega.__table__,
        Correccion.__table__,
        Actividad.__table__,
        CierreCursadaRun.__table__,
        CierreCursadaAlumno.__table__,
        AvanceSnapshot.__table__,
        AvanceAlumno.__table__,
    ]
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all, tables=tablas)

    session_maker = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with session_maker() as session:
        yield session

    await engine.dispose()


async def _crear_materia(db_session, *, codigo: str = "P1", nombre: str = "Programación 1") -> Materia:
    materia = Materia(universidad_id=UNIV_ID, nombre=nombre, codigo=codigo)
    db_session.add(materia)
    await db_session.flush()
    return materia


async def _crear_usuario(db_session, *, username: str, rol: str = "ADMIN") -> Usuario:
    # `rol` ya no se persiste en `Usuario` (Fase 6 multi-tenant): parámetro
    # conservado por compatibilidad de firma, no leído en este archivo.
    usuario = Usuario(username=username, password_hash="x", nombre=username.capitalize())
    db_session.add(usuario)
    await db_session.flush()
    return usuario


async def _crear_cuatrimestre(db_session) -> Cuatrimestre:
    cohorte = Cohorte(codigo="M26", nombre="M26")
    db_session.add(cohorte)
    await db_session.flush()
    cuatrimestre = Cuatrimestre(nombre="1C", cohorte_id=cohorte.id, numero=1)
    db_session.add(cuatrimestre)
    await db_session.flush()
    return cuatrimestre


def _alumno_cierre(comision_nombre: str | None, apellido: str = "Ap", nombre: str = "No") -> CierreCursadaAlumno:
    return CierreCursadaAlumno(
        moodle_user_id=1, nombre=nombre, apellido=apellido, email=f"{nombre}@x.com",
        comision_id=None, comision_nombre=comision_nombre, tutor_nombre=None,
        estado=EstadoCierreEnum.PROMOCIONA,
    )


def _alumno_avance(comision: str | None, apellido: str = "Ap", nombre: str = "No") -> AvanceAlumno:
    return AvanceAlumno(
        moodle_user_id=1, nombre=nombre, apellido=apellido, email=f"{nombre}@x.com",
        comision=comision, estado=EstadoAvanceEnum.AL_DIA,
    )


# ============================================================================
# 1. Unit — orden_natural_sql: los TRES términos aplican NULLS LAST explícito
# ============================================================================

_col = Column("nombre", String)


def test_orden_natural_sql_los_tres_terminos_tienen_nulls_last_explicito():
    """Tras el ajuste de la tarea 3.7, los TRES términos (prefijo, sufijo,
    desempate por columna completa) aplican `.nullslast()` explícito -- no sólo el
    sufijo -- para que "nulos al final" (Property 5) sea parte del contrato de la
    función, sin depender del comportamiento por defecto del motor de base de
    datos. **Validates: Requirements 3.6**"""
    expresiones = orden_natural_sql(_col)
    assert len(expresiones) == 3
    for expr in expresiones:
        compilado = str(expr.compile(compile_kwargs={"literal_binds": True}))
        assert "nulls last" in compilado.lower()


# ============================================================================
# 2. Unit — _agrupar_por_comision: una sola comisión
# ============================================================================


def test_agrupar_por_comision_una_sola_comision_no_agrega_bucket_sin_asignar():
    """Con una única comisión real (todos los alumnos la tienen asignada), el
    resultado tiene exactamente un bloque y NO se agrega el bucket "Sin comisión
    asignada" (sólo aparece cuando hay al menos un alumno sin comisión).
    **Validates: Requirements 2.2, 3.4**"""
    alumnos = [
        _alumno_cierre("COMI-5", "Bb", "Aa"),
        _alumno_cierre("COMI-5", "Dd", "Cc"),
    ]

    resultado = _agrupar_por_comision(alumnos)

    assert list(resultado.keys()) == ["COMI-5"]
    assert len(resultado["COMI-5"]) == 2


# ============================================================================
# 3. PBT — Property 4: duplicados / mismos (prefijo, n) → orden determinista y estable
# ============================================================================


@given(
    prefijo=st.sampled_from(["COMI", "A", "M26 C1"]),
    numeros=st.lists(st.integers(min_value=0, max_value=50), min_size=3, max_size=10),
)
def test_prop4_natural_key_orden_determinista_y_estable_con_duplicados(prefijo, numeros):
    """Para toda lista de nombres `PREFIJO-<n>` CON duplicados (mismos números
    repetidos), `sorted(key=natural_key)` es determinista (aplicarlo dos veces da
    el mismo resultado) y estable: entre elementos con clave igual, el orden
    relativo de las apariciones originales se conserva.
    **Validates: Requirements 3.2**"""
    nombres = [f"{prefijo}-{n}" for n in numeros]

    ordenado_1 = sorted(nombres, key=natural_key)
    ordenado_2 = sorted(nombres, key=natural_key)
    assert ordenado_1 == ordenado_2

    from itertools import groupby

    con_indices = sorted(enumerate(nombres), key=lambda par: natural_key(par[1]))
    for _, grupo in groupby(con_indices, key=lambda par: natural_key(par[1])):
        indices_del_grupo = [idx for idx, _ in grupo]
        assert indices_del_grupo == sorted(indices_del_grupo)


# ============================================================================
# 4. PBT — Property 5: mezcla aleatoria de None + nombres → None al final
# ============================================================================


@given(
    nombres=st.lists(
        st.one_of(
            st.none(),
            st.sampled_from(["COMI-1", "COMI-2", "COMI-10", "Teórica", "Práctica"]),
        ),
        min_size=1,
        max_size=10,
    )
)
def test_prop5_pbt_none_siempre_al_final_de_natural_key(nombres):
    """Para cualquier lista que mezcle `None` con nombres (con o sin sufijo
    numérico), `sorted(key=natural_key)` ubica TODOS los `None` estrictamente al
    final, sin importar cuántos haya ni en qué posición de llegada estén.
    **Validates: Requirements 3.6**"""
    ordenado = sorted(nombres, key=natural_key)

    primer_none = next((i for i, n in enumerate(ordenado) if n is None), len(ordenado))
    assert all(n is None for n in ordenado[primer_none:])
    assert all(n is not None for n in ordenado[:primer_none])


# ============================================================================
# 5. Integración — Superficie 2: get_all COMI-1..COMI-12 (dos años) + paginación +
#    filtro; get_by_materia / get_by_materia_con_tutores / get_by_tutor
# ============================================================================


@pytest.mark.asyncio
async def test_integracion_superficie2_get_all_comi1_a_comi12_dos_anios_paginado(db_session):
    """`get_all` con `COMI-1..COMI-12` sembradas en DOS años (2025 y 2026) ->
    `anio desc` primario + orden natural dentro de cada año; paginación
    (`per_page`) coherente (páginas concatenadas == orden sin paginar) y `total`
    correcto (24); filtro por año devuelve sólo esas 12 en orden natural.
    **Validates: Requirements 2.1, 3.3, 3.7**"""
    materia = await _crear_materia(db_session)
    db_session.add_all([
        Comision(universidad_id=UNIV_ID, materia_id=materia.id, nombre=f"COMI-{i}", anio=anio)
        for anio in (2025, 2026)
        for i in range(1, 13)
    ])
    await db_session.commit()

    repo = ComisionRepository(db_session)

    todas, total = await repo.get_all(materia_id=materia.id, per_page=100)
    assert total == 24
    esperado_natural = [f"COMI-{i}" for i in range(1, 13)]
    assert [c.nombre for c in todas if c.anio == 2026] == esperado_natural
    assert [c.nombre for c in todas if c.anio == 2025] == esperado_natural
    assert [c.anio for c in todas] == [2026] * 12 + [2025] * 12

    paginas: list[Comision] = []
    for page in range(1, 6):
        pagina, total_pagina = await repo.get_all(materia_id=materia.id, page=page, per_page=5)
        assert total_pagina == 24
        paginas.extend(pagina)
    assert [c.nombre for c in paginas] == [c.nombre for c in todas]
    assert [c.anio for c in paginas] == [c.anio for c in todas]

    filtradas, total_filtrado = await repo.get_all(materia_id=materia.id, anio=2025, per_page=100)
    assert total_filtrado == 12
    assert [c.nombre for c in filtradas] == esperado_natural


@pytest.mark.asyncio
async def test_integracion_superficie2_get_by_materia_con_tutores_y_tutor_orden_natural(db_session):
    """`get_by_materia`, `get_by_materia_con_tutores` y `get_by_tutor` -- no sólo
    `get_all` -- devuelven las comisiones en orden numérico natural.
    **Validates: Requirements 2.1, 3.3**"""
    materia = await _crear_materia(db_session)
    tutor = await _crear_usuario(db_session, username="tutorX", rol="TUTOR")
    comisiones = [
        Comision(universidad_id=UNIV_ID, materia_id=materia.id, nombre=nombre, anio=2026)
        for nombre in ["COMI-10", "COMI-1", "COMI-2"]
    ]
    db_session.add_all(comisiones)
    await db_session.flush()
    db_session.add_all([ComisionTutor(comision_id=c.id, tutor_id=tutor.id) for c in comisiones])
    await db_session.commit()

    repo = ComisionRepository(db_session)
    esperado = ["COMI-1", "COMI-2", "COMI-10"]

    por_materia = await repo.get_by_materia(materia.id)
    assert [c.nombre for c in por_materia] == esperado

    con_tutores = await repo.get_by_materia_con_tutores(materia.id)
    assert [c.nombre for c in con_tutores] == esperado

    por_tutor = await repo.get_by_tutor(tutor.id)
    assert [c.nombre for c in por_tutor] == esperado


# ============================================================================
# 6. Integración — Superficie 3: get_alumnos_de_run con M26 C1-01..C1-10 (+ nulos)
# ============================================================================


@pytest.mark.asyncio
async def test_integracion_superficie3_get_alumnos_de_run_c1_01_a_c1_10_con_nulos(db_session):
    """`get_alumnos_de_run` con alumnos en el rango completo `M26 C1-01..C1-10`
    (llegada en orden inverso) + dos alumnos sin comisión asignada -> orden
    natural por comisión, nulos al final, y entre los nulos, desempate por
    `(apellido, nombre)`. **Validates: Requirements 2.3, 3.5, 3.6**"""
    materia = await _crear_materia(db_session)
    cuatrimestre = await _crear_cuatrimestre(db_session)
    usuario = await _crear_usuario(db_session, username="admin_s3")

    nombres_comision = [f"M26 C1-{i:02d}" for i in range(1, 11)]
    alumnos = [
        _alumno_cierre(nombre, f"Ap{i}", f"No{i}")
        for i, nombre in enumerate(reversed(nombres_comision))
    ]
    alumnos += [
        _alumno_cierre(None, "Zzz", "Sin1"),
        _alumno_cierre(None, "Aaa", "Sin2"),
    ]

    run = CierreCursadaRun(universidad_id=UNIV_ID, 
        materia_id=materia.id, cuatrimestre_id=cuatrimestre.id,
        generado_por_id=usuario.id, total_alumnos=len(alumnos),
    )
    run.alumnos = alumnos
    db_session.add(run)
    await db_session.commit()

    repo = CierreCursadaRepository(db_session)
    resultado = await repo.get_alumnos_de_run(run.id)

    nombres_obtenidos = [a.comision_nombre for a in resultado]
    assert nombres_obtenidos == nombres_comision + [None, None]

    sin_comision = resultado[-2:]
    assert [(a.apellido, a.nombre) for a in sin_comision] == [("Aaa", "Sin2"), ("Zzz", "Sin1")]


# ============================================================================
# 7. Integración — Superficie 4: get_alumnos_por_estado con COMI-2..COMI-11 (+ nulos)
# ============================================================================


@pytest.mark.asyncio
async def test_integracion_superficie4_get_alumnos_por_estado_comi2_a_comi11_con_nulos(db_session):
    """`get_alumnos_por_estado` (no sólo `get_alumnos_de_snapshot`) con comisiones
    `COMI-2..COMI-11` (llegada en orden inverso) + un alumno sin comisión ->
    orden natural, nulos al final. **Validates: Requirements 2.4, 3.6**"""
    materia = await _crear_materia(db_session)
    nombres_comision = [f"COMI-{i}" for i in range(2, 12)]
    alumnos = [
        _alumno_avance(nombre, f"Ap{i}", f"No{i}")
        for i, nombre in enumerate(reversed(nombres_comision))
    ]
    alumnos.append(_alumno_avance(None, "Zzz", "SinX"))

    snapshot = AvanceSnapshot(universidad_id=UNIV_ID, materia_id=materia.id, total_alumnos=len(alumnos))
    snapshot.alumnos = alumnos
    db_session.add(snapshot)
    await db_session.commit()

    repo = AvanceRepository(db_session)
    resultado = await repo.get_alumnos_por_estado([snapshot.id], EstadoAvanceEnum.AL_DIA)

    observado = [a.comision for a in resultado]
    assert observado == nombres_comision + [None]


# ============================================================================
# 8. Integración — Superficie 1: camino completo generar_excel_cierre -> .xlsx ->
#    lectura del archivo, con comisiones de distinta longitud de sufijo
# ============================================================================


def test_integracion_superficie1_xlsx_bloques_orden_natural_multiples_sufijos():
    """Camino COMPLETO `generar_excel_cierre` -> bytes `.xlsx` -> `load_workbook`
    (no se llama a `_agrupar_por_comision` directamente), con comisiones de
    DISTINTA longitud de sufijo (`COMI-1, COMI-2, COMI-10, COMI-20`) + alumnos sin
    comisión asignada. Los títulos de bloque deben aparecer, en el archivo
    generado, en orden numérico natural, con "Sin comisión asignada" al final.
    **Validates: Requirements 2.2, 3.4**"""
    run = CierreCursadaRun(universidad_id=UNIV_ID, 
        id=1, materia_id=1, cuatrimestre_id=1, examenes_snapshot=[],
        generado_por_id=1, total_alumnos=5,
        total_promociona=5, total_regulariza=0, total_recursa=0,
    )
    run.alumnos = [
        _alumno_cierre("COMI-10", "Bb", "Aa"),
        _alumno_cierre(None, "Cc", "Dd"),
        _alumno_cierre("COMI-1", "Ee", "Ff"),
        _alumno_cierre("COMI-20", "Gg", "Hh"),
        _alumno_cierre("COMI-2", "Ii", "Jj"),
    ]

    data, _ = generar_excel_cierre("Programación 1", run)
    ws = load_workbook(BytesIO(data)).active

    # Las barras de bloque mergean desde la columna C (3); las columnas A y B de esa
    # fila quedan vacías (a diferencia de las filas de encabezado y de datos).
    titulos_bloque = [
        row[2].value
        for row in ws.iter_rows(min_row=6)
        if row[0].value is None and row[1].value is None and row[2].value is not None
    ]

    assert titulos_bloque == ["COMI-1", "COMI-2", "COMI-10", "COMI-20", "Sin comisión asignada"]
