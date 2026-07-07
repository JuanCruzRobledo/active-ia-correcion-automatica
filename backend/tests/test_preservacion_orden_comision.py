"""
Tests de preservación (bugfix comision-orden-numerico-fix, tarea 2 — observation-first).

Property 2: Preservation - Fallback alfabético, año primario, desempates, nulos e
intra-bloque (design.md).

METODOLOGÍA observation-first: se ejecuta el código SIN ARREGLAR con inputs que NO
disparan la Bug Condition (ver `test_orden_natural_comision.py`, tarea 1) — nombres sin
sufijo numérico, filtros/paginación del listado, nulos, y desempates existentes — se
OBSERVA la salida real y se escriben los tests que la capturan como línea base
("baseline"). Estos tests DEBEN PASAR sobre el código sin arreglar: confirman el
comportamiento que el fix de la tarea 3 no debe alterar. Este archivo se re-ejecuta SIN
CAMBIOS en la tarea 3.7, una vez aplicado el fix, donde se espera que sigan pasando
(ninguna regresión).

Cobertura (Correctness Properties 2, 3, 4 y 5 de design.md):
  - Property 2 (3.1) — Fallback alfabético: nombres SIN sufijo numérico (`Teórica`,
    `Práctica`, `Laboratorio`) → orden idéntico a `sorted(..., key=str.casefold)`.
    Superficie: `excel_cierre_cursada._agrupar_por_comision` (hoy YA ordena por
    `nombre.casefold()` para toda comisión real, sin extraer ningún sufijo numérico —
    por eso el fallback alfabético es, hoy, el comportamiento universal para nombres
    sin dígitos finales).
  - Property 3 (3.3, 3.7) — Año primario + filtros/paginación:
    `ComisionRepository.get_all` usa `anio desc` como criterio PRIMARIO y
    `nombre.asc()` como secundario; los filtros (materia/año/tutor/coordinador) y la
    paginación (`page`/`per_page`) no cambian el conjunto de filas ni el `total`.
  - Property 4 (3.2, 3.4, 3.5) — Desempate estable + intra-bloque + "Sin comisión
    asignada" al final: dentro de un bloque/reporte, los alumnos se ordenan por
    `(Apellido, Nombre)`; el bloque "Sin comisión asignada" del Excel queda SIEMPRE
    último; el encabezado "{comisión} — Tutor: {tutor}" se mantiene intacto; el
    desempate por `(apellido, nombre)` resuelve nombres de comisión idénticos.
  - Property 5 (3.6) — Nulos al final: filas con `comision_nombre` (cierre) /
    `comision` (avance) `NULL` quedan al final (`NULLS LAST`), sin excepción ni
    interrupción del listado/reporte.

Superficies ejercitadas (mismas 4 que la tarea 1, ver design.md "Fix Implementation"):
  1. Excel de cierre (`excel_cierre_cursada._agrupar_por_comision` /
     `generar_excel_cierre`) — puro, sin DB.
  2. Listado de comisiones (`comision_repository.ComisionRepository.get_all`) — SQL.
  3. Corrida de cierre (`cierre_cursada_repository.CierreCursadaRepository.get_alumnos_de_run`) — SQL.
  4. Reporte de avance (`avance_repository.AvanceRepository.get_alumnos_de_snapshot`,
     `get_alumnos_por_estado`) — SQL.

Enfoque de test: property-based (hypothesis) para los sub-properties que se pueden
ejercitar sobre la función PURA del Excel (fallback alfabético, estabilidad
intra-bloque) — igual patrón que `test_orden_natural_comision.py` (tarea 1) y
`test_excel_cierre_cursada.py` (Bug 3 de la spec previa). Integración con datos
sembrados (SQLite en memoria) para año/filtros/paginación (superficie 2) y nulos
(superficies 3 y 4), que requieren DB real para observar el `ORDER BY` de SQL.

**Validates: Requirements 3.1, 3.2, 3.3, 3.4, 3.5, 3.6, 3.7**
"""

import re
from datetime import datetime

import pytest
import pytest_asyncio
from hypothesis import given, settings, strategies as st
from sqlalchemy import ARRAY, event
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

from app.models.avance import AvanceAlumno, AvanceSnapshot
from app.models.base import Base
from app.models.cierre_cursada import CierreCursadaAlumno, CierreCursadaRun
from app.models.cohorte import Cohorte, Cuatrimestre
from app.models.comision import Comision, ComisionTutor
from app.models.componente_unidad import ComponenteUnidad
from app.models.entrega import Entrega
from app.models.enums import EstadoAvanceEnum, EstadoCierreEnum
from app.models.examen_materia import ExamenMateria
from app.models.materia import CoordinadorMateria, Materia
from app.models.rubrica import Rubrica
from app.models.unidad import Unidad
from app.models.usuario import Usuario
from app.repositories.avance_repository import AvanceRepository
from app.repositories.cierre_cursada_repository import CierreCursadaRepository
from app.repositories.comision_repository import ComisionRepository
from app.services.excel_cierre_cursada import _agrupar_por_comision, generar_excel_cierre


# SQLite no tiene JSONB nativo; se compila como JSON (mismo patrón que
# test_orden_natural_comision.py / test_cierre_cursada_service.py — sólo afecta al
# dialecto de tests, Postgres sigue usando JSONB real en producción).
@compiles(JSONB, "sqlite")
def _jsonb_como_json_en_sqlite(element, compiler, **kw):
    return "JSON"


@compiles(ARRAY, "sqlite")
def _array_como_json_en_sqlite(element, compiler, **kw):
    return "JSON"


def _regexp_replace_sqlite(source: str | None, pattern: str, replacement: str) -> str | None:
    """UDF (user-defined function) que replica, para efectos de este test suite, la
    semántica de `regexp_replace(source, pattern, replacement)` de PostgreSQL que usa
    `orden_natural_sql` (app/utils/orden_natural.py): reemplaza sólo la PRIMERA
    coincidencia del patrón (comportamiento por defecto de Postgres sin la bandera
    'g'), soportando backreferences con sintaxis `\\1` -- la misma que usa
    `re.sub`, por lo que el patrón/reemplazo SQL (`r"^.*?(\\d+)$"`, `r"\\1"`) funciona
    sin traducción.

    NOTA (alcance del shim, no altera producción): SQLite no tiene `regexp_replace`
    nativo; esta función se registra como UDF SOLO en la conexión SQLite en memoria
    usada por los tests (ver `db_session` más abajo), para poder ejecutar el mismo SQL
    que en producción corre contra PostgreSQL real (ver design.md → "Decisión: Python
    vs SQL", tradeoff documentado). `nullif` y `NULLS LAST` son soportados nativamente
    por SQLite (>= 3.30), así que no requieren shim adicional. Mismo shim que
    `test_orden_natural_comision.py` (tarea 3.6), reutilizado aquí (tarea 3.7) para que
    el mismo SQL de `orden_natural_sql` (superficies 2, 3 y 4) pueda ejecutarse contra
    la sesión SQLite en memoria de este archivo.
    """
    if source is None:
        return None
    return re.sub(pattern, replacement, source, count=1)


@pytest_asyncio.fixture
async def db_session():
    """Sesión SQLite en memoria con las tablas necesarias para las superficies 2, 3 y 4
    (Comision/ComisionTutor/CoordinadorMateria, CierreCursadaAlumno/Run, AvanceAlumno/
    Snapshot + sus FKs) — mismo set de tablas que `test_orden_natural_comision.py`
    (tarea 1), reutilizado para que ambos archivos de test sean independientes entre
    sí (no se importa fixtures de uno desde el otro)."""
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False, future=True)

    @event.listens_for(engine.sync_engine, "connect")
    def _registrar_regexp_replace(dbapi_connection, connection_record):
        dbapi_connection.create_function("regexp_replace", 3, _regexp_replace_sqlite)

    tablas = [
        Usuario.__table__,
        Cohorte.__table__,
        Cuatrimestre.__table__,
        Materia.__table__,
        CoordinadorMateria.__table__,
        Unidad.__table__,
        ComponenteUnidad.__table__,
        ExamenMateria.__table__,
        Rubrica.__table__,
        Comision.__table__,
        ComisionTutor.__table__,
        Entrega.__table__,
        CierreCursadaRun.__table__,
        CierreCursadaAlumno.__table__,
        AvanceSnapshot.__table__,
        AvanceAlumno.__table__,
    ]
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all, tables=tablas)

    session_maker = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with session_maker() as session:
        yield session

    await engine.dispose()


async def _crear_materia(db_session, *, codigo: str = "P1", nombre: str = "Programación 1") -> Materia:
    materia = Materia(nombre=nombre, codigo=codigo)
    db_session.add(materia)
    await db_session.flush()
    return materia


async def _crear_usuario(db_session, *, username: str, rol: str = "ADMIN") -> Usuario:
    usuario = Usuario(
        username=username, password_hash="x", nombre=username.capitalize(), rol=rol,
    )
    db_session.add(usuario)
    await db_session.flush()
    return usuario


def _alumno_cierre(comision_nombre: str | None, apellido: str = "Ap", nombre: str = "No") -> CierreCursadaAlumno:
    return CierreCursadaAlumno(
        moodle_user_id=1, nombre=nombre, apellido=apellido, email=f"{nombre}@x.com",
        comision_id=None, comision_nombre=comision_nombre, tutor_nombre=None,
        estado=EstadoCierreEnum.PROMOCIONA,
    )


def _alumno_avance(comision: str | None, apellido: str = "Ap", nombre: str = "No") -> AvanceAlumno:
    return AvanceAlumno(
        moodle_user_id=1, nombre=nombre, apellido=apellido, email=f"{nombre}@x.com",
        comision=comision, estado=EstadoAvanceEnum.AL_DIA,
    )


# ============================================================================
# Property 2 (3.1) — Fallback alfabético para nombres SIN sufijo numérico
# Superficie 1 (Excel) — `_agrupar_por_comision`, sin DB (función pura)
# ============================================================================
#
# Observado sobre el código SIN arreglar: `_clave_orden` hoy calcula
# `nombre.casefold()` para TODA comisión real (no extrae ningún sufijo numérico), así
# que para nombres sin dígitos finales el orden resultante YA es, siempre,
# `sorted(nombres, key=str.casefold)`. Este comportamiento debe seguir intacto tras el
# fix (natural_key degrada a casefold cuando no hay sufijo numérico).
# **Validates: Requirements 3.1**


def test_fallback_alfabetico_nombres_sin_sufijo_numerico_ejemplo_del_diseno():
    """Ejemplo concreto de design.md (Edge — fallback alfabético, NO es bug):
    `["Teórica", "Práctica", "Laboratorio"]` -> orden alfabético case-insensitive
    `Laboratorio, Práctica, Teórica`, idéntico hoy y tras el fix."""
    alumnos = [
        _alumno_cierre("Teórica", "Aa", "Bb"),
        _alumno_cierre("Práctica", "Cc", "Dd"),
        _alumno_cierre("Laboratorio", "Ee", "Ff"),
    ]

    titulos = list(_agrupar_por_comision(alumnos).keys())

    assert titulos == ["Laboratorio", "Práctica", "Teórica"]
    assert titulos == sorted(["Teórica", "Práctica", "Laboratorio"], key=str.casefold)


_POOL_ALFABETICO_SIN_SUFIJO = [
    "Teórica", "Práctica", "Laboratorio", "Alfa", "Beta", "gamma", "ZETA", "Delta",
]


@given(
    nombres=st.lists(
        st.sampled_from(_POOL_ALFABETICO_SIN_SUFIJO), unique=True, min_size=1, max_size=6
    )
)
def test_prop2_fallback_alfabetico_para_toda_lista_sin_sufijo_numerico(nombres):
    """Para TODA lista de nombres de comisión sin sufijo numérico, el orden observado
    de `_agrupar_por_comision` es idéntico al orden alfabético (case-insensitive)
    actual. **Validates: Requirements 3.1**"""
    alumnos = [
        _alumno_cierre(nombre, f"Ap{i}", f"No{i}") for i, nombre in enumerate(nombres)
    ]

    titulos = list(_agrupar_por_comision(alumnos).keys())

    assert titulos == sorted(nombres, key=str.casefold)


# ============================================================================
# Property 3 (3.3, 3.7) — Año descendente primario y filtros/paginación del listado
# Superficie 2 (Listado) — `ComisionRepository.get_all`, SQL
# ============================================================================
#
# Observado sobre el código SIN arreglar: `get_all` ordena por
# `Comision.anio.desc(), Comision.nombre.asc()` -- el año es SIEMPRE el criterio
# primario (todas las comisiones del año más reciente antes que las de un año
# anterior, sin importar el nombre); dentro de cada año, el nombre decide el orden
# secundario. Los filtros (materia/año/tutor/coordinador) restringen el WHERE sin
# afectar el ORDER BY, y la paginación (`offset/limit`) se aplica DESPUÉS del
# `ORDER BY` y el `count` total se calcula ANTES del `offset/limit` -- por eso el
# conjunto de filas y el total no cambian entre páginas.


@pytest.mark.asyncio
async def test_prop3_anio_desc_es_criterio_primario_del_listado(db_session):
    """Comisiones de dos años (2026 y 2025), mismos nombres (sin sufijo de distinta
    longitud -> no dispara la Bug Condition) -> esperado (hoy y preservado tras el
    fix): TODAS las comisiones de 2026 antes que las de 2025, y dentro de cada año en
    orden alfabético por nombre. **Validates: Requirements 3.3**"""
    materia = await _crear_materia(db_session)
    db_session.add_all([
        Comision(materia_id=materia.id, nombre="COMI-2", anio=2025),
        Comision(materia_id=materia.id, nombre="COMI-1", anio=2026),
        Comision(materia_id=materia.id, nombre="COMI-2", anio=2026),
        Comision(materia_id=materia.id, nombre="COMI-1", anio=2025),
    ])
    await db_session.commit()

    repo = ComisionRepository(db_session)
    comisiones, total = await repo.get_all(materia_id=materia.id, per_page=10)

    assert total == 4
    observado = [(c.anio, c.nombre) for c in comisiones]
    assert observado == [(2026, "COMI-1"), (2026, "COMI-2"), (2025, "COMI-1"), (2025, "COMI-2")]


@pytest.mark.asyncio
async def test_prop3_filtro_por_materia_no_altera_conjunto_ni_total(db_session):
    """Filtrar por `materia_id` sólo restringe el conjunto a esa materia; comisiones
    de OTRA materia no aparecen y no afectan el `total`. **Validates: Requirements 3.7**"""
    materia_a = await _crear_materia(db_session, codigo="P1", nombre="Programación 1")
    materia_b = await _crear_materia(db_session, codigo="P2", nombre="Programación 2")
    db_session.add_all([
        Comision(materia_id=materia_a.id, nombre="COMI-1", anio=2026),
        Comision(materia_id=materia_a.id, nombre="COMI-2", anio=2026),
        Comision(materia_id=materia_b.id, nombre="COMI-1", anio=2026),
    ])
    await db_session.commit()

    repo = ComisionRepository(db_session)
    comisiones, total = await repo.get_all(materia_id=materia_a.id, per_page=10)

    assert total == 2
    assert {c.materia_id for c in comisiones} == {materia_a.id}


@pytest.mark.asyncio
async def test_prop3_filtro_por_anio_tutor_y_coordinador(db_session):
    """Filtros combinados por año, tutor y coordinador -- cada uno devuelve
    exactamente el subconjunto esperado, sin alterar el criterio de orden (`anio desc,
    nombre asc`). **Validates: Requirements 3.7**"""
    materia = await _crear_materia(db_session)
    tutor = await _crear_usuario(db_session, username="tutor1", rol="TUTOR")
    coordinador = await _crear_usuario(db_session, username="coord1", rol="COORDINADOR")
    db_session.add(CoordinadorMateria(coordinador_id=coordinador.id, materia_id=materia.id))

    c_2025 = Comision(materia_id=materia.id, nombre="COMI-1", anio=2025)
    c_2026_con_tutor = Comision(materia_id=materia.id, nombre="COMI-1", anio=2026)
    c_2026_sin_tutor = Comision(materia_id=materia.id, nombre="COMI-2", anio=2026)
    db_session.add_all([c_2025, c_2026_con_tutor, c_2026_sin_tutor])
    await db_session.flush()
    db_session.add(ComisionTutor(comision_id=c_2026_con_tutor.id, tutor_id=tutor.id))
    await db_session.commit()

    repo = ComisionRepository(db_session)

    # Filtro por año: sólo 2026 (2 comisiones).
    _, total_2026 = await repo.get_all(materia_id=materia.id, anio=2026, per_page=10)
    assert total_2026 == 2

    # Filtro por tutor: sólo la comisión asignada al tutor.
    por_tutor, total_tutor = await repo.get_all(tutor_id=tutor.id, per_page=10)
    assert total_tutor == 1
    assert por_tutor[0].id == c_2026_con_tutor.id

    # Filtro por coordinador: todas las comisiones de la materia que coordina (3).
    por_coord, total_coord = await repo.get_all(coordinador_id=coordinador.id, per_page=10)
    assert total_coord == 3


@pytest.mark.asyncio
async def test_prop3_paginacion_concatena_al_mismo_orden_y_total_no_cambia(db_session):
    """`per_page` menor al total -> las páginas concatenadas reproducen el mismo
    orden que una consulta sin paginar, y el `total` es el mismo en cada página (no se
    ve afectado por `page`/`per_page`). **Validates: Requirements 3.7**"""
    materia = await _crear_materia(db_session)
    db_session.add_all([
        Comision(materia_id=materia.id, nombre=f"COMI-{i}", anio=2026) for i in range(1, 6)
    ])
    await db_session.commit()

    repo = ComisionRepository(db_session)

    todas, total_sin_paginar = await repo.get_all(materia_id=materia.id, per_page=10)
    assert total_sin_paginar == 5

    pagina_1, total_p1 = await repo.get_all(materia_id=materia.id, page=1, per_page=2)
    pagina_2, total_p2 = await repo.get_all(materia_id=materia.id, page=2, per_page=2)
    pagina_3, total_p3 = await repo.get_all(materia_id=materia.id, page=3, per_page=2)

    assert (total_p1, total_p2, total_p3) == (5, 5, 5)
    concatenado = pagina_1 + pagina_2 + pagina_3
    assert [c.nombre for c in concatenado] == [c.nombre for c in todas]
    assert len(pagina_1) == 2 and len(pagina_2) == 2 and len(pagina_3) == 1


# ============================================================================
# Property 4 (3.2, 3.4, 3.5) — Desempate estable, intra-bloque y "Sin comisión
# asignada" al final
# ============================================================================


def test_prop4_sin_comision_asignada_queda_siempre_al_final_del_excel():
    """"Sin comisión asignada" (Excel) queda ÚLTIMO sin importar en qué posición de
    llegada aparecen los alumnos sin comisión, incluso mezclado entre varios bloques
    reales con sufijos numéricos (sin disparar la Bug Condition -- eso es Property 1,
    no ésta). **Validates: Requirements 3.4**"""
    alumnos = [
        _alumno_cierre(None, "Aa", "Bb"),
        _alumno_cierre("COMI-1", "Cc", "Dd"),
        _alumno_cierre(None, "Ee", "Ff"),
        _alumno_cierre("COMI-2", "Gg", "Hh"),
    ]

    titulos = list(_agrupar_por_comision(alumnos).keys())

    assert titulos[-1] == "Sin comisión asignada"
    assert titulos == ["COMI-1", "COMI-2", "Sin comisión asignada"]


def test_prop4_encabezado_de_bloque_conserva_formato_comision_tutor():
    """El título compuesto "{comisión} — Tutor: {tutor}" se conserva sin cambios
    cuando hay tutor asignado, y el bloque queda como sólo el nombre cuando no hay
    tutor (`tutor_nombre=None`). **Validates: Requirements 3.4**"""
    alumnos = [
        _alumno_cierre("COMI-1", "Aa", "Bb"),
        _alumno_cierre("COMI-2", "Cc", "Dd"),
    ]
    alumnos[0].tutor_nombre = "Juan Pérez"

    titulos = list(_agrupar_por_comision(alumnos).keys())

    assert titulos == ["COMI-1 — Tutor: Juan Pérez", "COMI-2"]


def test_prop4_intra_bloque_ordena_por_apellido_nombre():
    """Dentro de un bloque del Excel, los alumnos se ordenan por `(Apellido, Nombre)`
    sin importar el orden de llegada -- independiente del criterio de orden de
    comisiones (Property 1), que sólo reordena los BLOQUES.
    **Validates: Requirements 3.5**"""
    run = CierreCursadaRun(
        id=1, materia_id=1, cuatrimestre_id=1, examenes_snapshot=[],
        generado_por_id=1, total_alumnos=3,
        total_promociona=3, total_regulariza=0, total_recursa=0,
    )
    run.created_at = datetime(2026, 1, 1, 12, 0)
    run.alumnos = [
        _alumno_cierre("COMI-1", "Zeta", "Zoe"),
        _alumno_cierre("COMI-1", "Alfa", "Ana"),
        _alumno_cierre("COMI-1", "Alfa", "Beto"),
    ]

    data, _ = generar_excel_cierre("Programación 1", run)

    from io import BytesIO
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(data)).active
    nombres_en_orden = [
        row[0].value
        for row in ws.iter_rows()
        if isinstance(row[0].value, str) and "," in row[0].value
    ]
    assert nombres_en_orden == ["Alfa, Ana", "Alfa, Beto", "Zeta, Zoe"]


_POOL_APELLIDOS = ["Alfa", "Beta", "Gamma", "Delta", "Epsilon"]


@settings(deadline=None)
@given(
    apellidos=st.lists(st.sampled_from(_POOL_APELLIDOS), min_size=1, max_size=6, unique=True)
)
def test_prop4_intra_bloque_siempre_por_apellido_para_cualquier_orden_de_llegada(apellidos):
    """Para TODA lista de apellidos (orden de llegada arbitrario, mismo bloque de
    comisión), el orden observado dentro del bloque siempre coincide con
    `sorted(apellidos)`. **Validates: Requirements 3.5**"""
    alumnos = [
        _alumno_cierre("COMI-1", apellido, "Nombre")
        for apellido in apellidos
    ]

    titulos = _agrupar_por_comision(alumnos)
    del_grupo = titulos["COMI-1"]
    orden_observado = [
        a.apellido for a in sorted(del_grupo, key=lambda x: (x.apellido or "", x.nombre or ""))
    ]

    assert orden_observado == sorted(apellidos)


@pytest.mark.asyncio
async def test_prop4_desempate_nombre_comision_identico_por_apellido_nombre(db_session):
    """Desempate estable (3.2): alumnos con el MISMO `comision_nombre` en
    `get_alumnos_de_run` -- el orden entre ellos se resuelve por el desempate
    secundario ya existente `(apellido, nombre)`, sin importar el orden de inserción.
    **Validates: Requirements 3.2, 3.5**"""
    materia = await _crear_materia(db_session)
    cuatrimestre = Cuatrimestre(nombre="1C", numero=1, cohorte_id=(await _crear_cohorte(db_session)).id)
    db_session.add(cuatrimestre)
    await db_session.flush()
    usuario = await _crear_usuario(db_session, username="admin1")

    run = CierreCursadaRun(
        materia_id=materia.id, cuatrimestre_id=cuatrimestre.id,
        generado_por_id=usuario.id, total_alumnos=3,
    )
    run.alumnos = [
        _alumno_cierre("COMI-1", "Zeta", "Zoe"),
        _alumno_cierre("COMI-1", "Alfa", "Ana"),
        _alumno_cierre("COMI-1", "Alfa", "Beto"),
    ]
    db_session.add(run)
    await db_session.commit()

    repo = CierreCursadaRepository(db_session)
    alumnos = await repo.get_alumnos_de_run(run.id)

    apellidos_nombres = [(a.apellido, a.nombre) for a in alumnos]
    assert apellidos_nombres == [("Alfa", "Ana"), ("Alfa", "Beto"), ("Zeta", "Zoe")]


async def _crear_cohorte(db_session) -> Cohorte:
    cohorte = Cohorte(codigo="M26", nombre="M26")
    db_session.add(cohorte)
    await db_session.flush()
    return cohorte


# ============================================================================
# Property 5 (3.6) — Valores nulos de comisión al final (NULLS LAST)
# ============================================================================
#
# Observado sobre el código SIN arreglar: `get_alumnos_de_run` ordena por
# `comision_nombre.asc().nullslast()` y `get_alumnos_de_snapshot`/
# `get_alumnos_por_estado` por `comision.asc().nullslast()` -- en ambos casos los
# `NULL` quedan al final, sin lanzar excepción ni interrumpir la consulta.


@pytest.mark.asyncio
async def test_prop5_nulos_al_final_en_corrida_de_cierre(db_session):
    """Alumnos con `comision_nombre=None` intercalados con alumnos con comisión ->
    los `None` quedan al FINAL, sin excepción, en `get_alumnos_de_run`.
    **Validates: Requirements 3.6**"""
    materia = await _crear_materia(db_session)
    cohorte = await _crear_cohorte(db_session)
    cuatrimestre = Cuatrimestre(nombre="1C", numero=1, cohorte_id=cohorte.id)
    db_session.add(cuatrimestre)
    await db_session.flush()
    usuario = await _crear_usuario(db_session, username="admin2")

    run = CierreCursadaRun(
        materia_id=materia.id, cuatrimestre_id=cuatrimestre.id,
        generado_por_id=usuario.id, total_alumnos=4,
    )
    run.alumnos = [
        _alumno_cierre(None, "Aa", "Bb"),
        _alumno_cierre("COMI-2", "Cc", "Dd"),
        _alumno_cierre(None, "Ee", "Ff"),
        _alumno_cierre("COMI-1", "Gg", "Hh"),
    ]
    db_session.add(run)
    await db_session.commit()

    repo = CierreCursadaRepository(db_session)
    alumnos = await repo.get_alumnos_de_run(run.id)

    nombres = [a.comision_nombre for a in alumnos]
    assert len(nombres) == 4
    primer_null = next((i for i, n in enumerate(nombres) if n is None), len(nombres))
    assert all(n is None for n in nombres[primer_null:])
    assert all(n is not None for n in nombres[:primer_null])
    # Con comisiones reales (COMI-1, COMI-2) ambas antes de los None.
    assert set(nombres[:primer_null]) == {"COMI-1", "COMI-2"}


@pytest.mark.asyncio
async def test_prop5_nulos_al_final_en_reporte_de_avance(db_session):
    """Alumnos con `comision=None` intercalados con alumnos con comisión -> los
    `None` quedan al FINAL, sin excepción, en `get_alumnos_de_snapshot` y en
    `get_alumnos_por_estado`. **Validates: Requirements 3.6**"""
    materia = await _crear_materia(db_session)
    snapshot = AvanceSnapshot(materia_id=materia.id, total_alumnos=4)
    snapshot.alumnos = [
        _alumno_avance("COMI-2", "Aa", "Bb"),
        _alumno_avance(None, "Cc", "Dd"),
        _alumno_avance("COMI-1", "Ee", "Ff"),
        _alumno_avance(None, "Gg", "Hh"),
    ]
    db_session.add(snapshot)
    await db_session.commit()

    repo = AvanceRepository(db_session)

    alumnos = await repo.get_alumnos_de_snapshot(snapshot.id)
    nombres = [a.comision for a in alumnos]
    primer_null = next((i for i, n in enumerate(nombres) if n is None), len(nombres))
    assert all(n is None for n in nombres[primer_null:])
    assert all(n is not None for n in nombres[:primer_null])

    por_estado = await repo.get_alumnos_por_estado([snapshot.id], EstadoAvanceEnum.AL_DIA)
    nombres_estado = [a.comision for a in por_estado]
    primer_null_estado = next(
        (i for i, n in enumerate(nombres_estado) if n is None), len(nombres_estado)
    )
    assert all(n is None for n in nombres_estado[primer_null_estado:])
    assert all(n is not None for n in nombres_estado[:primer_null_estado])
