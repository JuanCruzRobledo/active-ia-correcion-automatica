"""Integration tests para el router de cierre de cursada (bugfix cierre-cursada-fix,
tarea 14).

Nivel 1 (contrato HTTP) — sigue el patrón de `test_moodle_import.py`: no usa DB real
(el conftest SQLite no soporta JSONB/ARRAY/enums PG nativos con `Base.metadata` global),
se mockea `CierreCursadaService`/`MateriaRepository` a nivel de router
(`patch("app.routers.cierre_cursada.X")`), se sobreescriben `get_current_user`/`get_db`, y
se usa `AsyncClient` + `ASGITransport(app=app)`.

Nivel 2 (flujo end-to-end de `CierreCursadaService.generar`) — usa una sesión SQLite en
memoria con SOLO las tablas necesarias (mismo patrón que
`tests/unit/services/test_cierre_cursada_service.py`) y mockea únicamente el
`MoodleService` (I/O externo); todo el resto (repos, `examen_mapper`,
`calcular_estado_cierre`, persistencia append-only) corre con código real.
"""

import io
import re
from datetime import datetime
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
import pytest_asyncio
from fastapi import HTTPException, status
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import ARRAY, event
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

from app.core.dependencies import ContextoUniversidad, get_current_user, get_db, get_universidad_activa
from app.main import app
from app.models.base import Base
from app.models.actividad import Actividad
from app.models.cierre_cursada import CierreCursadaAlumno, CierreCursadaRun
from app.models.cohorte import Cohorte, Cuatrimestre
from app.models.comision import Comision, ComisionTutor
from app.models.componente_unidad import ComponenteUnidad
from app.models.correccion import Correccion
from app.models.entrega import Entrega
from app.models.enums import EstadoCierreEnum, RolEnum
from app.models.examen_materia import ExamenMateria
from app.models.materia import CoordinadorMateria, Materia
from app.models.rubrica import Rubrica
from app.models.unidad import Unidad
from app.models.usuario import Usuario
from app.repositories.cierre_cursada_repository import CierreCursadaRepository
from app.services.cierre_cursada_service import CierreCursadaService
from app.services.moodle_bulk_parser import construir_indice_nombre_cmid
from app.services.moodle_service import MoodleAuthError, MoodleConnectionError


# ===================================================================================
# Nivel 1 — contrato HTTP (mockeando el service completo)
# ===================================================================================


def _admin_user(**attrs):
    base = {
        "id": 1,
        "rol": RolEnum.ADMIN,
        "moodle_username": "coord1",
        "moodle_password_encrypted": "enc",
        "moodle_host": "https://moodle.example.com",
    }
    base.update(attrs)
    user = MagicMock(**base)
    app.dependency_overrides[get_current_user] = lambda: user
    app.dependency_overrides[get_db] = lambda: AsyncMock()
    app.dependency_overrides[get_universidad_activa] = lambda: ContextoUniversidad(
        universidad_id=1, rol=base["rol"], es_superadmin=False
    )
    return user


@pytest.fixture(autouse=True)
def _clear_overrides():
    yield
    app.dependency_overrides.clear()


def _run_response(**over):
    base = dict(
        id=1, materia_id=10, cuatrimestre_id=3, umbral_tp_pct=None,
        generado_por_id=1, total_alumnos=3,
        total_promociona=1, total_regulariza=1, total_recursa=1,
        created_at=datetime(2026, 1, 1, 12, 0),
    )
    base.update(over)
    return SimpleNamespace(**base)


@pytest.mark.asyncio
async def test_generar_exitoso_devuelve_200_con_conteos():
    _admin_user()
    run = _run_response()
    with patch("app.routers.cierre_cursada.CierreCursadaService") as cls:
        cls.return_value.generar = AsyncMock(return_value=run)
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/api/v1/cierre-cursada/materias/10/generar",
                json={"cuatrimestre_id": 3},
            )
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_promociona"] == 1
    assert body["total_regulariza"] == 1
    assert body["total_recursa"] == 1
    assert body["total_alumnos"] == 3
    # El payload enviado al service ya NO incluye umbral_tp_pct/reglas.
    cls.return_value.generar.assert_awaited_once()
    args = cls.return_value.generar.await_args.args
    assert args[0] == 10  # materia_id
    assert args[1] == 3  # cuatrimestre_id


@pytest.mark.asyncio
async def test_generar_materia_sin_examenes_devuelve_400():
    _admin_user()
    with patch("app.routers.cierre_cursada.CierreCursadaService") as cls:
        cls.return_value.generar = AsyncMock(
            side_effect=HTTPException(
                status.HTTP_400_BAD_REQUEST,
                detail="La materia no tiene exámenes configurados en el dashboard",
            )
        )
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/api/v1/cierre-cursada/materias/10/generar",
                json={"cuatrimestre_id": 3},
            )
    assert resp.status_code == 400
    assert "exámenes configurados" in resp.json()["detail"]


@pytest.mark.asyncio
async def test_generar_sin_credenciales_moodle_devuelve_424_sin_llamar_al_service():
    # El router chequea _requerir_credenciales_moodle ANTES de tocar el service.
    _admin_user(moodle_username=None, moodle_password_encrypted=None)
    with patch("app.routers.cierre_cursada.CierreCursadaService") as cls:
        cls.return_value.generar = AsyncMock()
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/api/v1/cierre-cursada/materias/10/generar",
                json={"cuatrimestre_id": 3},
            )
        cls.return_value.generar.assert_not_awaited()
    assert resp.status_code == 424


@pytest.mark.asyncio
async def test_generar_moodle_auth_error_devuelve_424():
    _admin_user()
    with patch("app.routers.cierre_cursada.CierreCursadaService") as cls:
        cls.return_value.generar = AsyncMock(side_effect=MoodleAuthError("Credenciales Moodle incorrectas"))
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/api/v1/cierre-cursada/materias/10/generar",
                json={"cuatrimestre_id": 3},
            )
    assert resp.status_code == 424


@pytest.mark.asyncio
async def test_generar_moodle_connection_error_devuelve_502():
    _admin_user()
    with patch("app.routers.cierre_cursada.CierreCursadaService") as cls:
        cls.return_value.generar = AsyncMock(side_effect=MoodleConnectionError("Timeout con Moodle"))
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.post(
                "/api/v1/cierre-cursada/materias/10/generar",
                json={"cuatrimestre_id": 3},
            )
    assert resp.status_code == 502


@pytest.mark.asyncio
async def test_excel_corrida_inexistente_devuelve_404():
    _admin_user()
    with patch("app.routers.cierre_cursada.CierreCursadaService") as cls:
        cls.return_value.obtener_run = AsyncMock(return_value=None)
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.get("/api/v1/cierre-cursada/runs/999/excel")
    assert resp.status_code == 404


def _alumno_excel(apellido, nombre, email, estado, *, p1=80.0, p2=60.0, global_valor=9.0, nota_final=8):
    return CierreCursadaAlumno(
        moodle_user_id=1, nombre=nombre, apellido=apellido, email=email,
        comision_id=None, comision_nombre="Comisión 1", tutor_nombre="Juan Pérez",
        resultados_examenes=[
            {"examen_id": 1, "valor_real": p1},
            {"examen_id": 2, "valor_real": p2},
            {"examen_id": 3, "valor_real": global_valor},
        ],
        global_valor=global_valor,
        estado=EstadoCierreEnum(estado),
        nota_final=nota_final,
    )


@pytest.mark.asyncio
async def test_excel_run_exitoso_devuelve_xlsx_con_layout_corregido():
    _admin_user()
    examenes_snapshot = [
        {"id": 1, "tipo": "PARCIAL", "moodle_cmid": 101, "modo_aprobacion": "NUMERICO",
         "nota_minima": 60, "recupera_examen_id": None, "orden": 1},
        {"id": 2, "tipo": "PARCIAL", "moodle_cmid": 102, "modo_aprobacion": "NUMERICO",
         "nota_minima": 60, "recupera_examen_id": None, "orden": 2},
        {"id": 3, "tipo": "GLOBAL", "moodle_cmid": 103, "modo_aprobacion": "NUMERICO",
         "nota_minima": 6, "recupera_examen_id": None, "orden": 3},
    ]
    alumno_con_nf = _alumno_excel("Gómez", "Ana", "ana@x.com", "PROMOCIONA", nota_final=8)
    alumno_sin_nf = _alumno_excel(
        "Pérez", "Bruno", "bruno@x.com", "REGULARIZA", global_valor=None, nota_final=None
    )
    run = SimpleNamespace(
        id=1, materia_id=10, cuatrimestre_id=3,
        examenes_snapshot=examenes_snapshot,
        generado_por_id=1, total_alumnos=2,
        total_promociona=1, total_regulariza=1, total_recursa=0,
        created_at=datetime(2026, 1, 1, 12, 0),
        alumnos=[alumno_con_nf, alumno_sin_nf],
    )
    materia = SimpleNamespace(nombre="Programación 1")

    with patch("app.routers.cierre_cursada.CierreCursadaService") as service_cls, \
         patch("app.routers.cierre_cursada.MateriaRepository") as materia_cls:
        service_cls.return_value.obtener_run = AsyncMock(return_value=run)
        materia_cls.return_value.get_by_id = AsyncMock(return_value=materia)
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.get("/api/v1/cierre-cursada/runs/1/excel")

    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    # Layout (Bug 2/4): columnas del modelo + Nota Final + la columna auxiliar
    # "Recuperable" como última (8ª), sin TPs/dona.
    fila_headers = next(
        [c.value for c in row] for row in ws.iter_rows() if row[0].value == "Nombre y Apellido"
    )
    assert fila_headers == [
        "Nombre y Apellido", "Email", "Parcial 1", "Parcial 2",
        "Global TPI", "Estado Alumno", "Nota Final", "Recuperable",
    ]
    valores = {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}
    assert not (valores & {"TPs Aprobados", "Autoeval OK", "TPI", "Habilitado para Final"})
    assert len(ws._charts) == 0

    fila_ana = next(row for row in ws.iter_rows() if row[0].value == "Gómez, Ana")
    fila_bruno = next(row for row in ws.iter_rows() if row[0].value == "Pérez, Bruno")
    # Nota Final es ahora la penúltima columna (la última es "Recuperable").
    assert fila_ana[-2].value == 8  # Nota Final numérica
    # Bug 4 (nota_final sólo para PROMOCIONA): REGULARIZA renderiza la celda "Nota
    # Final" en blanco (no "N/E") — openpyxl guarda la cadena vacía como None al
    # persistir el .xlsx.
    assert fila_bruno[-2].value in ("", None)
    assert fila_bruno[4].value == "N/E"  # Global TPI también N/E


@pytest.mark.asyncio
async def test_historial_devuelve_corridas_de_la_materia():
    _admin_user()
    runs = [_run_response(id=2), _run_response(id=1)]
    with patch("app.routers.cierre_cursada.CierreCursadaService") as cls:
        cls.return_value.listar_historial = AsyncMock(return_value=runs)
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            resp = await client.get("/api/v1/cierre-cursada/materias/10/historial")
    assert resp.status_code == 200
    body = resp.json()
    assert len(body["runs"]) == 2
    assert {r["id"] for r in body["runs"]} == {1, 2}


# ===================================================================================
# Nivel 2 — flujo end-to-end de CierreCursadaService.generar (Moodle mockeado)
# ===================================================================================
#
# SQLite no soporta JSONB/ARRAY/enums PG nativos con el metadata GLOBAL de la app
# (mismo problema documentado en test_cierre_cursada_service.py): se crea SOLO el
# subconjunto de tablas necesario y se compila JSONB -> JSON para sqlite.


@compiles(JSONB, "sqlite")
def _jsonb_como_json_en_sqlite(element, compiler, **kw):
    return "JSON"


@compiles(ARRAY, "sqlite")
def _array_como_json_en_sqlite(element, compiler, **kw):
    return "JSON"


def _regexp_replace_sqlite(source: str | None, pattern: str, replacement: str) -> str | None:
    """UDF que replica `regexp_replace` de PostgreSQL para que `orden_natural_sql`
    (usado por `ComisionRepository.get_by_materia_con_tutores`, ejercitado por
    `CierreCursadaService.generar` en este flujo end-to-end) pueda ejecutarse contra
    SQLite en memoria. Mismo shim que `test_orden_natural_comision.py` y
    `test_orden_natural_cobertura_adicional.py` (bugfix comision-orden-numerico-fix)."""
    if source is None:
        return None
    return re.sub(pattern, replacement, source, count=1)


@pytest_asyncio.fixture
async def db_session():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", echo=False, future=True)

    @event.listens_for(engine.sync_engine, "connect")
    def _registrar_regexp_replace(dbapi_connection, connection_record):
        dbapi_connection.create_function("regexp_replace", 3, _regexp_replace_sqlite)

    tablas = [
        Usuario.__table__,
        Cohorte.__table__,
        Cuatrimestre.__table__,
        Materia.__table__,
        CoordinadorMateria.__table__,
        ExamenMateria.__table__,
        Comision.__table__,
        ComisionTutor.__table__,
        Unidad.__table__,
        ComponenteUnidad.__table__,
        Rubrica.__table__,
        Entrega.__table__,
        Correccion.__table__,
        CierreCursadaRun.__table__,
        CierreCursadaAlumno.__table__,
        Actividad.__table__,
    ]
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all, tables=tablas)

    session_maker = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with session_maker() as session:
        yield session

    await engine.dispose()


def _secciones_curso():
    """3 quiz (2 parciales + 1 global) — modname 'quiz' evita la rama de grade
    estructural (assign), que no hace falta para este flujo."""
    return [
        {
            "id": 1,
            "modules": [
                {"id": 101, "name": "Parcial 1", "modname": "quiz"},
                {"id": 102, "name": "Parcial 2", "modname": "quiz"},
                {"id": 103, "name": "Global", "modname": "quiz"},
            ],
        }
    ]


def _csv_calificador_dual():
    """Export dual con 'real' para los 3 ítems: Ana promociona, Beto regulariza
    (banda 60-20=40: 45>=40 y 50>=40, sin global), Carla recursa (20 < 40)."""
    header = (
        "Nombre,Apellido(s),Número de ID,Institución,Departamento,"
        "Dirección de correo,Quiz:Parcial 1 (Real),Quiz:Parcial 2 (Real),Quiz:Global (Real)"
    )
    filas = [
        "Ana,Gomez,,,,ana@x.com,80,90,9.0",
        "Beto,Perez,,,,beto@x.com,45,50,-",
        "Carla,Diaz,,,,carla@x.com,20,10,-",
    ]
    return "\n".join([header, *filas])


def _students():
    grupos = [{"id": 342, "name": "G1"}]
    return [
        {"id": 1, "firstname": "Ana", "lastname": "Gomez", "email": "ana@x.com",
         "roles": [{"shortname": "student"}], "groups": grupos},
        {"id": 2, "firstname": "Beto", "lastname": "Perez", "email": "beto@x.com",
         "roles": [{"shortname": "student"}], "groups": grupos},
        {"id": 3, "firstname": "Carla", "lastname": "Diaz", "email": "carla@x.com",
         "roles": [{"shortname": "student"}], "groups": grupos},
    ]


def _mock_moodle_service():
    """Reemplazo de MoodleService: sólo lo que `generar` necesita, sin I/O real."""
    sesion = MagicMock()
    sesion.aclose = AsyncMock()

    mock = MagicMock()
    mock.get_token = AsyncMock(return_value="tok123")
    mock.crear_cliente_sesion = MagicMock(return_value=sesion)
    mock.login_sesion = AsyncMock(return_value=None)
    mock.get_enrolled_users_full = AsyncMock(return_value=_students())
    mock.get_course_contents = AsyncMock(return_value=_secciones_curso())
    mock.descargar_export_calificador = AsyncMock(return_value=_csv_calificador_dual())
    mock.get_grades_full = AsyncMock(return_value={})
    # Bug 3: fuente autoritativa de grupos (`core_group_*`). El curso tiene un único
    # grupo de comisión "G1" (id 342, moodle_group_id de Comisión 1) con los 3 alumnos
    # como miembros → el mapa autoritativo uid→grupos asocia a los 3 a Comisión 1.
    mock.get_course_groups = AsyncMock(return_value=[{"id": 342, "name": "G1"}])
    mock.get_group_members = AsyncMock(return_value=[1, 2, 3])
    return mock


async def _crear_fixture_materia_con_examenes(db_session) -> tuple[Materia, int, Usuario]:
    usuario = Usuario(
        username="coord1", password_hash="x", nombre="Coordinador", rol=RolEnum.ADMIN,
        moodle_username="coord1", moodle_password_encrypted="enc", moodle_host="https://moodle.example.com",
    )
    cohorte = Cohorte(codigo="M26", nombre="M26")
    db_session.add_all([usuario, cohorte])
    await db_session.flush()

    cuatrimestre = Cuatrimestre(nombre="1C", cohorte_id=cohorte.id, numero=1)
    materia = Materia(nombre="Programación 1", codigo="P1", moodle_course_id=555)
    db_session.add_all([cuatrimestre, materia])
    await db_session.flush()

    comision = Comision(materia_id=materia.id, nombre="Comisión 1", anio=2026, moodle_group_id=342)
    db_session.add(comision)
    await db_session.flush()
    db_session.add(ComisionTutor(comision_id=comision.id, tutor_id=usuario.id))

    examenes = [
        ExamenMateria(materia_id=materia.id, tipo="PARCIAL", moodle_cmid=101,
                       modo_aprobacion="NUMERICO", nota_minima=60, orden=1),
        ExamenMateria(materia_id=materia.id, tipo="PARCIAL", moodle_cmid=102,
                       modo_aprobacion="NUMERICO", nota_minima=60, orden=2),
        ExamenMateria(materia_id=materia.id, tipo="GLOBAL", moodle_cmid=103,
                       modo_aprobacion="NUMERICO", nota_minima=6, orden=3),
    ]
    db_session.add_all(examenes)
    await db_session.commit()
    return materia, cuatrimestre.id, usuario


@pytest.mark.asyncio
async def test_generar_end_to_end_conteos_y_comision_correctos(db_session):
    materia, cuatrimestre_id, usuario = await _crear_fixture_materia_con_examenes(db_session)

    service = CierreCursadaService(db_session)
    service.moodle = _mock_moodle_service()

    run = await service.generar(materia.id, cuatrimestre_id, usuario)

    assert run.total_alumnos == 3
    assert run.total_promociona == 1
    assert run.total_regulariza == 1
    assert run.total_recursa == 1

    # `crear_run` deja `run.alumnos` expirado tras el refresh; recargamos con
    # eager-load (igual que `obtener_run`) para poder acceder a la relación.
    run = await CierreCursadaRepository(db_session).get_run(run.id)
    por_email = {a.email: a for a in run.alumnos}
    assert por_email["ana@x.com"].estado == EstadoCierreEnum.PROMOCIONA
    assert por_email["beto@x.com"].estado == EstadoCierreEnum.REGULARIZA
    assert por_email["carla@x.com"].estado == EstadoCierreEnum.RECURSA

    # Agrupación por comisión: los 3 alumnos matchean el grupo "G1" (moodle_group_id=342) -> Comisión 1.
    for alumno in run.alumnos:
        assert alumno.comision_nombre == "Comisión 1"
        assert alumno.tutor_nombre == "Coordinador"


@pytest.mark.asyncio
async def test_generar_dos_veces_append_only_no_pisa_la_primera(db_session):
    materia, cuatrimestre_id, usuario = await _crear_fixture_materia_con_examenes(db_session)

    service = CierreCursadaService(db_session)
    service.moodle = _mock_moodle_service()
    primera = await service.generar(materia.id, cuatrimestre_id, usuario)

    service2 = CierreCursadaService(db_session)
    service2.moodle = _mock_moodle_service()
    segunda = await service2.generar(materia.id, cuatrimestre_id, usuario)

    assert primera.id != segunda.id

    repo = CierreCursadaRepository(db_session)
    runs = await repo.listar_runs(materia.id)
    assert len(runs) == 2
    assert {r.id for r in runs} == {primera.id, segunda.id}

    recargada = await repo.get_run(primera.id)
    assert recargada is not None
    assert recargada.total_alumnos == 3  # la primera corrida sigue intacta
