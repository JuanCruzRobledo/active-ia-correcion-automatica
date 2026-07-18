# app/services/excel_cierre_cursada.py
"""
Genera el .xlsx de un cierre de cursada (CierreCursadaRun + CierreCursadaAlumno ya
calculados y persistidos — este módulo NO calcula nada, solo escribe lo que ya está en
la corrida).

Layout = modelo `docs/modelos/modelo planilla de cierre CORREGIDO.xlsx` (ver diseño,
sección "Layout exacto del modelo CORREGIDO") + una columna `Nota Final` agregada al
final:
  R1: "TOTAL MATERIA {NOMBRE}" (merge A1:B1, mayúsculas)
  R2-R6: resumen PROMOCIONADOS/REGULARES/RECURSANTES/RECUPERABLES/ABANDONOS con el conteo
         como fórmula nativa `COUNTIF` (A=etiqueta, B=fórmula)
  R7: vacía (separador)
  R8+: por comisión, barra "{comisión} — Tutor: {tutor}" (merge desde la columna C hasta
       la última) + encabezados + filas de alumnos.

Columnas dinámicas, derivadas de `run.examenes_snapshot` (config de exámenes congelada
en la corrida): `Nombre y Apellido | Email | Parcial n… | Global TPI | Estado Alumno |
Nota Final`. Sin columnas de TPs y sin gráfico de dona (a diferencia del sistema viejo).

Fórmulas nativas y locale:
    Las fórmulas nativas que este módulo escribe (columna "Recuperable" por fila y los
    conteos del resumen) usan la sintaxis del **formato de archivo** de Excel, que es
    siempre en inglés (`IF`, `AND`, `IFERROR`, `VALUE`, `COUNTIF`) y con coma `,` como
    separador de argumentos, independientemente del idioma del Excel del usuario. Es lo
    que openpyxl persiste literalmente en el XML de la celda (string que empieza con `=`
    en `cell.value`). Escribirlas en español (`SI`/`Y`/`VALOR`/`SI.ERROR`) o con `;`
    hace que Excel las rechace con `#NAME?` en cualquier locale. Al abrir el archivo,
    un Excel es-AR las **muestra** traducidas (`SI`/`Y`/`VALOR`/`SI.ERROR`, `;`), pero en
    disco quedan siempre en inglés + coma.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.core.fecha import ahora_ar
from app.models.cierre_cursada import CierreCursadaAlumno, CierreCursadaRun
from app.services import excel_estilos
from app.services.excel_estilos import (
    CENTER,
    FILL_BLOQUE,
    FONT_BLOQUE,
    FONT_SECCION,
    banda_titulo,
    celda_header,
    fila_datos,
)
from app.utils.orden_natural import natural_key


def _fmt_valor(valor: float | int | None) -> str | float | int:
    """Valor numérico (`Parcial n` / `Global TPI`) o `N/E` si no hay nota."""
    return "N/E" if valor is None else valor


def _fmt_nota_final(valor: int | None) -> str | int:
    """Nota Final: celda en blanco (`""`) si no hay nota (alumno no promociona).
    Diverge a propósito de `_fmt_valor` (columnas de examen), que sigue devolviendo
    `"N/E"` ante `None` (preserva 3.10)."""
    return "" if valor is None else valor


def _formula_recuperable(estado_col: str, p1_col: str, p2_col: str, fila: int) -> str:
    """Fórmula nativa es-AR "Recuperable" para la fila `fila` (1-based, número de fila de
    Excel del alumno).

    Marca "RECUPERABLE CON PARCIAL 2" (falta rendir el parcial 2) o "RECUPERABLE CON
    PARCIAL 1" (falta el parcial 1) sólo para alumnos RECURSA con un parcial `>=40` y el
    otro `<40`/N/E (los `N/E` se tratan como 0 vía `SI.ERROR(VALOR(...);0)`); vacío en
    cualquier otro caso.

    Las columnas se reciben como letras (`estado_col`/`p1_col`/`p2_col`) para que la
    fórmula quede correcta en ambas hojas ("por Comisiones": F/C/D; "Crudo": H/E/F). Usa
    la sintaxis nativa de Excel (formato de archivo): nombres de función en inglés
    (`IF`/`AND`/`IFERROR`/`VALUE`) y coma `,` como separador — ver la nota de locale del
    módulo. Excel la mostrará traducida (`SI`/`Y`/`VALOR`/`SI.ERROR`, `;`) en un Excel
    es-AR.
    """
    return (
        f'=IF({estado_col}{fila}<>"RECURSA","",'
        f"IF(AND(IFERROR(VALUE({p1_col}{fila}),0)>=40,IFERROR(VALUE({p2_col}{fila}),0)<40),"
        f'"RECUPERABLE CON PARCIAL 2",'
        f"IF(AND(IFERROR(VALUE({p2_col}{fila}),0)>=40,IFERROR(VALUE({p1_col}{fila}),0)<40),"
        f'"RECUPERABLE CON PARCIAL 1","")))'
    )


def _formulas_conteo_resumen(
    estado_letra: str, recuperable_letra: str, recuperable_criterio: str
) -> list[tuple[str, str]]:
    """[(etiqueta, fórmula `COUNTIF`)] del resumen — conteos como fórmulas nativas es-AR
    (no enteros estáticos), que recalculan al editar la columna "Estado Alumno".

    Los estados se cuentan sobre la columna Estado (`estado_letra`) y los recuperables
    sobre la columna auxiliar `Recuperable` (`recuperable_letra`). El criterio de
    recuperable cambia por hoja: `"RECUPERABLE*"` en "por Comisiones", `"RECUPERABLE CON*"`
    en "Crudo" (`recuperable_criterio`). Sintaxis nativa de Excel: `COUNTIF` (inglés) y
    coma `,` como separador — ver la nota de locale del módulo (2.11, 2.12, 2.15)."""
    def _por_estado(patron: str) -> str:
        return f'=COUNTIF({estado_letra}:{estado_letra},"{patron}")'

    return [
        ("PROMOCIONADOS", _por_estado("PROMOCIONA")),
        ("REGULARES", _por_estado("REGULARIZA")),
        ("RECURSANTES", _por_estado("RECURSA")),
        (
            "RECUPERABLES",
            f'=COUNTIF({recuperable_letra}:{recuperable_letra},"{recuperable_criterio}")',
        ),
        ("ABANDONOS", _por_estado("ABANDONO")),
    ]


def _recuperable_para_fila(estado_letra: str, cols_parciales: list[int], fila: int) -> str:
    """Valor de la celda `Recuperable` de la fila `fila`: la fórmula nativa cuando hay al
    menos dos parciales (usa los dos primeros como P1/P2), o `""` si la materia no tiene
    dos parciales (no aplica la lógica de recuperable). `cols_parciales` son los índices
    1-based de columna de los parciales, en orden."""
    if len(cols_parciales) < 2:
        return ""
    return _formula_recuperable(
        estado_letra,
        get_column_letter(cols_parciales[0]),
        get_column_letter(cols_parciales[1]),
        fila,
    )


def _parciales_y_global(
    examenes_snapshot: list[dict] | None,
) -> tuple[list[dict], dict | None]:
    """(parciales ordenados por orden/id, examen GLOBAL o None) desde la config congelada."""
    snapshot = examenes_snapshot or []
    parciales = sorted(
        (e for e in snapshot if e.get("tipo") == "PARCIAL"),
        key=lambda e: (e.get("orden") if e.get("orden") is not None else 0, e.get("id") or 0),
    )
    globales = [e for e in snapshot if e.get("tipo") == "GLOBAL"]
    return parciales, (globales[0] if globales else None)


def _headers(parciales: list[dict], global_examen: dict | None) -> list[str]:
    """Encabezados de la hoja "por Comisiones". Incluye la columna auxiliar `Recuperable`
    al final (col H en el layout estándar 2 parciales + 1 global): A Nombre y Apellido |
    B Email | C Parcial 1 | D Parcial 2 | E Global TPI | F Estado Alumno | G Nota Final |
    H Recuperable. La fórmula por fila de `Recuperable` la escribe `_escribir_detalle`."""
    headers = ["Nombre y Apellido", "Email"]
    headers += [f"Parcial {i}" for i in range(1, len(parciales) + 1)]
    if global_examen is not None:
        headers.append("Global TPI")
    headers += ["Estado Alumno", "Nota Final", "Recuperable"]
    return headers


def _valor_parcial(alumno: CierreCursadaAlumno, examen_id: int) -> float | None:
    """`valor_real` del examen `examen_id` en `alumno.resultados_examenes` (matcheo por id)."""
    for r in alumno.resultados_examenes or []:
        if r.get("examen_id") == examen_id:
            return r.get("valor_real")
    return None


def _barra_bloque_desde_c(ws, fila: int, texto: str, ncols: int) -> None:
    """Barra de bloque por comisión, igual a `excel_estilos.barra_bloque` pero mergeando
    desde la columna **C** (no desde A, a diferencia del helper estándar) hasta la última
    columna (`Nota Final`). Reutiliza los estilos de la casa (`FILL_BLOQUE`/`FONT_BLOQUE`);
    NO modifica la firma pública de `excel_estilos.barra_bloque` (la usan otros reportes).
    """
    ws.merge_cells(start_row=fila, start_column=3, end_row=fila, end_column=ncols)
    for col in range(3, ncols + 1):
        ws.cell(fila, col).fill = FILL_BLOQUE
    c = ws.cell(fila, 3, texto)
    c.font = FONT_BLOQUE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 20


_SIN_COMISION = "Sin comisión asignada"


def _agrupar_por_comision(alumnos: list[CierreCursadaAlumno]) -> dict[str, list[CierreCursadaAlumno]]:
    """{titulo_bloque: [alumnos]} — un bloque por comisión (+ tutor), ORDENADO por el
    orden numérico natural de `comision_nombre` (alfabético si no hay sufijo numérico),
    con "Sin comisión asignada" siempre al final. El título compuesto ("{comisión} —
    Tutor: {tutor}") se conserva por bloque; el orden se calcula sobre el nombre de
    comisión subyacente, no sobre ese título compuesto.
    """
    grupos: dict[str, tuple[str, list[CierreCursadaAlumno]]] = {}
    for a in alumnos:
        nombre = a.comision_nombre or _SIN_COMISION
        titulo = f"{nombre} — Tutor: {a.tutor_nombre}" if a.tutor_nombre else nombre
        if nombre not in grupos:
            grupos[nombre] = (titulo, [])
        grupos[nombre][1].append(a)

    def _clave_orden(item: tuple[str, tuple[str, list[CierreCursadaAlumno]]]):
        nombre = item[0]
        if nombre == _SIN_COMISION:
            return (1, ())
        return (0, natural_key(nombre))

    return {
        titulo: alumnos_del_grupo
        for _, (titulo, alumnos_del_grupo) in sorted(grupos.items(), key=_clave_orden)
    }


def _escribir_resumen(ws, materia_nombre: str, estado_letra: str, recuperable_letra: str) -> None:
    """R1: título (merge A1:B1, mayúsculas). R2-R6: conteo PROMOCIONADOS/REGULARES/
    RECURSANTES/RECUPERABLES/ABANDONOS en A:B, con la celda de conteo (col B) como
    fórmula nativa `COUNTIF` (no un entero estático) que recalcula al editar la columna
    "Estado Alumno". En el layout estándar Estado es F y Recuperable H (2.11, 2.12, 2.15)."""
    banda_titulo(ws, f"TOTAL MATERIA {materia_nombre.upper()}", ncols=2)

    resumen = _formulas_conteo_resumen(estado_letra, recuperable_letra, "RECUPERABLE*")
    for fila, (etiqueta, formula) in enumerate(resumen, start=2):
        ws.cell(fila, 1, etiqueta).font = FONT_SECCION
        celda_conteo = ws.cell(fila, 2, formula)
        celda_conteo.alignment = CENTER


def _escribir_detalle(
    ws,
    alumnos: list[CierreCursadaAlumno],
    parciales: list[dict],
    global_examen: dict | None,
    headers: list[str],
    fila_inicio: int,
) -> int:
    """Escribe los bloques por comisión desde `fila_inicio`. Devuelve la última fila usada.

    La última columna es `Recuperable`: en cada fila de alumno se escribe la fórmula
    nativa es-AR (`_formula_recuperable`) referenciando la fila de Excel de ese alumno,
    con las letras de columna calculadas del layout (estado, primeros dos parciales,
    recuperable) → en el layout estándar F/C/D/H (2.13, 2.14)."""
    ncols = len(headers)
    # Layout: …, Estado Alumno (antepenúltima), Nota Final (penúltima), Recuperable (última).
    col_estado = ncols - 2
    # Parciales arrancan en la col 3 (tras Nombre y Apellido / Email).
    cols_parciales = [3 + i for i in range(len(parciales))]
    estado_letra = get_column_letter(col_estado)
    fila = fila_inicio
    for titulo, del_grupo in _agrupar_por_comision(alumnos).items():
        fila += 1
        _barra_bloque_desde_c(ws, fila, titulo, ncols)
        fila += 1
        for col, h in enumerate(headers, 1):
            celda_header(ws, fila, col, h)
        fila += 1
        for i, a in enumerate(sorted(del_grupo, key=lambda x: (x.apellido or "", x.nombre or ""))):
            estado = a.estado.value
            valores = [f"{a.apellido}, {a.nombre}".strip(", "), a.email]
            valores += [_fmt_valor(_valor_parcial(a, p["id"])) for p in parciales]
            if global_examen is not None:
                valores.append(_fmt_valor(a.global_valor))
            valores.append(estado)
            valores.append(_fmt_nota_final(a.nota_final))
            valores.append(_recuperable_para_fila(estado_letra, cols_parciales, fila))
            fila_datos(
                ws, fila, valores,
                zebra=(i % 2 == 1),
                resaltar_col=col_estado if estado == "RECURSA" else None,
                center_cols=tuple(range(3, ncols + 1)),
            )
            fila += 1
        fila += 1  # separador entre bloques
    return fila


def _headers_crudo(parciales: list[dict], global_examen: dict | None) -> list[str]:
    """Encabezados de la hoja "Crudo": agrega `Comision`/`Tutor` (C/D) respecto de la
    hoja "por Comisiones" y una columna `Recuperable` al final.

    Layout (2 parciales + 1 global): A Nombre y Apellido | B Email | C Comision |
    D Tutor | E Parcial 1 | F Parcial 2 | G Global TPI | H Estado Alumno | I Nota Final |
    J Recuperable (coincide con `docs/modelos/Cierre_Programacin_3 FIX.xlsx`)."""
    headers = ["Nombre y Apellido", "Email", "Comision", "Tutor"]
    headers += [f"Parcial {i}" for i in range(1, len(parciales) + 1)]
    if global_examen is not None:
        headers.append("Global TPI")
    headers += ["Estado Alumno", "Nota Final", "Recuperable"]
    return headers


def _escribir_resumen_crudo(ws, materia_nombre: str, estado_letra: str, recuperable_letra: str) -> None:
    """R1: título (merge A1:B1, mayúsculas). R2-R6: conteo PROMOCIONADOS/REGULARES/
    RECURSANTES/RECUPERABLES/ABANDONOS en A:B, con la celda de conteo (col B) como
    fórmula nativa `COUNTIF` que recalcula al editar la columna "Estado Alumno". En el
    layout estándar Estado es H y Recuperable J; el criterio de recuperable es
    `"RECUPERABLE CON*"`. Reutiliza `banda_titulo`/`FONT_SECCION` (estilos de la casa →
    3.10) (2.11, 2.12, 2.15)."""
    banda_titulo(ws, f"TOTAL MATERIA {materia_nombre.upper()}", ncols=2)

    resumen = _formulas_conteo_resumen(estado_letra, recuperable_letra, "RECUPERABLE CON*")
    for fila, (etiqueta, formula) in enumerate(resumen, start=2):
        ws.cell(fila, 1, etiqueta).font = FONT_SECCION
        celda_conteo = ws.cell(fila, 2, formula)
        celda_conteo.alignment = CENTER


def _escribir_hoja_cruda(
    ws,
    materia_nombre: str,
    run: CierreCursadaRun,
    parciales: list[dict],
    global_examen: dict | None,
) -> None:
    """Hoja 2 "Crudo": resumen (R1-R6) + lista PLANA de TODOS los alumnos ordenados
    alfabéticamente por `(apellido, nombre)` — sin cuadros por comisión —, agregando las
    columnas `Comision` ("Sin comisión asignada" si aplica) y `Tutor` del alumno.

    Las columnas de examen muestran `N/E` cuando no hay nota (`_fmt_valor` → 3.8) y se
    usan los helpers de `excel_estilos.py` (`banda_titulo`/`celda_header`/`fila_datos`
    → 3.10). La columna `Recuperable` lleva la fórmula nativa es-AR por fila
    (`_formula_recuperable`) y los conteos del resumen quedan como placeholders
    (tarea 10.4).
    """
    headers = _headers_crudo(parciales, global_examen)
    ncols = len(headers)
    # Estado Alumno es la antepenúltima columna (le siguen Nota Final y Recuperable).
    col_estado = ncols - 2
    estado_letra = get_column_letter(col_estado)
    # Parciales arrancan en la col 5 (tras Nombre y Apellido / Email / Comision / Tutor).
    cols_parciales = [5 + i for i in range(len(parciales))]

    _escribir_resumen_crudo(ws, materia_nombre, estado_letra, get_column_letter(ncols))

    fila_header = 8  # R7 queda como separador entre el resumen y la tabla plana.
    for col, h in enumerate(headers, 1):
        celda_header(ws, fila_header, col, h)

    alumnos_ordenados = sorted(
        run.alumnos, key=lambda x: (x.apellido or "", x.nombre or "")
    )
    fila = fila_header + 1
    for i, a in enumerate(alumnos_ordenados):
        estado = a.estado.value
        valores = [
            f"{a.apellido}, {a.nombre}".strip(", "),
            a.email,
            a.comision_nombre or _SIN_COMISION,
            a.tutor_nombre or "",
        ]
        valores += [_fmt_valor(_valor_parcial(a, p["id"])) for p in parciales]
        if global_examen is not None:
            valores.append(_fmt_valor(a.global_valor))
        valores.append(estado)
        valores.append(_fmt_nota_final(a.nota_final))
        # Recuperable: fórmula nativa es-AR por fila (col J en el layout estándar; estado
        # H, parciales E/F). Se referencia la fila de Excel `fila` del alumno (2.13, 2.14).
        valores.append(_recuperable_para_fila(estado_letra, cols_parciales, fila))
        fila_datos(
            ws, fila, valores,
            zebra=(i % 2 == 1),
            resaltar_col=col_estado if estado == "RECURSA" else None,
            center_cols=tuple(range(5, ncols + 1)),
        )
        fila += 1

    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26 if col <= 2 else 16


def _escribir_hoja_por_comisiones(ws, materia_nombre: str, run: CierreCursadaRun) -> None:
    """Hoja 1 "por Comisiones": resumen (R1-R4) + bloques por comisión (R6+) sobre `ws`.

    Concentra el layout histórico del cierre (resumen + detalle agrupado por comisión) y
    reutiliza los helpers de la casa (`banda_titulo`/`celda_header`/`fila_datos`, vía
    `_escribir_resumen`/`_escribir_detalle`) → preserva estilos y orden (3.6, 3.10).
    """
    parciales, global_examen = _parciales_y_global(run.examenes_snapshot)
    headers = _headers(parciales, global_examen)
    ncols = len(headers)
    # Layout: …, Estado (ncols-2), Nota Final (ncols-1), Recuperable (ncols) → estándar F/H.
    estado_letra = get_column_letter(ncols - 2)
    recuperable_letra = get_column_letter(ncols)

    _escribir_resumen(ws, materia_nombre, estado_letra, recuperable_letra)
    # Resumen ocupa R1 (título) + R2-R6 (5 conteos); R7 queda como separador y los
    # bloques por comisión arrancan en R8 (bar) — igual patrón "una fila en blanco" que
    # el layout original de 3 conteos.
    _escribir_detalle(ws, run.alumnos, parciales, global_examen, headers, fila_inicio=7)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26 if col <= 2 else 16


def generar_excel_cierre(materia_nombre: str, run: CierreCursadaRun) -> tuple[bytes, str]:
    """(bytes, filename) del .xlsx de un cierre ya generado. `run.alumnos` debe venir cargado.

    Orquesta DOS hojas (modelo `docs/modelos/Cierre_Programacin_3 FIX.xlsx`):
      - Hoja 1 "{materia} por Comisiones": cuadros por comisión (ver
        `_escribir_hoja_por_comisiones`).
      - Hoja 2 "{materia} Crudo": lista plana alfabética con columnas Comision/Tutor
        (contenido en `_escribir_hoja_cruda`, tarea 10.2). Acá se crea la hoja con el
        nombre correcto para dejar el orquestador cableado.
    """
    wb = Workbook()
    parciales, global_examen = _parciales_y_global(run.examenes_snapshot)

    ws1 = wb.active
    ws1.title = excel_estilos.sheet_title(f"{materia_nombre} por Comisiones", default="Cierre")
    _escribir_hoja_por_comisiones(ws1, materia_nombre, run)

    # Hoja 2 "Crudo": lista plana alfabética con columnas Comision/Tutor (tarea 10.2).
    ws2 = wb.create_sheet(excel_estilos.sheet_title(f"{materia_nombre} Crudo", default="Crudo"))
    _escribir_hoja_cruda(ws2, materia_nombre, run, parciales, global_examen)

    buffer = io.BytesIO()
    wb.save(buffer)
    fecha = ahora_ar().strftime("%Y%m%d")
    filename = excel_estilos.sanitize_filename(f"Cierre_{materia_nombre}_{fecha}.xlsx")
    return buffer.getvalue(), filename
