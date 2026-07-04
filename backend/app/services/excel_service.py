# app/services/excel_service.py
"""
Excel service for Active-IA.

Business logic for exporting grades to Excel format.
Uses openpyxl to create .xlsx files with student grades and correction status.

Ref: docs/specs/03-REQUISITOS-FUNCIONALES.md seccion 10.4
"""

import io
from datetime import datetime

from app.core.fecha import ahora_ar, fmt_fecha_ar
from app.services import excel_estilos

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.comision import Comision
from app.models.entrega import Entrega as EntregaModel
from app.repositories.entrega_repository import EntregaRepository


class ExcelService:
    """Service for generating Excel documents."""

    def __init__(self, db: AsyncSession):
        """Initialize Excel service."""
        self.db = db
        self.entrega_repo = EntregaRepository(db)

    async def exportar_notas_excel(
        self,
        comision_id: int,
        rubrica_id: int,
    ) -> tuple[bytes, str]:
        """
        Export grades to Excel format.

        Args:
            comision_id: ID of the comision.
            rubrica_id: ID of the rubrica.

        Returns:
            Tuple of (Excel file bytes, suggested filename).

        Raises:
            HTTPException 404: No entregas found.
        """
        from fastapi import HTTPException, status

        # Get all entregas for this comision and rubrica
        entregas_list, _ = await self.entrega_repo.get_all(
            comision_id=comision_id,
            rubrica_id=rubrica_id,
            page=1,
            per_page=1000,  # Get all
        )

        if not entregas_list:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No hay entregas para esta comisión y rúbrica",
            )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Notas"

        # Re-fetch first entrega with full nested relations (comision -> materia, rubrica)
        first_id = entregas_list[0].id
        stmt = (
            select(EntregaModel)
            .options(
                selectinload(EntregaModel.comision).selectinload(Comision.materia),
                selectinload(EntregaModel.rubrica),
            )
            .where(EntregaModel.id == first_id)
        )
        result = await self.db.execute(stmt)
        first_entrega = result.scalar_one()

        materia_nombre = first_entrega.comision.materia.nombre
        materia_codigo = first_entrega.comision.materia.codigo
        comision_nombre = first_entrega.comision.nombre
        rubrica_nombre = first_entrega.rubrica.titulo
        # Use .value to get raw string from SQLAlchemy enum (avoids "TipoRubricaEnum.TP" in filename)
        rubrica_tipo = first_entrega.rubrica.tipo.value if hasattr(first_entrega.rubrica.tipo, 'value') else str(first_entrega.rubrica.tipo)
        rubrica_numero = str(first_entrega.rubrica.numero)

        # Add title — row 1: materia name
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"{materia_nombre} ({materia_codigo})"
        title_cell.font = Font(size=14, bold=True, color="1F4788")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(
            start_color="E7F3FF", end_color="E7F3FF", fill_type="solid"
        )

        # Row 2: comision and rubrica info
        ws.merge_cells("A2:F2")
        subtitle_cell = ws["A2"]
        subtitle_cell.value = (
            f"Comisión: {comision_nombre}  |  "
            f"TP: {rubrica_tipo} {rubrica_numero} - {rubrica_nombre}"
        )
        subtitle_cell.font = Font(size=11, italic=True, color="444444")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        subtitle_cell.fill = PatternFill(
            start_color="F0F7FF", end_color="F0F7FF", fill_type="solid"
        )

        # Add headers — row 4 (one row spacing after subtitle)
        headers = ["Alumno", "Nota", "Estado", "Fecha Corrección", "Editado", "Comentario General"]
        header_row = 4

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="1F4788", end_color="1F4788", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Add data rows
        current_row = header_row + 1

        for entrega in entregas_list:
            # Alumno
            ws.cell(row=current_row, column=1).value = entrega.alumno_nombre

            # Nota, Estado, Fecha, Editado, Comentario
            if entrega.correccion:
                nota = entrega.correccion.nota
                estado = "CORREGIDA"
                fecha = entrega.correccion.created_at.strftime("%d/%m/%Y %H:%M")
                editado = "Sí" if entrega.correccion.editado_manualmente else "No"
                comentario = entrega.correccion.comentario_general or ""

                # Nota cell with color
                nota_cell = ws.cell(row=current_row, column=2)
                nota_cell.value = nota
                nota_cell.fill = self._get_nota_fill(nota)
                nota_cell.alignment = Alignment(horizontal="center")
            else:
                # No corregida
                estado_map = {
                    "SUBIDA": "PENDIENTE",
                    "PENDIENTE": "EN PROCESO",
                    "ERROR": "ERROR",
                }
                estado = estado_map.get(entrega.estado, entrega.estado)
                fecha = "-"
                editado = "-"
                comentario = ""

                # Nota vacía
                ws.cell(row=current_row, column=2).value = "-"
                ws.cell(row=current_row, column=2).alignment = Alignment(
                    horizontal="center"
                )

            # Estado
            estado_cell = ws.cell(row=current_row, column=3)
            estado_cell.value = estado
            estado_cell.alignment = Alignment(horizontal="center")

            # Fecha
            ws.cell(row=current_row, column=4).value = fecha
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="center")

            # Editado
            ws.cell(row=current_row, column=5).value = editado
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="center")

            # Comentario general
            comentario_cell = ws.cell(row=current_row, column=6)
            comentario_cell.value = comentario
            comentario_cell.alignment = Alignment(horizontal="left", wrap_text=True)

            # Add borders to all cells
            for col_num in range(1, 7):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

            current_row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 32  # Alumno
        ws.column_dimensions["B"].width = 10  # Nota
        ws.column_dimensions["C"].width = 15  # Estado
        ws.column_dimensions["D"].width = 20  # Fecha Corrección
        ws.column_dimensions["E"].width = 10  # Editado
        ws.column_dimensions["F"].width = 60  # Comentario General

        # Add summary row
        summary_row = current_row + 1
        ws.merge_cells(f"A{summary_row}:C{summary_row}")
        summary_cell = ws.cell(row=summary_row, column=1)
        summary_cell.value = f"Total de entregas: {len(entregas_list)}"
        summary_cell.font = Font(bold=True)

        # Freeze header rows
        ws.freeze_panes = "A5"

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()
        excel_buffer.close()

        # Build suggested filename using text code (materia.codigo), not ID
        fecha = datetime.now().strftime("%Y%m%d")

        excel_filename = f"notas_{materia_codigo}_{comision_nombre}_{rubrica_tipo}{rubrica_numero}_{fecha}.xlsx"
        excel_filename = self._sanitize_filename(excel_filename)

        return excel_bytes, excel_filename

    # =========================================================================
    # Gestión (pantalla rol GESTOR) — export multi-hoja
    # =========================================================================

    GESTION_HEADERS = ["Nombre", "Apellido", "Email", "Regional", "Comisión", "Tiempo de inactividad"]

    def exportar_gestion(
        self,
        resultado,
        materia_nombre: str,
        agrupar_por: str = "regional",
    ) -> tuple[bytes, str]:
        """Arma el .xlsx de la pantalla Gestión, con las columnas pedidas, total por
        hoja y una hoja "Resumen" (total por grupo + global).

        `agrupar_por`:
        - "regional" (Excel para Nexos): una hoja por Regional, ordenada por comisión.
        - "comision" (Excel para Tutores): una hoja por Comisión, ordenada por regional.

        `resultado` es un ResultadoGestion (items: AlumnoGestion, total: int). No toca
        la DB: se construye sólo a partir del resultado ya consultado.
        """
        por_comision = agrupar_por == "comision"
        clave = (lambda a: a.comision) if por_comision else (lambda a: a.regional)
        orden_interno = (
            (lambda a: (a.regional.lower(), a.apellido.lower(), a.nombre.lower()))
            if por_comision
            else (lambda a: (a.comision, a.apellido.lower(), a.nombre.lower()))
        )
        etiqueta = "Comisión" if por_comision else "Regional"
        variante = "tutores" if por_comision else "nexos"

        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        grupos: dict[str, list] = {}
        for a in resultado.items:
            grupos.setdefault(clave(a), []).append(a)

        wb = Workbook()
        resumen = wb.active
        resumen.title = "Resumen"

        # --- Una hoja por grupo (regional o comisión) ---
        for grupo in sorted(grupos):
            alumnos = sorted(grupos[grupo], key=orden_interno)
            ws = wb.create_sheet(title=self._sheet_title(grupo))

            ws.merge_cells("A1:F1")
            titulo = ws["A1"]
            titulo.value = f"{etiqueta}: {grupo}"
            titulo.font = Font(size=13, bold=True, color="1F4788")
            titulo.alignment = Alignment(horizontal="center", vertical="center")

            for col, h in enumerate(self.GESTION_HEADERS, 1):
                c = ws.cell(row=2, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border

            row = 3
            for a in alumnos:
                valores = [a.nombre, a.apellido, a.email, a.regional, a.comision, a.tiempo_inactividad]
                for col, val in enumerate(valores, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = border
                row += 1

            total_cell = ws.cell(row=row + 1, column=1, value=f"Total de alumnos: {len(alumnos)}")
            total_cell.font = Font(bold=True)

            for col, width in enumerate([20, 20, 32, 24, 14, 20], 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            ws.freeze_panes = "A3"

        # --- Hoja Resumen ---
        resumen["A1"] = f"Gestión — {materia_nombre}"
        resumen["A1"].font = Font(size=14, bold=True, color="1F4788")
        resumen["A2"] = f"Documento generado el {fmt_fecha_ar(ahora_ar())}"
        resumen["A2"].font = Font(italic=True, size=10)
        for col, h in enumerate([etiqueta, "Alumnos"], 1):
            c = resumen.cell(row=3, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        r = 4
        for grupo in sorted(grupos):
            resumen.cell(row=r, column=1, value=grupo)
            resumen.cell(row=r, column=2, value=len(grupos[grupo]))
            r += 1
        resumen.cell(row=r, column=1, value="Total").font = Font(bold=True)
        resumen.cell(row=r, column=2, value=resultado.total).font = Font(bold=True)
        resumen.column_dimensions["A"].width = 26
        resumen.column_dimensions["B"].width = 12

        buffer = io.BytesIO()
        wb.save(buffer)
        data = buffer.getvalue()
        buffer.close()

        fecha = datetime.now().strftime("%Y%m%d")
        filename = self._sanitize_filename(f"gestion_{variante}_{materia_nombre}_{fecha}.xlsx")
        return data, filename

    def _sheet_title(self, regional: str) -> str:
        """Nombre de hoja válido para Excel: sin caracteres prohibidos y ≤ 31 chars."""
        return excel_estilos.sheet_title(regional, default="Sin regional")

    def exportar_pendientes(self, pendientes, materia_nombre: str, agrupar_por: str = "trabajo") -> tuple[bytes, str]:
        """Arma el .xlsx de "entregas pendientes" (entregadas en Moodle sin corregir).

        `agrupar_por`:
        - "trabajo" (Pendientes por Práctico): hoja por trabajo, columnas
          Comisión·Nombre·Apellido·Email·Fecha; resumen = pendientes por práctico.
        - "comision" (Pendientes por Comisión): hoja por comisión, columnas
          Trabajo·Nombre·Apellido·Email·Fecha; resumen = pendientes por comisión.

        La primera hoja siempre es "Resumen". `pendientes` es una lista plana de
        EntregaPendiente (trabajo, comision, nombre, apellido, email, fecha_entrega).
        """
        por_comision = agrupar_por == "comision"
        if por_comision:
            # Hoja por comisión: el tutor es constante por hoja → va en título y Resumen.
            clave = lambda p: p.comision  # noqa: E731
            etiqueta = "Comisión"
            headers = ["Trabajo", "Nombre", "Apellido", "Email", "Fecha de entrega"]
            fila = lambda p: [p.trabajo, p.nombre, p.apellido, p.email, p.fecha_entrega]  # noqa: E731
            orden = lambda p: (p.trabajo, p.apellido.lower(), p.nombre.lower())  # noqa: E731
            anchos = [24, 22, 22, 30, 20]
            variante = "comision"
        else:
            # Hoja por trabajo: la comisión varía por fila → tutor como columna al lado.
            clave = lambda p: p.trabajo  # noqa: E731
            etiqueta = "Trabajo"
            headers = ["Comisión", "Tutor", "Nombre", "Apellido", "Email", "Fecha de entrega"]
            fila = lambda p: [p.comision, p.tutor, p.nombre, p.apellido, p.email, p.fecha_entrega]  # noqa: E731
            orden = lambda p: (p.comision, p.apellido.lower(), p.nombre.lower())  # noqa: E731
            anchos = [18, 26, 22, 22, 30, 20]
            variante = "practico"

        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        grupos: dict[str, list] = {}
        for p in pendientes:
            grupos.setdefault(clave(p), []).append(p)

        # Tutor por comisión (constante por hoja): lo toma de la primera fila del grupo.
        tutor_de: dict[str, str] = {}
        if por_comision:
            for grupo, filas in grupos.items():
                tutor_de[grupo] = filas[0].tutor if filas else ""

        wb = Workbook()
        resumen = wb.active
        resumen.title = "Resumen"

        # --- Hoja Resumen (primera) ---
        resumen["A1"] = f"Pendientes — {materia_nombre}"
        resumen["A1"].font = Font(size=14, bold=True, color="1F4788")
        resumen["A2"] = f"Documento generado el {fmt_fecha_ar(ahora_ar())}"
        resumen["A2"].font = Font(italic=True, size=10)
        resumen_headers = [etiqueta, "Tutor", "Pendientes"] if por_comision else [etiqueta, "Pendientes"]
        col_total = len(resumen_headers)  # columna donde va el conteo
        for col, h in enumerate(resumen_headers, 1):
            c = resumen.cell(row=3, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        rr = 4
        for grupo in sorted(grupos):
            resumen.cell(row=rr, column=1, value=grupo)
            if por_comision:
                resumen.cell(row=rr, column=2, value=tutor_de.get(grupo, ""))
            resumen.cell(row=rr, column=col_total, value=len(grupos[grupo]))
            rr += 1
        resumen.cell(row=rr, column=1, value="Total").font = Font(bold=True)
        resumen.cell(row=rr, column=col_total, value=len(pendientes)).font = Font(bold=True)
        resumen.column_dimensions["A"].width = 28
        if por_comision:
            resumen.column_dimensions["B"].width = 26
            resumen.column_dimensions["C"].width = 12
        else:
            resumen.column_dimensions["B"].width = 12

        # --- Una hoja por grupo (trabajo o comisión) ---
        usados: set[str] = {"Resumen"}
        merge_end = get_column_letter(len(headers))
        for grupo in sorted(grupos):
            ws = wb.create_sheet(title=self._titulo_unico(grupo, usados))

            ws.merge_cells(f"A1:{merge_end}1")
            t = ws["A1"]
            titulo = f"{etiqueta}: {grupo}"
            if por_comision and tutor_de.get(grupo):
                titulo += f" — Tutor: {tutor_de[grupo]}"
            t.value = titulo
            t.font = Font(size=13, bold=True, color="1F4788")
            t.alignment = Alignment(horizontal="center", vertical="center")

            for col, h in enumerate(headers, 1):
                c = ws.cell(row=2, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border

            row = 3
            for p in sorted(grupos[grupo], key=orden):
                for col, val in enumerate(fila(p), 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = border
                row += 1

            ws.cell(row=row + 1, column=1, value=f"Total pendientes: {len(grupos[grupo])}").font = Font(bold=True)

            for col, width in enumerate(anchos, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            ws.freeze_panes = "A3"

        buffer = io.BytesIO()
        wb.save(buffer)
        data = buffer.getvalue()
        buffer.close()

        fecha = datetime.now().strftime("%Y%m%d")
        filename = self._sanitize_filename(f"pendientes_{variante}_{materia_nombre}_{fecha}.xlsx")
        return data, filename

    def _titulo_unico(self, titulo: str, usados: set[str]) -> str:
        """Nombre de hoja válido y único (Excel no permite hojas con el mismo nombre)."""
        return excel_estilos.titulo_unico(titulo, usados)

    def _get_nota_fill(self, nota: float) -> PatternFill:
        """
        Get background color for nota cell based on value.

        Args:
            nota: Grade value (0-100).

        Returns:
            PatternFill with appropriate color.
        """
        if nota >= 80:
            # Light green
            return PatternFill(
                start_color="D1F2EB", end_color="D1F2EB", fill_type="solid"
            )
        elif nota >= 60:
            # Light yellow
            return PatternFill(
                start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"
            )
        else:
            # Light red
            return PatternFill(
                start_color="FADBD8", end_color="FADBD8", fill_type="solid"
            )

    def _sanitize_filename(self, filename: str) -> str:
        """
        Sanitize filename by removing invalid characters.

        Args:
            filename: Original filename.

        Returns:
            Sanitized filename.
        """
        return excel_estilos.sanitize_filename(filename)
