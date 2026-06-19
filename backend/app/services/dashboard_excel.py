# app/services/dashboard_excel.py
"""
Armado del Excel de avance académico (Dashboard de Gestores).

Hoja 1 "Resumen": tabla por estado (con chips de color) + gráfico de dona.
Hoja 2 "Exámenes": estado de cada parcial/global (aprobado en cualquier instancia /
desaprobado / ausente) con una torta por examen.
Hojas 3-6: desglose de alumnos por estado (Al día / Riesgo medio / Riesgo alto / Sin actividad).

Función pura (recibe datos ya consultados, devuelve bytes). El diseño (paleta, bordes,
banding, barras) se comparte con los Excel de tutores vía excel_estilos. Ref:
PLAN_DASHBOARD_GESTORES.md.
"""

import io

from openpyxl import Workbook
from openpyxl.chart import Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.avance import AvanceAlumno
from app.models.enums import EstadoAvanceEnum
from app.services import excel_estilos as xl
from app.services.notificacion_render import (
    formatear_examenes,
    formatear_faltantes,
    subtitulo_documento,
    titulo_con_corte,
)

_LABEL: dict[EstadoAvanceEnum, str] = {
    EstadoAvanceEnum.AL_DIA: "Al día",
    EstadoAvanceEnum.RIESGO_MEDIO: "Riesgo medio",
    EstadoAvanceEnum.RIESGO_ALTO: "Riesgo alto",
    EstadoAvanceEnum.SIN_ACTIVIDAD: "Sin actividad",
}
_ORDEN: list[EstadoAvanceEnum] = [
    EstadoAvanceEnum.AL_DIA,
    EstadoAvanceEnum.RIESGO_MEDIO,
    EstadoAvanceEnum.RIESGO_ALTO,
    EstadoAvanceEnum.SIN_ACTIVIDAD,
]
# Mismos colores que el gráfico del dashboard web (verde/naranja/rojo/gris).
_COLOR: dict[EstadoAvanceEnum, str] = {
    EstadoAvanceEnum.AL_DIA: "16A34A",
    EstadoAvanceEnum.RIESGO_MEDIO: "D97706",
    EstadoAvanceEnum.RIESGO_ALTO: "DC2626",
    EstadoAvanceEnum.SIN_ACTIVIDAD: "6B7280",
}


# Estado de un examen → etiqueta visible y color, en orden de tabla.
_RESULTADO_FMT: list[tuple[str, str, str]] = [
    ("aprobado", "Aprobado", "16A34A"),         # verde
    ("desaprobado", "Desaprobado", "DC2626"),   # rojo
    ("sin_corregir", "Sin corregir", "9CA3AF"),  # gris claro: entregó, espera corrección
    ("ausente", "Ausente", "6B7280"),           # gris oscuro: no entregó / no rindió
]
# Filas (incluida la barra del bloque y el total) que ocupa cada examen, con aire
# para que la torta de 7,5cm no pise el bloque siguiente.
_ALTO_BLOQUE_EXAMEN = 16


def agregar_examenes(alumnos) -> list[dict]:
    """Cuenta, por examen principal (parcial/global), cuántos alumnos aprobaron /
    desaprobaron / estuvieron ausentes.

    Un alumno cuenta como APROBADO si aprobó el examen en CUALQUIER instancia
    (recuperatorio / extensión / extraordinaria): el rescate ya viene aplicado en el
    campo `resultado` del snapshot, así que acá solo se agrega. Devuelve, en orden de
    aparición, [{etiqueta, aprobado, desaprobado, ausente, total}, ...].
    """
    orden: list[str] = []
    conteos: dict[str, dict[str, int]] = {}
    for a in alumnos:
        examenes = (getattr(a, "resultados_examenes", None) or {}).get("examenes") or []
        for e in examenes:
            etiqueta = e.get("etiqueta")
            if not etiqueta:
                continue
            if etiqueta not in conteos:
                orden.append(etiqueta)
                conteos[etiqueta] = {
                    "aprobado": 0, "desaprobado": 0, "sin_corregir": 0, "ausente": 0,
                }
            res = e.get("resultado")
            if res in conteos[etiqueta]:
                conteos[etiqueta][res] += 1
    filas = []
    for etiqueta in orden:
        c = conteos[etiqueta]
        filas.append({"etiqueta": etiqueta, **c, "total": sum(c.values())})
    return filas


def _hoja_examenes(wb, alumnos, titulo, etiqueta, unidad_actual, corte_examen) -> None:
    """Hoja 'Exámenes': un bloque por parcial/global con su tabla y su torta.

    No se crea si la materia no tiene exámenes configurados (filas vacías).
    """
    filas = agregar_examenes(alumnos)
    if not filas:
        return
    ws = wb.create_sheet("Exámenes")
    xl.banda_titulo(
        ws,
        titulo_con_corte(titulo, etiqueta, unidad_actual, corte_examen),
        8,
        subtitulo=subtitulo_documento(),
    )

    fila = 4
    for ex in filas:
        xl.barra_bloque(ws, fila, ex["etiqueta"], 2)
        header = fila + 1
        xl.celda_header(ws, header, 1, "Resultado")
        xl.celda_header(ws, header, 2, "Alumnos")
        primera = header + 1
        for i, (clave, label, color) in enumerate(_RESULTADO_FMT):
            chip = ws.cell(primera + i, 1, label)
            chip.fill = PatternFill("solid", fgColor=color)
            chip.font = Font(bold=True, color="FFFFFF")
            chip.border = xl.BORDER
            cnt = ws.cell(primera + i, 2, ex[clave])
            cnt.alignment = xl.CENTER
            cnt.border = xl.BORDER
        ultima = primera + len(_RESULTADO_FMT) - 1

        total_row = ultima + 1
        tl = ws.cell(total_row, 1, "Total")
        tv = ws.cell(total_row, 2, ex["total"])
        for c in (tl, tv):
            c.font = Font(bold=True)
            c.border = xl.BORDER
            c.fill = xl.FILL_TITULO
        tv.alignment = xl.CENTER

        xl.grafico_dona(
            ws, f"D{fila}", ex["etiqueta"],
            cat_ref=Reference(ws, min_col=1, min_row=primera, max_row=ultima),
            data_ref=Reference(ws, min_col=2, min_row=header, max_row=ultima),  # con header
            colores=[color for _, _, color in _RESULTADO_FMT],
        )
        fila += _ALTO_BLOQUE_EXAMEN

    _autoancho(ws, [16, 10])
    ws.freeze_panes = "A3"


def _headers(etiqueta: str) -> list[str]:
    return [
        "Apellido", "Nombre", "Email", "Comisión",
        f"{etiqueta} alcanzada", "Le falta", "Exámenes",
    ]


def _autoancho(ws, anchos: list[int]) -> None:
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _hoja_resumen(ws, titulo, conteos, etiqueta, unidad_actual, corte_examen) -> None:
    """Hoja Resumen: banda de título (con corte al lado) + tabla por estado + dona."""
    # ncols=7 para que la banda y el gráfico tengan aire a la derecha de la tabla.
    xl.banda_titulo(
        ws,
        titulo_con_corte(titulo, etiqueta, unidad_actual, corte_examen),
        7,
        subtitulo=subtitulo_documento(),
    )

    # Tabla por estado (header en R4; un chip de color por estado).
    xl.celda_header(ws, 4, 1, "Estado")
    xl.celda_header(ws, 4, 2, "Alumnos")
    primera = 5
    for i, estado in enumerate(_ORDEN):
        chip = ws.cell(primera + i, 1, _LABEL[estado])
        chip.fill = PatternFill("solid", fgColor=_COLOR[estado])
        chip.font = Font(bold=True, color="FFFFFF")
        chip.border = xl.BORDER
        cnt = ws.cell(primera + i, 2, conteos.get(estado, 0))
        cnt.alignment = xl.CENTER
        cnt.border = xl.BORDER
    ultima = primera + len(_ORDEN) - 1

    total_row = ultima + 1
    tlabel = ws.cell(total_row, 1, "Total")
    tval = ws.cell(total_row, 2, sum(conteos.get(e, 0) for e in _ORDEN))
    for c in (tlabel, tval):
        c.font = Font(bold=True)
        c.border = xl.BORDER
        c.fill = xl.FILL_TITULO
    tval.alignment = xl.CENTER

    # Gráfico de DONA con porcentajes + leyenda, coloreado por estado.
    xl.grafico_dona(
        ws, "D4", "Distribución por estado",
        cat_ref=Reference(ws, min_col=1, min_row=primera, max_row=ultima),
        data_ref=Reference(ws, min_col=2, min_row=4, max_row=ultima),  # incluye header
        colores=[_COLOR[e] for e in _ORDEN],
    )

    _autoancho(ws, [16, 10])
    ws.freeze_panes = "A3"


def _hoja_estado(wb, estado, alumnos, etiqueta) -> None:
    """Hoja de desglose de un estado: header azul + filas con borde, banding y resaltado."""
    hoja = wb.create_sheet(_LABEL[estado][:31])
    headers = _headers(etiqueta)
    for col, header in enumerate(headers, 1):
        xl.celda_header(hoja, 1, col, header)

    fila = 2
    for i, a in enumerate(alumnos):
        # "Le falta": las deudas del alumno (igual que los emails).
        le_falta = formatear_faltantes(a.actividades_faltantes, etiqueta)
        xl.fila_datos(
            hoja, fila,
            [a.apellido, a.nombre, a.email, a.comision, a.unidad_alcanzada,
             le_falta, formatear_examenes(a.resultados_examenes)],
            zebra=bool(i % 2),
            center_cols=(4, 5),
            # Si la actividad de la unidad alcanzada quedó DESAPROBADA, se pinta "Le falta".
            resaltar_col=6 if a.actividad_actual_desaprobada else None,
        )
        fila += 1

    if alumnos:
        total = hoja.cell(fila + 1, 1, f"Total: {len(alumnos)} alumnos")
        total.font = Font(bold=True)

    hoja.freeze_panes = "A2"
    _autoancho(hoja, [22, 22, 30, 14, 16, 42, 40])


def construir_excel_avance(
    titulo: str,
    conteos: dict[EstadoAvanceEnum, int],
    alumnos_por_estado: dict[EstadoAvanceEnum, list[AvanceAlumno]],
    etiqueta: str = "Unidad",
    unidad_actual: int | None = None,
    corte_examen: str | None = None,
) -> bytes:
    """Arma el .xlsx de avance (Resumen + 4 hojas de desglose). Devuelve los bytes.

    `etiqueta` ("Unidad"|"Semana") nombra la progresión en los headers y en la columna
    'Le falta' (deudas del alumno, ej. "Semana 8: Quiz, TP"). `unidad_actual` + `corte_examen`
    forman la línea de "Tareas requeridas" del informe (ej. "Unidad 8, Parcial 2").
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    _hoja_resumen(ws, titulo, conteos, etiqueta, unidad_actual, corte_examen)
    todos = [a for lista in alumnos_por_estado.values() for a in lista]
    _hoja_examenes(wb, todos, titulo, etiqueta, unidad_actual, corte_examen)
    for estado in _ORDEN:
        _hoja_estado(wb, estado, alumnos_por_estado.get(estado, []), etiqueta)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
