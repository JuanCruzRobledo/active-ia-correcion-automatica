# app/services/excel_estilos.py
"""
Lenguaje visual compartido de los Excel (reportes de tutores/nexos y de gestores).

Fuente ÚNICA del diseño: paleta, bordes, rellenos y helpers de celda. Así los tres
reportes (avance de tutor nexo/académico y dashboard de gestores) se ven idénticos y
un cambio de estilo se hace en un solo lugar.
"""

from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ===================== paleta =====================

AZUL_TITULO = "1F4E78"   # títulos y encabezados de columna
AZUL_BLOQUE = "2E75B6"   # barra de cada bloque (unidad/comisión/estado)
ROJO_TXT = "C00000"      # texto de "desaprobado"
GRIS_FECHA = "808080"

_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

FILL_HEADER = PatternFill("solid", fgColor=AZUL_TITULO)
FILL_TITULO = PatternFill("solid", fgColor="DCE6F1")    # banda suave del título
FILL_BLOQUE = PatternFill("solid", fgColor=AZUL_BLOQUE)
FILL_ZEBRA = PatternFill("solid", fgColor="F2F6FA")     # fila alterna
FILL_DESAPROB = PatternFill("solid", fgColor="FCE4E4")  # celda con desaprobado

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Estado de avance (clave = EstadoAvanceEnum.value) → etiqueta y color del gráfico.
# Compartido por el Excel de avance (tutores) y el de gestores.
ESTADO_ORDEN = ["AL_DIA", "RIESGO_MEDIO", "RIESGO_ALTO", "SIN_ACTIVIDAD"]
ESTADO_LABEL = {
    "AL_DIA": "Al día",
    "RIESGO_MEDIO": "Riesgo medio",
    "RIESGO_ALTO": "Riesgo alto",
    "SIN_ACTIVIDAD": "Sin actividad",
}
ESTADO_COLOR = {
    "AL_DIA": "16A34A",       # verde
    "RIESGO_MEDIO": "D97706",  # naranja
    "RIESGO_ALTO": "DC2626",   # rojo
    "SIN_ACTIVIDAD": "6B7280",  # gris
}

FONT_TITULO = Font(bold=True, size=16, color=AZUL_TITULO)
FONT_SUBTITULO = Font(italic=True, size=10, color=GRIS_FECHA)
FONT_SECCION = Font(bold=True, size=12, color=AZUL_TITULO)
FONT_BLOQUE = Font(bold=True, size=11, color="FFFFFF")
FONT_HEADER = Font(bold=True, color="FFFFFF")


# ===================== helpers de celda =====================


def celda_header(ws, row: int, col: int, valor) -> None:
    """Celda de cabecera de columna (azul, blanca, negrita, centrada, con borde)."""
    c = ws.cell(row, col, valor)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = CENTER
    c.border = BORDER


def banda_titulo(ws, titulo: str, ncols: int, *, subtitulo: str | None = None) -> None:
    """Título sobre banda azul clara (R1) + subtítulo opcional (R2), abarcando ncols."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = titulo
    ws["A1"].font = FONT_TITULO
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, ncols + 1):
        ws.cell(1, col).fill = FILL_TITULO
    ws.row_dimensions[1].height = 28
    if subtitulo is not None:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws["A2"] = subtitulo
        ws["A2"].font = FONT_SUBTITULO


def barra_bloque(ws, fila: int, texto: str, ncols: int) -> None:
    """Barra de encabezado de un bloque (merge + fill azul + texto blanco)."""
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
    for col in range(1, ncols + 1):
        ws.cell(fila, col).fill = FILL_BLOQUE
    c = ws.cell(fila, 1, texto)
    c.font = FONT_BLOQUE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 20


def grafico_dona(ws, anchor: str, titulo: str, *, cat_ref, data_ref, colores: list[str]) -> None:
    """Gráfico de DONA con porcentajes + leyenda, coloreado punto a punto.

    cat_ref/data_ref: Reference de openpyxl (categorías y datos, este último con header).
    colores: hex (sin #) en el mismo orden que las categorías.
    """
    ch = DoughnutChart()
    ch.title = titulo
    ch.height = 7.5
    ch.width = 11
    ch.holeSize = 55
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cat_ref)
    if ch.series:
        serie = ch.series[0]
        for i, color in enumerate(colores):
            punto = DataPoint(idx=i)
            punto.graphicalProperties.solidFill = color
            serie.data_points.append(punto)
    ch.dataLabels = DataLabelList()
    ch.dataLabels.showPercent = True
    ws.add_chart(ch, anchor)


def tabla_torta_estado(ws, conteos: dict, *, col: int = 4, fila_header: int = 4, anchor: str = "G4") -> None:
    """Tabla 'Estado | Alumnos' con chips de color + torta de dona al lado.

    conteos: {estado_value: n} (claves = EstadoAvanceEnum.value). Se dibujan los 4 estados
    en el orden canónico (los ausentes cuentan 0). La tabla ocupa las columnas (col, col+1)
    desde fila_header; la dona se ancla en `anchor`.
    """
    c0, c1 = col, col + 1
    celda_header(ws, fila_header, c0, "Estado")
    celda_header(ws, fila_header, c1, "Alumnos")
    fila = fila_header + 1
    for estado in ESTADO_ORDEN:
        chip = ws.cell(fila, c0, ESTADO_LABEL[estado])
        chip.fill = PatternFill("solid", fgColor=ESTADO_COLOR[estado])
        chip.font = Font(bold=True, color="FFFFFF")
        chip.border = BORDER
        cnt = ws.cell(fila, c1, conteos.get(estado, 0))
        cnt.alignment = CENTER
        cnt.border = BORDER
        fila += 1
    ultima = fila - 1
    grafico_dona(
        ws, anchor, "Distribución por estado",
        cat_ref=Reference(ws, min_col=c0, min_row=fila_header + 1, max_row=ultima),
        data_ref=Reference(ws, min_col=c1, min_row=fila_header, max_row=ultima),
        colores=[ESTADO_COLOR[e] for e in ESTADO_ORDEN],
    )


def fila_datos(
    ws,
    fila: int,
    valores: list,
    *,
    zebra: bool = False,
    resaltar_col: int | None = None,
    center_cols: tuple[int, ...] = (),
) -> None:
    """Escribe una fila de datos con borde, banding y resaltado opcional.

    - zebra: fondo gris-azulado claro (filas alternas).
    - center_cols: columnas (1-based) centradas; el resto va a la izquierda con wrap.
    - resaltar_col: columna (1-based) a pintar de rojo claro (deuda desaprobada).
    """
    for col, val in enumerate(valores, 1):
        c = ws.cell(fila, col, val)
        c.border = BORDER
        c.alignment = CENTER if col in center_cols else LEFT
        if zebra:
            c.fill = FILL_ZEBRA
    if resaltar_col is not None:
        u = ws.cell(fila, resaltar_col)
        u.fill = FILL_DESAPROB
        u.font = Font(color=ROJO_TXT)
