# app/services/notificacion_render.py
"""
Render de los adjuntos/cuerpos de las notificaciones (PLAN_NOTIFICACIONES_EMAIL.md §5.1).

Funciones PURAS (reciben datos ya resueltos, devuelven str/bytes):
- formatear_faltantes: JSONB de deudas → texto legible.
- construir_excel_avance: Excel (openpyxl) con el formato de análisis por materia,
  agrupado por unidad/semana (tutor nexo) o por comisión (tutor académico).

El HTML del alumno se arma acá también, reutilizando formatear_faltantes.
"""

import io
import re

from app.core.fecha import ahora_ar, fmt_fecha_ar
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services import excel_estilos as xl


# ===================== texto =====================


def _unir_y(items: list[str]) -> str:
    """['4'] → '4'; ['4','5'] → '4 y 5'; ['4','5','6'] → '4, 5 y 6'."""
    items = [i for i in items if i]
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " y " + items[-1]


_LABEL_DEUDA = {
    "TP": "TP",
    "QUIZ": "Quiz",
    "AUTOEVALUACION": "Autoevaluación",
    "CIERRE": "Cierre",
}


def formatear_faltantes(faltantes: dict | None, etiqueta: str = "Unidad") -> str:
    """Convierte el JSONB de deudas (§9.bis F) en texto para el reporte, agrupado por unidad.

    Formato esperado: {"deudas": [{"unidad", "tipo": "TP"|"QUIZ"|"AUTOEVALUACION"|"CIERRE", "estado"}]}.
    `etiqueta` es cómo se llama la progresión en esta materia ("Unidad" por defecto,
    "Semana" para materias por semana como PYE).
    Ej (etiqueta="Unidad"): "Unidad 3: TP (desaprobado), Autoevaluación · Unidad 5: Cierre".
    Ej (etiqueta="Semana"): "Semana 3: Quiz · Semana 5: Cierre". "" si no hay deudas.
    """
    if not faltantes:
        return ""
    deudas = faltantes.get("deudas") or []
    if not deudas:
        return ""
    por_unidad: dict[int, list[str]] = {}
    for d in deudas:
        tipo = _LABEL_DEUDA.get(d.get("tipo"), d.get("tipo") or "?")
        if d.get("estado") == "desaprobado":
            tipo += " (desaprobado)"
        por_unidad.setdefault(d.get("unidad"), []).append(tipo)
    partes = [
        f"{etiqueta} {u}: {', '.join(items)}"
        for u, items in sorted(por_unidad.items(), key=lambda kv: kv[0] if kv[0] is not None else 0)
    ]
    return " · ".join(partes)


_LABEL_RESULTADO = {"aprobado": "Aprobó", "desaprobado": "Desaprobó", "ausente": "Ausente"}


def formatear_examenes(resultados: dict | None) -> str:
    """Resultados de exámenes (JSONB) → texto para el reporte.

    Formato esperado: {"examenes": [{"etiqueta", "resultado", "rescatado"}, ...]}.
    Ej: "Parcial 1: Aprobó · Parcial 2: Aprobó (recuperado) · Global 1: Ausente".
    "" si no hay exámenes. `rescatado` marca un parcial reprobado/ausente que su
    recuperatorio/extensión rescató.
    """
    if not resultados:
        return ""
    examenes = resultados.get("examenes") or []
    if not examenes:
        return ""
    partes = []
    for e in examenes:
        label = _LABEL_RESULTADO.get(e.get("resultado"), e.get("resultado") or "?")
        if e.get("rescatado"):
            label += " (recuperado)"
        partes.append(f"{e.get('etiqueta')}: {label}")
    return " · ".join(partes)


def tiene_examen_desaprobado(resultados: dict | None) -> bool:
    """True si algún examen quedó DESAPROBADO (con rescate ya aplicado).

    Dispara la inclusión del alumno en los reportes aunque esté al día. NO usa "ausente":
    antes de que se tome el examen todos figuran ausentes y dispararía a toda la materia.
    """
    if not resultados:
        return False
    return any(
        e.get("resultado") == "desaprobado" for e in (resultados.get("examenes") or [])
    )


def texto_tareas_requeridas(etiqueta: str, unidad_actual: int | None, corte_examen: str | None) -> str:
    """'Tareas requeridas: Unidad 8, Parcial 2' (o '' si no hay corte)."""
    partes: list[str] = []
    if unidad_actual is not None:
        partes.append(f"{etiqueta} {unidad_actual}")
    if corte_examen:
        partes.append(corte_examen)
    return f"Tareas requeridas: {', '.join(partes)}" if partes else ""


def titulo_con_corte(
    nombre: str, etiqueta: str = "Unidad", unidad_actual: int | None = None,
    corte_examen: str | None = None,
) -> str:
    """Título de la hoja con el corte al lado, mismo peso: 'Materia — Tareas requeridas: ...'."""
    tr = texto_tareas_requeridas(etiqueta, unidad_actual, corte_examen)
    return f"{nombre}  —  {tr}" if tr else nombre


def subtitulo_documento() -> str:
    """Metadata chica bajo el título: 'Documento generado el {fecha}'."""
    return f"Documento generado el {fmt_fecha_ar(ahora_ar())}"


def examenes_pendientes(resultados: dict | None) -> list[str]:
    """Etiquetas de los exámenes que el alumno DEBE atender (desaprobado o ausente).

    El rescatado ya quedó 'aprobado' (no entra). Ej: ['Parcial 1', 'Parcial 2'].
    """
    if not resultados:
        return []
    return [
        e.get("etiqueta")
        for e in (resultados.get("examenes") or [])
        if e.get("resultado") in ("desaprobado", "ausente") and e.get("etiqueta")
    ]


def examen_para_atender(resultados: dict | None) -> bool:
    """True si algún examen quedó DESAPROBADO o AUSENTE (con rescate ya aplicado).

    Lo usa la matriz de exámenes del PDF del tutor: muestra a quienes tienen algo para
    atender (reprobó o no rindió). El rescatado queda 'aprobado' → no entra.
    """
    if not resultados:
        return False
    return any(
        e.get("resultado") in ("desaprobado", "ausente")
        for e in (resultados.get("examenes") or [])
    )


def label_resultado_examen(e: dict | None) -> str:
    """Resultado de UN examen → 'Aprobó' / 'Desaprobó' / 'Ausente' (+ ' (R)' si rescatado)."""
    if not e:
        return "—"
    label = _LABEL_RESULTADO.get(e.get("resultado"), e.get("resultado") or "?")
    if e.get("rescatado"):
        label += " (R)"
    return label


# ===================== helpers de texto (escape HTML) =====================


def _escape(text: str) -> str:
    """Escapa &, <, > para no inyectar markup en el HTML del email."""
    if not text:
        return ""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# ===================== helpers de Excel (compartidos) =====================

_INVALIDOS_HOJA = re.compile(r"[:\\/?*\[\]]")


def _safe_sheet_name(nombre: str, usados: set[str]) -> str:
    """Nombre de hoja válido para Excel: sin caracteres prohibidos, <=31, único."""
    base = _INVALIDOS_HOJA.sub(" ", (nombre or "Hoja")).strip()[:31] or "Hoja"
    nombre_final = base
    i = 2
    while nombre_final in usados:
        sufijo = f" ({i})"
        nombre_final = base[: 31 - len(sufijo)] + sufijo
        i += 1
    usados.add(nombre_final)
    return nombre_final


# ===================== Excel de avance (formato análisis por regional/materia) =====================
# El diseño (paleta, bordes, banding, barras) vive en excel_estilos (xl), compartido con
# el Excel del Dashboard de Gestores para que los tres reportes se vean idénticos.


def _nombre_alumno(a: dict) -> str:
    """'APELLIDO, NOMBRE' (como en el análisis de referencia)."""
    return f"{a.get('apellido', '')}, {a.get('nombre', '')}".strip(", ")


def _tipos_de_deudas(deudas: list[dict]) -> str:
    """Tipos de un conjunto de deudas (de UNA unidad) → 'TP, Autoevaluación, Quiz (desaprobado)'."""
    partes = []
    for d in deudas:
        tipo = _LABEL_DEUDA.get(d.get("tipo"), d.get("tipo") or "?")
        if d.get("estado") == "desaprobado":
            tipo += " (desaprobado)"
        partes.append(tipo)
    return ", ".join(partes)


def _agrupar_avance(alumnos: list[dict], por_unidad: bool) -> list[tuple]:
    """Agrupa los alumnos con deudas. Devuelve [(clave_raw, [alumnos])] ordenado.

    - por_unidad: clave = número de unidad; un alumno aparece en CADA unidad donde debe.
    - por comisión: clave = comisión; un alumno aparece una vez (con todas sus deudas).
    """
    grupos: dict = {}
    if por_unidad:
        for a in alumnos:
            for u in sorted({d.get("unidad") for d in a.get("deudas", []) if d.get("unidad") is not None}):
                grupos.setdefault(u, []).append(a)
    else:
        for a in alumnos:
            grupos.setdefault(a.get("comision") or "—", []).append(a)
    return sorted(grupos.items(), key=lambda kv: kv[0])


def _hoja_avance(
    wb: Workbook, mat: dict, usados: set[str], por_unidad: bool, *, prefijo: str = ""
) -> None:
    """Arma la hoja de UNA materia con el formato de referencia (resumen + listado).

    `prefijo`: se antepone al nombre de la PESTAÑA (no al título) para distinguir las dos
    hojas de una misma materia cuando se generan ambas vistas (académico).
    """
    etiqueta = mat.get("etiqueta") or "Unidad"
    nombre_mat = mat.get("materia") or "Materia"
    ws = wb.create_sheet(_safe_sheet_name(f"{prefijo}{nombre_mat}", usados))

    todos = mat.get("alumnos", [])
    # Al listado van los que deben actividades O tienen parciales para atender.
    deudores = [a for a in todos if a.get("deudas") or a.get("examenes")]
    con_examen = [a for a in todos if a.get("examenes")]
    # Vista por unidad: agrupar por unidad a los que deben actividades (los de solo-examen
    # van al bloque "Exámenes"). Vista por comisión: agrupar TODOS los deudores por comisión.
    if por_unidad:
        grupos = _agrupar_avance([a for a in deudores if a.get("deudas")], True)
    else:
        grupos = _agrupar_avance(deudores, False)

    def _plural(n: int) -> str:
        return f"{n} alumno" if n == 1 else f"{n} alumnos"

    def _bloque(fila: int, encabezado: str, headers: list[str]) -> int:
        """Barra de bloque + headers de columna. Devuelve la fila de la 1ª fila de datos."""
        xl.barra_bloque(ws, fila, encabezado, len(headers))
        for col, h in enumerate(headers, 1):
            xl.celda_header(ws, fila + 1, col, h)
        return fila + 2

    def _orden_comision(x):
        return (x.get("comision") or "", x.get("apellido") or "", x.get("nombre") or "")

    # --- Título con el corte al lado (R1) + fecha (R2) ---
    xl.banda_titulo(
        ws,
        titulo_con_corte(nombre_mat, etiqueta, mat.get("unidad_actual"), mat.get("corte_examen")),
        7,
        subtitulo=subtitulo_documento(),
    )

    # --- Resumen de deudas (A-B): por unidad/comisión (+ fila Exámenes en vista por unidad) ---
    fila = 4
    xl.celda_header(ws, fila, 1, etiqueta if por_unidad else "Comisión")
    xl.celda_header(ws, fila, 2, "Alumnos que deben")
    fila += 1
    for i, (clave, miembros) in enumerate(grupos):
        clave_txt = f"{etiqueta} {clave}" if por_unidad else str(clave)
        xl.fila_datos(ws, fila, [clave_txt, len(miembros)], zebra=bool(i % 2), center_cols=(2,))
        fila += 1
    if por_unidad and con_examen:
        xl.fila_datos(ws, fila, ["Exámenes", len(con_examen)], zebra=bool(len(grupos) % 2), center_cols=(2,))
        fila += 1

    # --- Tabla + torta por estado (D-E / G4), con TODOS los alumnos (incluye al día) ---
    conteo_estado: dict = {}
    for a in todos:
        est = a.get("estado")
        if est:
            conteo_estado[est] = conteo_estado.get(est, 0) + 1
    xl.tabla_torta_estado(ws, conteo_estado)

    # --- Listado detallado (debajo de las tablas y la torta) ---
    fila = max(fila, 10) + 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    titulo = (
        f"LISTADO DE ALUMNOS POR {etiqueta.upper()} (por comisión y alfabético)"
        if por_unidad
        else "LISTADO DE ALUMNOS POR COMISIÓN (alfabético)"
    )
    ws.cell(fila, 1, titulo).font = xl.FONT_SECCION
    fila += 2

    if por_unidad:
        for clave, miembros in grupos:
            fila = _bloque(fila, f"{etiqueta} {clave}  ({_plural(len(miembros))})",
                           ["Comisión", "Alumno", "Pendiente"])
            for i, a in enumerate(sorted(miembros, key=_orden_comision)):
                deudas_u = [d for d in a.get("deudas", []) if d.get("unidad") == clave]
                desaprob = any(d.get("estado") == "desaprobado" for d in deudas_u)
                xl.fila_datos(
                    ws, fila, [a.get("comision") or "—", _nombre_alumno(a), _tipos_de_deudas(deudas_u)],
                    zebra=bool(i % 2), center_cols=(1,), resaltar_col=3 if desaprob else None,
                )
                fila += 1
            fila += 1
        # Parciales "como otra unidad": bloque Exámenes con los alumnos que deben parciales.
        if con_examen:
            fila = _bloque(fila, f"Exámenes  ({_plural(len(con_examen))})",
                           ["Comisión", "Alumno", "Parciales"])
            for i, a in enumerate(sorted(con_examen, key=_orden_comision)):
                xl.fila_datos(
                    ws, fila, [a.get("comision") or "—", _nombre_alumno(a), ", ".join(a.get("examenes", []))],
                    zebra=bool(i % 2), center_cols=(1,),
                )
                fila += 1
            fila += 1
        anchos = (20, 40, 46, 16, 10)
    else:
        for clave, miembros in grupos:
            fila = _bloque(fila, f"Comisión {clave}  ({_plural(len(miembros))})",
                           ["Alumno", "Pendiente", "Exámenes"])
            for i, a in enumerate(sorted(miembros, key=lambda x: (x.get("apellido") or "", x.get("nombre") or ""))):
                pendiente = formatear_faltantes({"deudas": a.get("deudas", [])}, etiqueta)
                desaprob = any(d.get("estado") == "desaprobado" for d in a.get("deudas", []))
                xl.fila_datos(
                    ws, fila, [_nombre_alumno(a), pendiente, ", ".join(a.get("examenes", []))],
                    zebra=bool(i % 2), resaltar_col=2 if desaprob else None,
                )
                fila += 1
            fila += 1
        anchos = (40, 50, 30, 16, 10)

    ws.freeze_panes = "A3"
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


# Pasos (por_unidad) y prefijo de pestaña según el modo de agrupación.
_PREFIJO_UNIDAD = "Por unidad — "
_PREFIJO_COMISION = "Por comisión — "
_PASOS_AGRUPAR: dict[str, list[bool]] = {
    "unidad": [True],
    "comision": [False],
    "ambos": [True, False],  # 2 hojas por materia (académico): unidad + comisión
}


def construir_excel_avance(materias: list[dict], *, agrupar: str = "unidad") -> bytes:
    """Excel de avance con el formato del análisis por materia.

    materias: [{"materia": str, "etiqueta": str, "alumnos": [
        {"comision": str, "apellido": str, "nombre": str,
         "deudas": [{"unidad": int, "tipo": str, "estado": str}, ...]}]}]
    agrupar:
      - "unidad" (tutor nexo): 1 hoja/materia por unidad/semana; columnas Comisión|Alumno|Pendiente.
        El Pendiente de cada bloque son las deudas de ESA unidad.
      - "comision" (vista por comisión): 1 hoja/materia por comisión; columnas Alumno|Pendiente.
        El Pendiente trae todas las unidades del alumno (formatear_faltantes).
      - "ambos" (tutor académico): 2 hojas/materia — primero por unidad (formato nexo) y luego
        por comisión —, con la pestaña prefijada para distinguirlas.
    Solo se listan alumnos con ≥1 deuda. Respeta la etiqueta (Unidad/Semana) de la materia.
    """
    pasos = _PASOS_AGRUPAR.get(agrupar, [True])
    dos_vistas = len(pasos) > 1
    wb = Workbook()
    wb.remove(wb.active)
    usados: set[str] = set()
    if not materias:
        ws = wb.create_sheet(_safe_sheet_name("Sin datos", usados))
        ws["A1"] = "Sin alumnos con pendientes"
    for mat in materias:
        for por_unidad in pasos:
            # Solo se prefija la pestaña cuando hay DOS vistas de la misma materia.
            prefijo = ""
            if dos_vistas:
                prefijo = _PREFIJO_UNIDAD if por_unidad else _PREFIJO_COMISION
            _hoja_avance(wb, mat, usados, por_unidad, prefijo=prefijo)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ===================== HTML (cuerpo del email del alumno) =====================

# Estilos INLINE: los clientes de correo ignoran <style>/CSS externo.
_TH = (
    "background:#1F4E78;color:#ffffff;text-align:left;"
    "padding:8px 10px;border:1px solid #cccccc;font-size:13px"
)
_TD = "padding:8px 10px;border:1px solid #cccccc;font-size:13px"
_WRAP = "font-family:Arial,Helvetica,sans-serif;font-size:14px;color:#333333"


def construir_html_alumno(nombre_alumno: str, filas: list[dict]) -> str:
    """Tabla HTML inline (Comisión / Materia / Actividad) para el cuerpo del email del alumno.

    filas: [{"comision", "materia", "actividad"}] — 'actividad' ya formateada (texto).
    Sin filas → mensaje "estás al día" (no debería ocurrir: el caller filtra, T6).
    """
    nombre = _escape(nombre_alumno or "")
    if not filas:
        return (
            f'<div style="{_WRAP}">'
            f"<p>Hola {nombre},</p>"
            "<p>¡Estás al día! No registramos actividades pendientes. 🎉</p>"
            "</div>"
        )

    filas_html = "".join(
        "<tr>"
        f'<td style="{_TD}">{_escape(f.get("comision") or "—")}</td>'
        f'<td style="{_TD}">{_escape(f.get("materia") or "—")}</td>'
        f'<td style="{_TD}">{_escape(f.get("corte") or "—")}</td>'
        f'<td style="{_TD}">{_escape(f.get("actividad") or "—")}</td>'
        f'<td style="{_TD}">{_escape(f.get("examenes") or "—")}</td>'
        "</tr>"
        for f in filas
    )
    generado = fmt_fecha_ar(ahora_ar())
    return (
        f'<div style="{_WRAP}">'
        f"<p>Hola {nombre},</p>"
        "<p>Este es tu resumen de actividades pendientes y de tus exámenes. La columna "
        "<strong>Hasta</strong> indica hasta qué unidad/semana se evalúa cada materia. Te "
        "recomendamos completar lo pendiente para no atrasarte:</p>"
        '<table style="border-collapse:collapse;width:100%;max-width:760px">'
        "<thead><tr>"
        f'<th style="{_TH}">Comisión</th>'
        f'<th style="{_TH}">Materia</th>'
        f'<th style="{_TH}">Hasta</th>'
        f'<th style="{_TH}">Actividad</th>'
        f'<th style="{_TH}">Exámenes</th>'
        "</tr></thead>"
        f"<tbody>{filas_html}</tbody>"
        "</table>"
        f'<p style="color:#999999;font-size:11px;margin-top:10px">Documento generado el {generado}</p>'
        '<p style="color:#666666;font-size:12px;margin-top:14px">'
        "Ante cualquier duda, escribile a tu tutor. ¡Éxitos! 💪</p>"
        "</div>"
    )
